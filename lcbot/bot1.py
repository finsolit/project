import telebot
import time
import threading
import datetime
import copy
import string
from datetime import datetime
import schedule
from telebot import types
import _pickle as pickle
import os
import _thread
import urllib
import random
import sqlite3
import openpyxl
global inadmin, password, admin_id, oprcall, oprs, add_opr, work_with_opr, add_vopros_admin, filter,txtrassilki, new_group, ngroup, change_group,add_poi, vvod_koda, what_change
class group:
    id=0
    name=''
    tema=''
    time=0
what_change=''	
def chek_pass():
    conn = sqlite3.connect('CONV.sqlite')
    cursor = conn.cursor()
    cursor.execute("SELECT * FROM PASS")
    results = cursor.fetchall()
    conn.close()	
    print(results[0][0])
    return(results[0][0])	
	
vvod_koda=[]
change_group=''
add_poi=[]
add_vopros_admin=0
oprcall=1
add_opr=0
new_group=0
inadmin=[]
filter=[]
txtrassilki=''
password=chek_pass()
input = open('admin.pkl', 'rb')
admin_id = pickle.load(input)
input.close()
input = open('oprs.pkl', 'rb')
oprs = pickle.load(input)
input.close()
print(admin_id)
wb = openpyxl.load_workbook(filename = 'leng.xlsx')
sheet = wb['test']
val = sheet.cell(row=1, column=1)
print(val.value)
def  lengstr(leng,strn):
    bb=sheet.cell(row=strn, column=leng).value
    return str(bb)
TOKEN = '507631866:AAHIe_Lc8b2llPPEinLXACFvu0rN2ncZzCE'
bot = telebot.TeleBot(TOKEN)

# Handle '/start' and '/help'
@bot.message_handler(commands=['help', 'start'])
def start(message):
    gg=start_prov(message.chat.id)
    if gg==0:
        keyboard = types.InlineKeyboardMarkup(row_width=2)
        keyboard.add(types.InlineKeyboardButton(text=lengstr(1,3),callback_data='%'),
                     types.InlineKeyboardButton(text=lengstr(1,4),callback_data='@'))
        msg = bot.send_message(message.chat.id, lengstr(1,2),reply_markup=keyboard)
    elif gg==1:
        keyboard = types.ReplyKeyboardMarkup(resize_keyboard=True)
        keyboard.add(*[types.InlineKeyboardButton(text=name,callback_data=name) for name in [lengstr(1,6)]])
        msg = bot.send_message(message.chat.id, lengstr(1,5),reply_markup=keyboard) 
    else:        
        keyboard = types.ReplyKeyboardMarkup(resize_keyboard=True)
        keyboard.add(*[types.InlineKeyboardButton(text=name,callback_data=name) for name in [lengstr(1,8),lengstr(1,9)]])
        msg = bot.send_message(message.chat.id, lengstr(1,7),reply_markup=keyboard) 
	
	
@bot.message_handler(commands=['admin'])	
def admin(m):
    inadmin.append(m.chat.id)
    keyboard = types.InlineKeyboardMarkup(row_width=2)
    keyboard.add(types.InlineKeyboardButton(text=lengstr(1,11),callback_data='!'))
    msg = bot.send_message(m.chat.id, lengstr(1,10),reply_markup=keyboard)    
    return()	
	
@bot.message_handler(commands=['chat_id'])	
def chat_idr(m):
    global new_group, ngroup, admin_id
    if new_group==1 and m.from_user.id==admin_id:
        try:
            ngroup.id=m.chat.id  
            ngroup.name=m.chat.title
            new_group=2
            msg = bot.send_message(admin_id, lengstr(1,47))
        except Exception:
            msg = bot.send_message(admin_id, lengstr(1,48)) 
    return()	

	
@bot.message_handler(content_types=["text"])
def repeat_all_messages(message): 
    name(message)	
	
	
	
@bot.callback_query_handler(func=lambda c:True)
def inline(c):
    global oprs, add_opr, work_with_opr, add_vopros_admin, filter,txtrassilki, new_group, ngroup, change_group,add_poi, what_change
    bib=c.message.message_id
    if 'otv' in c.data:
        print(c.data)
        for i in range(0,len(c.data)):
            if c.data[i]==':':
                lst=i+1
                otvet_id=c.data[3:i]
            if c.data[i]=='†':
                opr_id=c.data[lst:i]
                vopros_id=c.data[i+1:]
        add_otvet_in(c.message.chat.id,opr_id,vopros_id,otvet_id)
        print(c.message.chat.id,opr_id,vopros_id,otvet_id)
        chek_next_question(c.message.chat.id,bib,opr_id,vopros_id)
        return
    if c.data=='!':
        inadmin.remove(c.message.chat.id)
        msg = bot.send_message(c.message.chat.id, lengstr(1,14))   
        return		
    if c.data=='%':
        add_urist(c.message.chat.id)
        keyboard = types.ReplyKeyboardMarkup(resize_keyboard=True)
        keyboard.add(*[types.InlineKeyboardButton(text=name,callback_data=name) for name in [lengstr(1,6)]])
        msg = bot.send_message(c.message.chat.id, lengstr(1,5),reply_markup=keyboard) 
        return		
    if c.data=='@':
        add_client(c.message.chat.id)
        keyboard = types.ReplyKeyboardMarkup(resize_keyboard=True)
        keyboard.add(*[types.InlineKeyboardButton(text=name,callback_data=name) for name in [lengstr(1,8),lengstr(1,9)]])
        msg = bot.send_message(c.message.chat.id, lengstr(1,7),reply_markup=keyboard) 
        return
    if c.data=='addopr':
        oprs+=1
        output = open('oprs.pkl', 'wb')
        pickle.dump(oprs, output, 2)
        output.close()	
        add_opr=1
        msg = bot.send_message(c.message.chat.id, lengstr(1,20))  
        return
    if c.data=='extxt':
        msg = bot.send_message(c.message.chat.id, lengstr(1,41))   
        txtrassilki=''		
        return
######################## Vvod teksta rassilki
    if c.data=='txtras':
        keyboard = types.InlineKeyboardMarkup(row_width=2)
        keyboard.add(*[types.InlineKeyboardButton(text=name,callback_data='extxt') for name in [lengstr(1,40)]])
        msg = bot.send_message(c.message.chat.id, lengstr(1,39),reply_markup=keyboard)   
        txtrassilki=1		
        return
################################################
    if 'ddelopr' in c.data:
            opr_id=int(c.data[:-7])	 
            delete_opros(opr_id)
            dsa=all_oprosi()
            keyboard = types.InlineKeyboardMarkup(row_width=1)
            for i in range(0,len(dsa)):
                keyboard.add(*[types.InlineKeyboardButton(text=name,callback_data='cop'+str(dsa[i][0])) for name in [dsa[i][1]]]) 
            keyboard.add(*[types.InlineKeyboardButton(text=name,callback_data='addopr') for name in [lengstr(1,19)]])                 
            msg = bot.send_message(c.message.chat.id, lengstr(1,18)+' '+str(len(dsa)),reply_markup=keyboard) 
            return()			
    if 'delopr' in c.data:
        opr_id=int(c.data[:-6])	
        keyboard = types.InlineKeyboardMarkup(row_width=2)
        keyboard.add(types.InlineKeyboardButton(text='Нет',callback_data='cop'+str(opr_id)),
                    types.InlineKeyboardButton(text='Да',callback_data=str(opr_id)+'ddelopr'))
        msg = bot.send_message(c.message.chat.id, lengstr(1,24),reply_markup=keyboard)
        return
    if 'addvopr' in c.data:
        opr_id=int(c.data[:-7])	
        add_vopros_admin=1
        work_with_opr=opr_id
        keyboard = types.InlineKeyboardMarkup(row_width=1)
        keyboard.add(types.InlineKeyboardButton(text='Отмена',callback_data='cop'+str(opr_id)))
        msg = bot.send_message(c.message.chat.id, lengstr(1,25),reply_markup=keyboard)
        return
    if 'razoslat' in c.data:
        opr_id=int(c.data[:-8])	
        keyboard = types.InlineKeyboardMarkup(row_width=2)
        keyboard.add(types.InlineKeyboardButton(text='Нет',callback_data='cop'+str(opr_id)),
                    types.InlineKeyboardButton(text='Да',callback_data=str(opr_id)+'razuslat'))
        msg = bot.send_message(c.message.chat.id, lengstr(1,32),reply_markup=keyboard)
        return    
    if 'razuslat' in c.data:
        opr_id=int(c.data[:-8])	
        rassilka_oprosa(opr_id)	
        msg = bot.send_message(c.message.chat.id, lengstr(1,29))
    if 'delvopr' in c.data:
        opr_id=int(c.data[:-7])
        zz=''		
        spisok=spisok_voprosov(opr_id)
        keyboard = types.InlineKeyboardMarkup(row_width=1)
        for i in range(len(spisok)):
            zz+=str(i+1)+') '+spisok[i][0]+'\n'
            keyboard.add(types.InlineKeyboardButton(text=str(i+1),callback_data='dzx'+str(i)+':'+str(opr_id)))             
        keyboard.add(types.InlineKeyboardButton(text='Отмена',callback_data='cop'+str(opr_id)))
        msg = bot.send_message(c.message.chat.id, lengstr(1,27)+'\n\n'+zz,reply_markup=keyboard) 
        return		
    if 'dzx' in c.data:
        for i in range(0,len(c.data)):
            if c.data[i]==':':
                num=c.data[3:i]
                opr_id=c.data[i+1:]
        delete_vopros(num,opr_id)
        msg = bot.send_message(c.message.chat.id, lengstr(1,28))	
        return		
    if 'cop' in c.data:
        opr_id=int(c.data[3:])
        zz='Error'
        add_vopros_admin=0
        zz=chek_opros_full(opr_id)
        keyboard = types.InlineKeyboardMarkup(row_width=2)
        keyboard.add(types.InlineKeyboardButton(text='Удалить вопрос',callback_data=str(opr_id)+'delvopr'),
                    types.InlineKeyboardButton(text='Добавить вопрос',callback_data=str(opr_id)+'addvopr'),
                    types.InlineKeyboardButton(text='Удалить опрос',callback_data=str(opr_id)+'delopr'),
                    types.InlineKeyboardButton(text='Разослать опрос',callback_data=str(opr_id)+'razoslat'))
        msg = bot.send_message(c.message.chat.id, zz,reply_markup=keyboard) 
        return		
    if 'stt' in c.data:
        opr_id=int(c.data[3:])
        zz='Error'
        zz=stat_opros(opr_id)
        msg = bot.send_message(c.message.chat.id, zz) 
        return	
############################### sozdanie filtra
    if 'ras' in c.data:
        opr_id=int(c.data[3:])
        work_with_opr=opr_id
        kr=chek_op(opr_id)
        if kr==1:
           zz=lengstr(1,37)
        else:
           zz=lengstr(1,36)
           msg = bot.send_message(c.message.chat.id, zz) 
           return	
        create_filter(opr_id)
        keyboard=keyboard_for_filter(opr_id)
                        
        msg = bot.send_message(c.message.chat.id, zz,reply_markup=keyboard) 
        return	
    if 'ifm' in c.data:
        for i in range(0,len(c.data)):
            if c.data[i]==':':
                su=int(c.data[3:i])
                tt=i+1
            if c.data[i]=='†':
                zu=int(c.data[tt:i])
                opr_id=int(c.data[i+1:])         	
        if filter[su][zu]==1:
           filter[su][zu]=0
        else:
           filter[su][zu]=1	
        print(filter)
        keyboard=keyboard_for_filter(opr_id)
        msg = bot.edit_message_reply_markup(c.message.chat.id, bib,reply_markup=keyboard) 
        return	
    if c.data=='addgroup':
        new_group=1  
        ngroup=group()
        msg=  bot.send_message(c.message.chat.id, lengstr(1,45))      
        return	
    if 'grp' in c.data:
        zcf=chek_group(int(c.data[3:]))	
        change_group=''
        keyboard = types.InlineKeyboardMarkup(row_width=2)	
        keyboard.add(types.InlineKeyboardButton(text='Изменить задержку',callback_data='gzadr'+str(zcf[0])),
                    types.InlineKeyboardButton(text='Изменить тему',callback_data='gtema'+str(zcf[0])),
                    types.InlineKeyboardButton(text='Удалить группу',callback_data='gdelt'+str(zcf[0])))	
        msg = bot.send_message(c.message.chat.id, 'Id: '+str(zcf[0])+'\nНазвание: '+zcf[1]+'\nТема: '+zcf[2]+'\nЗадержка: '+str(zcf[3]),reply_markup=keyboard)     
    if 'gzadr' in c.data:
        change_group='z'+c.data[5:]
        keyboard = types.InlineKeyboardMarkup(row_width=1)	
        keyboard.add(types.InlineKeyboardButton(text=lengstr(1,40),callback_data='grp'+c.data[5:])) 
        msg=  bot.send_message(c.message.chat.id, lengstr(1,50),reply_markup=keyboard)    
    if 'gtema' in c.data:
        change_group='t'+c.data[5:]
        keyboard = types.InlineKeyboardMarkup(row_width=1)	
        keyboard.add(types.InlineKeyboardButton(text=lengstr(1,40),callback_data='grp'+c.data[5:])) 
        msg=  bot.send_message(c.message.chat.id, lengstr(1,52),reply_markup=keyboard)       
    if 'gdelt' in c.data:
        keyboard = types.InlineKeyboardMarkup(row_width=1)	
        keyboard.add(types.InlineKeyboardButton(text='Нет',callback_data='grp'+c.data[5:]),types.InlineKeyboardButton(text='Да',callback_data='ddt'+c.data[5:])) 
        msg=  bot.send_message(c.message.chat.id, lengstr(1,53),reply_markup=keyboard)
    if 'ddt' in c.data:
        delete_group(int(c.data[3:]))
        msg=  bot.send_message(c.message.chat.id, lengstr(1,54)) 
    if 'poisk' in c.data:
        tema=c.data[5:]
        dobavit_poi(c.message.chat.id,tema)	
        add_poi.append(c.message.chat.id)
        msg=  bot.send_message(c.message.chat.id, lengstr(1,57)) 	
    if 'zaprs' in c.data:
        zp_id=int(c.data[5:])
        zaprosi=zaprosi_cl(c.message.chat.id)
        keyboard = types.InlineKeyboardMarkup(row_width=1)
        keyboard.add(types.InlineKeyboardButton(text=lengstr(1,60),callback_data='zpdel'+str(zp_id)))		
        msg=  bot.send_message(c.message.chat.id, lengstr(1,59)+'\n'+zaprosi[zp_id][3],reply_markup=keyboard) 	
    if 'zpdel' in c.data:
        zp_id=int(c.data[5:])
        zaprosi=zaprosi_delete(c.message.chat.id,zp_id)	
        msg=  bot.send_message(c.message.chat.id, lengstr(1,61)) 	
    if 'rznxt' in c.data:
        zp_id=int(c.data[5:])
        razgovor_go(c.message.chat.id,zp_id)
        keyboard = types.InlineKeyboardMarkup(row_width=1)
        msg=  bot.edit_message_reply_markup(c.message.chat.id,bib,keyboard) 
    if 'rzend' in c.data:
        zp_id=int(c.data[5:])
        keyboard = types.InlineKeyboardMarkup(row_width=1)
        msg=  bot.edit_message_reply_markup(c.message.chat.id,bib,keyboard) 
        msg= bot.send_message(c.message.chat.id,lengstr(1,68)) 	
        msg= bot.send_message(zp_id,lengstr(1,68))		
    if c.data=='opciipass':
        what_change='p'
        msg= bot.send_message(c.message.chat.id,lengstr(1,72))    
    if c.data=='opciitext':
        what_change='t'
        msg = bot.send_document(c.message.chat.id,open('leng.xlsx', 'rb') ,caption=lengstr(1,74))		
    return()		
		
		
		
		
		
		
		
		
		
		
		
		
		
		
def name(m):
    global admin_id, add_opr, oprs, work_with_opr, add_vopros_admin, filter,txtrassilki, new_group, ngroup, change_group,add_poi, vvod_koda, what_change, password
    razgovori=from_us()
    if m.chat.id in razgovori:
        cl_id=chek_razgovor(m.chat.id)
        keyboard = types.InlineKeyboardMarkup(row_width=1)
        keyboard.add(types.InlineKeyboardButton(text=lengstr(1,66),callback_data='rznxt'+str(m.chat.id)))
        keyboard.add(types.InlineKeyboardButton(text=lengstr(1,67),callback_data='rzend'+str(m.chat.id)))
        msg = bot.send_message(cl_id, lengstr(1,65)+str(m.chat.id)+'\n\n'+m.text,reply_markup=keyboard)
    if m.chat.id in inadmin:
        if m.text==password:
            admin_id=m.chat.id
            output = open('admin.pkl', 'wb')
            pickle.dump(admin_id, output, 2)
            output.close()
            inadmin.remove(m.chat.id)
            keyboard = types.ReplyKeyboardMarkup(resize_keyboard=True)
            keyboard.add(*[types.InlineKeyboardButton(text=name,callback_data=name) for name in [lengstr(1,15),lengstr(1,16),lengstr(1,17),lengstr(1,33),lengstr(1,42)]])
            msg = bot.send_message(m.chat.id, lengstr(1,12),reply_markup=keyboard)  
        else:
            keyboard = types.InlineKeyboardMarkup(row_width=2)
            keyboard.add(types.InlineKeyboardButton(text=lengstr(1,11),callback_data='!'))
            msg = bot.send_message(m.chat.id, lengstr(1,13),reply_markup=keyboard) 
    if m.chat.id in vvod_koda:
        vvod_koda.remove(m.chat.id)
        kod=m.text
        id_client=chek_kod(kod)
        if id_client==0:
            msg = bot.send_message(m.chat.id, lengstr(1,63))
        else:
            msg = bot.send_message(m.chat.id, lengstr(1,64))
            razgovor_go(m.chat.id,id_client)			
    if m.text==lengstr(1,8):
            temi=groups_tems()
            keyboard = types.InlineKeyboardMarkup(row_width=1)
            for i in range(0,len(temi)):
                keyboard.add(*[types.InlineKeyboardButton(text=str(name),callback_data='poisk'+str(name)) for name in [temi[i][0]]])                 
            msg = bot.send_message(m.chat.id, lengstr(1,55),reply_markup=keyboard) 
    if m.text==lengstr(1,6):       
            vvod_koda.append(m.chat.id)	
            msg = bot.send_message(m.chat.id, lengstr(1,62)) 
    if m.text==lengstr(1,9):
            zaprosi=zaprosi_cl(m.chat.id)
            keyboard = types.InlineKeyboardMarkup(row_width=1)
            for i in range(0,len(zaprosi)):
                keyboard.add(*[types.InlineKeyboardButton(text=str(name),callback_data='zaprs'+str(i)) for name in [zaprosi[i][3][:29]]])                 
            msg = bot.send_message(m.chat.id, lengstr(1,58),reply_markup=keyboard) 
    if m.chat.id in add_poi:
        add_poi.remove(m.chat.id)	
        ref=refund1()
        dobavit_poi1(m.chat.id,m.text,ref)
        msg = bot.send_message(m.chat.id, lengstr(1,56)) 
    if add_vopros_admin==1 and m.chat.id==admin_id:
        zxvopros=m.text
        for i in range(0,len(zxvopros)):
            if zxvopros[i]=='\n':
                vopros=zxvopros[:i]
                stroka=zxvopros[i+1:]+'\n'
                break
        otvet=[]
        last=0
        for i in range(1,len(stroka)):
            if stroka[i]=='\n':
                otvet.append(stroka[last:i])
                last=i+1
        otvet=pickle.dumps(otvet)
        add_vopros_opt(vopros,otvet,work_with_opr)       
        add_vopros_admin=0
        opr_id=work_with_opr
        zz='Error'
        zz=chek_opros_full(opr_id)
        keyboard = types.InlineKeyboardMarkup(row_width=2)
        keyboard.add(types.InlineKeyboardButton(text='Удалить вопрос',callback_data=str(opr_id)+'delvopr'),
                    types.InlineKeyboardButton(text='Добавить вопрос',callback_data=str(opr_id)+'addvopr'),
                    types.InlineKeyboardButton(text='Удалить опрос',callback_data=str(opr_id)+'delopr'),
                    types.InlineKeyboardButton(text='Разослать опрос',callback_data=str(opr_id)+'razoslat'))
        msg = bot.send_message(m.chat.id, zz,reply_markup=keyboard) 
        return
######################### rassilka texta po filtram oprosa
    if txtrassilki==1:
        txtrassilki=m.text
        print(filter,txtrassilki,work_with_opr)
        rassilka_po_filtru(m.text,work_with_opr)
        txtrassilki=''

#######################
 
    if add_opr==1 and m.chat.id==admin_id:
        add_opr_in_db(m.text,oprs)
        work_with_opr=oprs
        add_opr=2
        msg = bot.send_message(m.chat.id, lengstr(1,21))   
        return
    if add_opr==2 and m.chat.id==admin_id:
        add_vopros(m.text,work_with_opr)
        add_opr=3
        msg = bot.send_message(m.chat.id, lengstr(1,22)) 
        return	
    if add_opr==3 and m.chat.id==admin_id:
        otvet=[]
        stroka=m.text+'\n'
        last=0
        for i in range(1,len(stroka)):
            if stroka[i]=='\n':
                otvet.append(stroka[last:i])
                last=i+1
        otvet=pickle.dumps(otvet)
        add_otvet(otvet,work_with_opr)
        add_opr=0
        msg = bot.send_message(m.chat.id, lengstr(1,23)) 
        return			
    if lengstr(1,15)==m.text and m.chat.id==admin_id:
            dsa=all_oprosi()
            keyboard = types.InlineKeyboardMarkup(row_width=1)
            for i in range(0,len(dsa)):
                keyboard.add(*[types.InlineKeyboardButton(text=name,callback_data='cop'+str(dsa[i][0])) for name in [dsa[i][1]]]) 
            keyboard.add(*[types.InlineKeyboardButton(text=name,callback_data='addopr') for name in [lengstr(1,19)]])                 
            msg = bot.send_message(m.chat.id, lengstr(1,18)+' '+str(len(dsa)),reply_markup=keyboard)      
    if lengstr(1,33)==m.text and m.chat.id==admin_id:
            dsa=all_oprosi()
            keyboard = types.InlineKeyboardMarkup(row_width=1)
            for i in range(0,len(dsa)):
                keyboard.add(*[types.InlineKeyboardButton(text=name,callback_data='stt'+str(dsa[i][0])) for name in [dsa[i][1]]])                 
            msg = bot.send_message(m.chat.id, lengstr(1,35),reply_markup=keyboard) 	
    if lengstr(1,16)==m.text and m.chat.id==admin_id:
            dsa=all_oprosi()
            keyboard = types.InlineKeyboardMarkup(row_width=1)
            for i in range(0,len(dsa)):
                keyboard.add(*[types.InlineKeyboardButton(text=name,callback_data='ras'+str(dsa[i][0])) for name in [dsa[i][1]]])                 
            msg = bot.send_message(m.chat.id, lengstr(1,34),reply_markup=keyboard) 		
    if lengstr(1,17)==m.text and m.chat.id==admin_id:
            dsa=all_groups()
            keyboard = types.InlineKeyboardMarkup(row_width=1)
            for i in range(0,len(dsa)):
                keyboard.add(*[types.InlineKeyboardButton(text=name,callback_data='grp'+str(dsa[i][0])) for name in [dsa[i][1][:29]]])   
            keyboard.add(*[types.InlineKeyboardButton(text=name,callback_data='addgroup') for name in [lengstr(1,44)]])                 
            msg = bot.send_message(m.chat.id, lengstr(1,43),reply_markup=keyboard) 	  
    if new_group==2 and m.chat.id==admin_id:
            ngroup.tema=m.text 
            new_group=3			
            msg = bot.send_message(m.chat.id, lengstr(1,46))
            return			
    if new_group==3 and m.chat.id==admin_id:
            if m.text.isdigit():			
               new_group=0	
               ngroup.time=m.text			   
               add_group(ngroup)			   
               msg = bot.send_message(m.chat.id, lengstr(1,49))	
            else:
               msg = bot.send_message(m.chat.id, lengstr(1,48))
    if change_group!='' and m.chat.id==admin_id:
        lisd=change_group[0]
        grid=int(change_group[1:])
        change_group=''
        if lisd=='z':
            if m.text.isdigit():
                ch_zaderjku(grid,m.text)
                msg = bot.send_message(m.chat.id, lengstr(1,51))
            else:
                msg = bot.send_message(m.chat.id, lengstr(1,48))  
        if lisd=='t':
                ch_tema(grid,m.text)
                msg = bot.send_message(m.chat.id, lengstr(1,51))	
    print(m.text, admin_id, m.chat.id)
    if m.text==lengstr(1,42) and m.chat.id==admin_id:
            keyboard = types.InlineKeyboardMarkup(row_width=1)  
            keyboard.add(*[types.InlineKeyboardButton(text=name,callback_data='opciitext') for name in [lengstr(1,70)]])        
            keyboard.add(*[types.InlineKeyboardButton(text=name,callback_data='opciipass') for name in [lengstr(1,71)]])  			
            msg = bot.send_message(m.chat.id, lengstr(1,69),reply_markup=keyboard) 	  
    if what_change=='p' and m.chat.id==admin_id:
        change_pass(m.text)
        password=m.text	
        what_change=''		
        msg = bot.send_message(m.chat.id, lengstr(1,73)) 
    return

	
def change_pass(text):
    conn = sqlite3.connect('CONV.sqlite')
    cursor = conn.cursor()
    cursor.execute("DELETE FROM PASS")
    conn.commit()
    conn.close() 
    conn = sqlite3.connect('CONV.sqlite')
    cursor = conn.cursor()
    cursor.execute("insert into PASS values (:a1) ", {"a1": text})
    conn.commit()
    conn.close()     	

	
def start_prov(id):
    results1=[]
    results2=[]
    conn = sqlite3.connect('BD.sqlite')
    cursor = conn.cursor()
    cursor.execute("SELECT * FROM USERS_CL")
    results = cursor.fetchall()   
    conn.close()
    for i in range(0,len(results)):
        results1.append(results[i][0])
    conn = sqlite3.connect('BD.sqlite')
    cursor = conn.cursor()
    cursor.execute("SELECT * FROM USERS_UR")
    results = cursor.fetchall()   
    conn.close()
    for i in range(0,len(results)):
        results2.append(results[i][0])	
    if id in results1:
       return(2)	
    if id in results2:
       return(1)	
    return(0) 
    
	
def chek_razgovor(id):
    conn = sqlite3.connect('CONV.sqlite')
    cursor = conn.cursor()
    cursor.execute("SELECT * FROM RAZ")
    results = cursor.fetchall()
    conn.close()
    for i in range(0,len(results)):
        if results[i][0]==id:
            conn = sqlite3.connect('CONV.sqlite')
            cursor = conn.cursor()
            cursor.execute("DELETE FROM RAZ WHERE FROM1 = :cod AND TO1=:da", {"cod": id,"da": results[i][1]})
            conn.commit()
            conn.close()            
            return(results[i][1])	
	
def from_us():	
    conn = sqlite3.connect('CONV.sqlite')
    cursor = conn.cursor()
    cursor.execute("SELECT * FROM RAZ")
    results = cursor.fetchall()
    conn.close()
    bs=[]
    for i in range(0,len(results)):
        bs.append(results[i][0])
    return(bs)   

   
def razgovor_go(from1,to1):
    conn = sqlite3.connect('CONV.sqlite')
    cursor = conn.cursor()
    cursor.execute("insert into RAZ values (:a1, :a2) ", {"a1": from1,"a2": to1})
    conn.commit()
    conn.close()     

def chek_kod(kod):
    conn = sqlite3.connect('CONV.sqlite')
    cursor = conn.cursor()
    cursor.execute("SELECT CL_ID FROM ZAI WHERE CODE=:id",{"id": kod})
    results = cursor.fetchall()
    conn.close()
    if len(results)==0:
       return(0)
    else:
       return(results[0][0])	


	
	
def zaprosi_delete(id,zp_id):		
    conn = sqlite3.connect('CONV.sqlite')
    cursor = conn.cursor()
    cursor.execute("SELECT * FROM ZAI WHERE CL_ID=:id",{"id": id})
    results = cursor.fetchall()
    conn.close()
    text=results[zp_id][3]	
    conn = sqlite3.connect('CONV.sqlite')
    cursor = conn.cursor()
    cursor.execute("DELETE FROM ZAI WHERE CL_ID=:id AND TEXT = :text", {"id": id,"text": text})
    conn.commit()
    conn.close() 
    conn = sqlite3.connect('CONV.sqlite')
    cursor = conn.cursor()
    cursor.execute("DELETE FROM Thread WHERE CL_ID=:id AND TEXT = :text", {"id": id,"text": text})
    conn.commit()
    conn.close() 	
    conn = sqlite3.connect('CONV.sqlite')
    cursor = conn.cursor()
    cursor.execute("DELETE FROM ALL_COD WHERE COD = :cod", {"cod": results[zp_id][1]})
    conn.commit()
    conn.close() 
    mid=pickle.loads(results[zp_id][2])
    for i in range(0,len(mid[0])):
        try:
            msg = bot.delete_message(mid[0][i],mid[1][i])
        except Exception:
            aa=1
	
def zaprosi_cl(id):	
    conn = sqlite3.connect('CONV.sqlite')
    cursor = conn.cursor()
    cursor.execute("SELECT * FROM ZAI WHERE CL_ID=:id",{"id": id})
    results = cursor.fetchall()
    conn.close()
    return(results)	
	
def dobavit_poi1(id,text,ref):
    conn = sqlite3.connect('CONV.sqlite')
    cursor = conn.cursor()
    cursor.execute("SELECT TEMA FROM ZAI WHERE CL_ID=:id AND TEXT='' ",{"id": id})
    results = cursor.fetchall()
    conn.close()
    tema=results[0][0]
    conn = sqlite3.connect('CONV.sqlite')
    cursor = conn.cursor()
    cursor.execute("SELECT * FROM GROUPS WHERE TEMA =:id",{"id": tema})
    results = cursor.fetchall()
    conn.close() 
    send_array=[]
    message_array=[]
    for i in range(0,len(results)):
        if results[i][3]==0:
            send_array.append(results[i][0])
            msg = bot.send_message(results[i][0],'Новый клиент\nОписание:\n'+text+'\n\nКод для связи с клиентом: '+ref)	
            message_array.append(msg.message_id)	
        else:
            tt=int(time.time())
            tt=tt+(results[i][3]*3600)
            conn = sqlite3.connect('CONV.sqlite')
            cursor = conn.cursor()
            cursor.execute("insert into Thread values (:a1, :a2, :a3, :a4, :a5) ", {"a1": text,"a2": tt,"a3": id, "a4":results[i][0],"a5": ref})
            conn.commit()
            conn.close()            			
    print(send_array)
    print(message_array)
    mid=[]
    mid.append(send_array)
    mid.append(message_array)
    mid = pickle.dumps(mid)
    conn = sqlite3.connect('CONV.sqlite')
    cursor = conn.cursor()
    cursor.execute("UPDATE ZAI SET TEXT = :time, CODE = :ref, MESSAGE_ID = :mid WHERE CL_ID =:id AND TEXT = '' ",{"id": id,"time": text,"ref": ref,"mid": mid})
    conn.commit()
    conn.close()  	
	
def refund1():
    conn = sqlite3.connect('CONV.sqlite')
    cursor = conn.cursor()
    cursor.execute("SELECT COD FROM ALL_COD ")
    results = cursor.fetchall()
    conn.close()
    reflist=results
    gh=1
    while gh==1:
            chars=string.ascii_uppercase + string.digits
            zz=''.join(random.choice(chars) for _ in range(8))   
            if zz not in reflist:
                conn = sqlite3.connect('CONV.sqlite')
                cursor = conn.cursor()
                cursor.execute("insert into ALL_COD values (:a1) ", {"a1": zz})
                conn.commit()
                conn.close()                
                return(zz)	
	            
	
def   dobavit_poi(id,tema):
    conn = sqlite3.connect('CONV.sqlite')
    cursor = conn.cursor()
    cursor.execute("insert into ZAI values (:a1, :a2, :a3, :a4, :a5) ", {"a1": id,"a2": '',"a3": '',"a4": '',"a5": tema})
    conn.commit()
    conn.close() 
	
	
def groups_tems():	
    conn = sqlite3.connect('CONV.sqlite')
    cursor = conn.cursor()
    cursor.execute("SELECT TEMA FROM GROUPS")
    results = cursor.fetchall()
    conn.close()
    zz=list(set(results))
    return(zz)   



	
def delete_group(id):
    conn = sqlite3.connect('CONV.sqlite')
    cursor = conn.cursor()
    cursor.execute("DELETE FROM GROUPS WHERE GROUP_ID=:id", {"id": id})
    conn.commit()
    conn.close()  	
	
	
def ch_tema(id,time):
    conn = sqlite3.connect('CONV.sqlite')
    cursor = conn.cursor()
    cursor.execute("UPDATE GROUPS SET TEMA = :time WHERE GROUP_ID =:id",{"id": id,"time": time})
    conn.commit()
    conn.close()  	
	
def ch_zaderjku(id,time):
    conn = sqlite3.connect('CONV.sqlite')
    cursor = conn.cursor()
    cursor.execute("UPDATE GROUPS SET TIME = :time WHERE GROUP_ID =:id",{"id": id,"time": time})
    conn.commit()
    conn.close()  	
	
def chek_group(id):
    conn = sqlite3.connect('CONV.sqlite')
    cursor = conn.cursor()
    cursor.execute("SELECT * FROM GROUPS WHERE GROUP_ID =:id",{"id": id})
    results = cursor.fetchall()
    conn.close()
    zz=results[0]
    return(zz)     	
	
def add_group(grou):
    conn = sqlite3.connect('CONV.sqlite')
    cursor = conn.cursor()
    cursor.execute("insert into GROUPS values (:a1, :a2, :a3, :a4) ", {"a1": grou.id,"a2": grou.name,"a3": grou.tema,"a4": grou.time,})
    conn.commit()
    conn.close()    	
	

def all_oprosi():
    conn = sqlite3.connect('BD.sqlite')
    cursor = conn.cursor()
    cursor.execute("SELECT * FROM OPROSI")
    results = cursor.fetchall()
    conn.close()
    print(results)
    return(results)   

def add_urist(user_id):
    conn = sqlite3.connect('BD.sqlite')
    cursor = conn.cursor()
    cursor.execute("insert into USERS_UR values (:post_id) ", {"post_id": user_id})
    conn.commit()
    conn.close()
	
def add_opr_in_db(name,id):
    conn = sqlite3.connect('BD.sqlite')
    cursor = conn.cursor()
    cursor.execute("insert into OPROSI values (:post_id, :name) ", {"post_id": id,"name": name})
    conn.commit()
    conn.close()	

	
def add_vopros(vopros,id):
    conn = sqlite3.connect('BD.sqlite')
    cursor = conn.cursor()
    cursor.execute("insert into VOPROSI values (:post_id, :name, :otvet) ", {"post_id": id,"name": vopros,"otvet": ''})
    conn.commit()
    conn.close()
	
def add_otvet(vopros,id):
    conn = sqlite3.connect('BD.sqlite')
    cursor = conn.cursor()
    cursor.execute("UPDATE VOPROSI SET OTVETI = :name WHERE OPROS_ID = :id ", {"name": vopros,"id": id})
    conn.commit()
    conn.close()	

def  add_vopros_opt(vopros,otvet,id) :
    conn = sqlite3.connect('BD.sqlite')
    cursor = conn.cursor()
    cursor.execute("insert into VOPROSI values (:post_id, :name, :otvet) ", {"post_id": id,"name": vopros,"otvet": otvet})
    conn.commit()
    conn.close()

	
def add_client(user_id):
    conn = sqlite3.connect('BD.sqlite')
    cursor = conn.cursor()
    cursor.execute("insert into USERS_CL values (:post_id) ", {"post_id": user_id})
    conn.commit()
    conn.close()

	
def chek_opros_full(id):
    res='Название опроса: '
    conn = sqlite3.connect('BD.sqlite')
    cursor = conn.cursor()
    cursor.execute("SELECT NAME FROM OPROSI WHERE ID=:id", {"id": id})
    results = cursor.fetchall()
    conn.close()  
    print(results)
    res+=results[0][0]+'\n'
    conn = sqlite3.connect('BD.sqlite')
    cursor = conn.cursor()
    cursor.execute("SELECT TEXT FROM VOPROSI WHERE OPROS_ID=:id", {"id": id})
    results1 = cursor.fetchall()
    cursor.execute("SELECT OTVETI FROM VOPROSI WHERE OPROS_ID=:id", {"id": id})
    results2 = cursor.fetchall()
    conn.close()    
    for i in range(0,len(results1)):
        res+='\nВопрос №'+str(i+1)+' : '+results1[i][0]+'\nВарианты ответов:\n'
        otveti=pickle.loads(results2[i][0])
        for k in range(0,len(otveti)):
                res+=otveti[k]+'\n'
    return(res)			
	
def delete_opros(id):
    conn = sqlite3.connect('BD.sqlite')
    cursor = conn.cursor()
    cursor.execute("DELETE FROM OPROSI WHERE ID=:id", {"id": id})
    conn.commit()
    conn.close()  
    conn = sqlite3.connect('BD.sqlite')
    cursor = conn.cursor()
    cursor.execute("DELETE FROM VOPROSI WHERE OPROS_ID=:id", {"id": id})
    conn.commit()
    conn.close()    
    return()	
	
	
	
	
def spisok_voprosov(id):
    conn = sqlite3.connect('BD.sqlite')
    cursor = conn.cursor()
    cursor.execute("SELECT TEXT FROM VOPROSI WHERE OPROS_ID = :id", {"id": id})
    results = cursor.fetchall()
    conn.close()
    return(results)    

def   delete_vopros(num,opr_id):
    conn = sqlite3.connect('BD.sqlite')
    cursor = conn.cursor()
    cursor.execute("SELECT TEXT FROM VOPROSI WHERE OPROS_ID = :id", {"id": opr_id})
    results = cursor.fetchall()
    conn.close()
    tt=results[int(num)][0]
    conn = sqlite3.connect('BD.sqlite')
    cursor = conn.cursor()
    cursor.execute("DELETE FROM VOPROSI WHERE TEXT=:id", {"id": tt})
    conn.commit()
    conn.close()         





def  rassilka_oprosa(opr_id):
    conn = sqlite3.connect('BD.sqlite')
    cursor = conn.cursor()
    cursor.execute("DELETE FROM OTVETI WHERE OPROS_ID=:id", {"id": opr_id})
    conn.commit()
    conn.close()     
    conn = sqlite3.connect('BD.sqlite')
    cursor = conn.cursor()
    cursor.execute("SELECT TEXT FROM VOPROSI WHERE OPROS_ID = :id", {"id": opr_id})
    results = cursor.fetchall()
    conn.close()
    conn = sqlite3.connect('BD.sqlite')
    cursor = conn.cursor()
    aa=[]
    aa=pickle.dumps(aa)
    for i in range(0,len(results)):
        cursor.execute("insert into OTVETI values (:post_id, :name, :otvet, :otvet1) ", {"post_id": opr_id,"name": i,"otvet": aa, "otvet1": aa})
        conn.commit()
    conn.close()
    t = threading.Thread(target=clock, args=(opr_id,))
    t.start()




	
def clock(opr_id):
    conn = sqlite3.connect('BD.sqlite')
    cursor = conn.cursor()
    cursor.execute("SELECT * FROM USERS_CL")
    results = cursor.fetchall()
    conn.close()    
    conn = sqlite3.connect('BD.sqlite')
    cursor = conn.cursor()
    cursor.execute("SELECT * FROM VOPROSI WHERE OPROS_ID = :id", {"id": opr_id})
    results1 = cursor.fetchall()
    conn.close()    
    for i in range(0,len(results)):
        if i % 15==0:
            time.sleep(1)
        keyboard = types.InlineKeyboardMarkup(row_width=1)
        otveti=pickle.loads(results1[i][2])  
        keyboard.add(*[types.InlineKeyboardButton(text=name,callback_data='otv'+str(otveti.index(name))+':'+str(opr_id)+'†'+str(0)) for name in otveti])            		
        msg=bot.send_message(results[i][0],lengstr(1,30)+'\n\n'+results1[0][1],reply_markup=keyboard)



def add_otvet_in(user_id,opr_id,vopros_id,otvet_id):
    conn = sqlite3.connect('BD.sqlite')
    cursor = conn.cursor()
    cursor.execute("SELECT * FROM OTVETI WHERE OPROS_ID = :opr_id and VOPROS_ID = :vopros_id", {"opr_id": opr_id,"vopros_id": vopros_id})
    results = cursor.fetchall()
    conn.close()
    users=pickle.loads(results[0][2])
    print(users)
    otveti=pickle.loads(results[0][3])	
    print(otveti)
    users.append(user_id)
    otveti.append(otvet_id)
    users=pickle.dumps(users)
    otveti=pickle.dumps(otveti)
    conn = sqlite3.connect('BD.sqlite')
    cursor = conn.cursor()
    cursor.execute("UPDATE OTVETI SET USERS=:users WHERE OPROS_ID = :opr_id and VOPROS_ID = :vopros_id ", {"users": users,"opr_id": opr_id,"vopros_id": vopros_id})
    cursor.execute("UPDATE OTVETI SET ANSW=:otveti WHERE OPROS_ID = :opr_id and VOPROS_ID = :vopros_id ", {"otveti": otveti,"opr_id": opr_id,"vopros_id": vopros_id})
    conn.commit()
    conn.close()    

	
	
def   chek_next_question(user_id,bib,opr_id,vopros_id):    
    conn = sqlite3.connect('BD.sqlite')
    cursor = conn.cursor()
    vopros_id=int(vopros_id)
    cursor.execute("SELECT * FROM VOPROSI WHERE OPROS_ID = :id", {"id": opr_id})
    results1 = cursor.fetchall()
    conn.close()  
    vopros_id+=1
    if len(results1)==vopros_id:
           msg = bot.edit_message_text(chat_id=user_id, message_id=bib, text=lengstr(1,31)) 	
           return
    keyboard = types.InlineKeyboardMarkup(row_width=1)
    otveti=pickle.loads(results1[vopros_id][2])  
    keyboard.add(*[types.InlineKeyboardButton(text=name,callback_data='otv'+str(otveti.index(name))+':'+str(opr_id)+'†'+str(vopros_id)) for name in otveti])            		
    msg=bot.edit_message_text(chat_id=user_id, message_id=bib,text=results1[vopros_id][1],reply_markup=keyboard)
	
	
	
def stat_opros(opr_id):
    conn = sqlite3.connect('BD.sqlite')
    cursor = conn.cursor()
    cursor.execute("SELECT * FROM OTVETI WHERE OPROS_ID = :opr_id", {"opr_id": opr_id})
    results1 = cursor.fetchall()
    conn.close()
    conn = sqlite3.connect('BD.sqlite')
    cursor = conn.cursor()
    cursor.execute("SELECT * FROM VOPROSI WHERE OPROS_ID = :id", {"id": opr_id})
    results2 = cursor.fetchall()
    conn.close()
    res=''
    if len(results1)==0:
        res=lengstr(1,36)
        return(res)
    for i in range(0,len(results2)):
            col_vo=len(pickle.loads(results1[i][3]))
            otv=pickle.loads(results1[i][3])
            res+='Вопрос №'+str(i+1)+'\n'+results2[i][1]+'\nВсего ответило: '+str(col_vo)+'\n\n'
            col_ot=len(pickle.loads(results2[i][2]))
            var_ot=pickle.loads(results2[i][2])
            for k in range(0,col_ot):
                sum=0
                for j in range(0,col_vo):
                    if str(otv[j])==str(k):
                        sum+=1
                res+=var_ot[k]+' - '+str(sum)+'\n'
            res+='\n'
    return(res)
  
def chek_op(opr_id):
    conn = sqlite3.connect('BD.sqlite')
    cursor = conn.cursor()
    cursor.execute("SELECT * FROM OTVETI WHERE OPROS_ID = :opr_id", {"opr_id": opr_id})
    results1 = cursor.fetchall()
    conn.close()
    if len(results1)==0:
        res=0
    else:
        res=1
    return(res)
  
  
  
  
def create_filter(opr_id):
    global filter
    filter=[]
    conn = sqlite3.connect('BD.sqlite')
    cursor = conn.cursor()
    cursor.execute("SELECT TEXT FROM VOPROSI WHERE OPROS_ID=:id", {"id": opr_id})
    results1 = cursor.fetchall()
    cursor.execute("SELECT OTVETI FROM VOPROSI WHERE OPROS_ID=:id", {"id": opr_id})
    results2 = cursor.fetchall()
    conn.close()    
    for i in range(0,len(results1)):
        filter.append([])
        otveti=pickle.loads(results2[i][0])
        for k in range(0,len(otveti)):
                filter[-1].append(1)
    print(filter)	    
	
	
	
	
	
def keyboard_for_filter(opr_id):
    global filter	
    res=''

    fulter_emoji=copy.deepcopy(filter)
    for i in range(0,len(filter)):
        for k in range(0,len(filter[i])):
            if filter[i][k]==1:
                fulter_emoji[i][k]='✅'
            else:
                fulter_emoji[i][k]='❌'

    keyboard = types.InlineKeyboardMarkup(row_width=2)
    conn = sqlite3.connect('BD.sqlite')
    cursor = conn.cursor()
    cursor.execute("SELECT TEXT FROM VOPROSI WHERE OPROS_ID=:id", {"id": opr_id})
    results1 = cursor.fetchall()
    cursor.execute("SELECT OTVETI FROM VOPROSI WHERE OPROS_ID=:id", {"id": opr_id})
    results2 = cursor.fetchall()
    conn.close() 

    for i in range(0,len(results1)):
        keyboard.add(types.InlineKeyboardButton(text='Вопрос №'+str(i+1),callback_data='„'))
        otveti=pickle.loads(results2[i][0])
        keyboard.add(*[types.InlineKeyboardButton(text=fulter_emoji[i][otveti.index(name)]+' '+name,callback_data='ifm'+str(i)+':'+str(otveti.index(name))+'†'+str(opr_id)) for name in otveti])
    keyboard.add(types.InlineKeyboardButton(text=lengstr(1,38),callback_data='txtras'))
    return(keyboard)	
	
	
def  rassilka_po_filtru(txt,opr_id):
    global filter	
    conn = sqlite3.connect('BD.sqlite')
    cursor = conn.cursor()
    cursor.execute("SELECT * FROM OTVETI WHERE OPROS_ID = :opr_id", {"opr_id": opr_id})
    results = cursor.fetchall()
    conn.close()
    users_big=[]
    for i in range(0,len(filter)):
        spisokall=pickle.loads(results[i][2])
        for i in spisokall:
            users_big.append(i)
    users_big=list(set(users_big))
    for i in range(0,len(filter)):
        users_small=[]
        spisokall=pickle.loads(results[i][2])
        answersall=pickle.loads(results[i][3])
        for k in range(0,len(filter[i])):
            if filter[i][k]==0:
                continue
            else:
                for j in range(0,len(answersall)):
               	    if str(answersall[j])==str(k):
                        users_small.append(spisokall[j])	
        print(users_small)
        for m in range(0,len(users_big)):
            try:
             if users_big[m] in users_small:
                continue
             else:
                del users_big[m]
            except Exception:
             continue			
    print(users_big)
    t = threading.Thread(target=clock1, args=(txt,users_big))
    t.start()
 


def all_groups():
    conn = sqlite3.connect('CONV.sqlite')
    cursor = conn.cursor()
    cursor.execute("SELECT * FROM GROUPS")
    results = cursor.fetchall()
    conn.close() 
    return(results)    
      	
def clock1(txt,arr):
    for i in range(0,len(arr)):
        if i % 15==0:
            time.sleep(1)          		
        msg=bot.send_message(arr[i],txt)
	
	
def forserer():
    conn = sqlite3.connect('CONV.sqlite')
    cursor = conn.cursor()
    array=[]
    now_time=int(time.time())
    cursor.execute("SELECT * FROM Thread WHERE TIME <= :time",{"time":now_time})
    results = cursor.fetchall()	
    print(results)
    cursor.execute("DELETE FROM Thread WHERE TIME <= :time",{"time":now_time})
    conn.commit()    
    conn.close()
    for i in range(0,len(results)):
            client_id=results[i][2]
            grid=results[i][3]
            msg = bot.send_message(results[i][3],'Новый клиент\nОписание:\n'+results[i][0]+'\n\nКод для связи с клиентом: '+results[i][4])	
            msid=msg.message_id	 
            conn = sqlite3.connect('CONV.sqlite')
            cursor = conn.cursor()
            cursor.execute("SELECT * FROM ZAI WHERE CL_ID=:id AND TEXT=:text ",{"id": client_id,"text": results[i][0]})
            results1 = cursor.fetchall()
            conn.close()
            mid=results1[0][2]		
            mid=pickle.loads(mid)
            mid[0].append(grid)
            mid[1].append(msid)
            print(mid)
            mid = pickle.dumps(mid)
            conn = sqlite3.connect('CONV.sqlite')
            cursor = conn.cursor()
            cursor.execute("UPDATE ZAI SET MESSAGE_ID = :mid WHERE CL_ID =:id AND TEXT = :text ",{"id": client_id,"mid": mid,"text": results[i][0],"mid": mid})
            conn.commit()
            conn.close()            			

			
schedule.every(1).minutes.do(forserer)
def lal():
    while 1:
        schedule.run_pending()
        time.sleep(1)
_thread.start_new_thread(lal,())	
	






@bot.message_handler(content_types=['document'])
def photoget(message):
    global wb, wb1, sheet, sheet1, what_change, admin_id
    if what_change=='t'	and message.chat.id==admin_id:
        what_change=''
        fileid=(message.document.file_id)
        bb=bot.get_file(fileid)
        bb=bb.file_path
        logo = urllib.request.urlopen("https://api.telegram.org/file/bot"+TOKEN+"/"+bb).read()
        f = open('leng.xlsx', "wb")
        f.write(logo)
        f.close()
        wb1 = openpyxl.load_workbook(filename = 'leng.xlsx')
        sheet1 = wb1['test']
        keyboard = types.ReplyKeyboardMarkup(resize_keyboard=True)
        keyboard.add(*[types.InlineKeyboardButton(text=name,callback_data=name) for name in [lengstr(1,15),lengstr(1,16),lengstr(1,17),lengstr(1,33),lengstr(1,42)]])
        msg = bot.send_message(message.chat.id, lengstr(1,73),reply_markup=keyboard) 
	
if __name__ == '__main__':

            bot.polling(none_stop=True)
