import telebot
import time
import threading
import datetime
import copy
import re
import string
from datetime import datetime
import schedule
from telebot import types
import _pickle as pickle
import os
import _thread
import urllib
import random
import sqlite3
import openpyxl
global o_chat
global inadmin, password, admin_id, oprcall, oprs, add_opr, work_with_opr, add_vopros_admin, filter,txtrassilki, new_group, ngroup, change_group,add_poi, vvod_koda, what_change, opcii_menu, otz_g,add_otz,prs_otz, txtrassilkiua
class group:
    id=0
    name=''
    tema=''
    time=0
what_change=''
txtrassilkiua=''	
def chek_pass():
    conn = sqlite3.connect('CONV.sqlite')
    cursor = conn.cursor()
    cursor.execute("SELECT * FROM PASS")
    results = cursor.fetchall()
    conn.close()	
    print(results[0][0])
    return(results[0][0])	
input = open('opt.pkl', 'rb')
opcii_menu = pickle.load(input)
input.close()
#output = open('opt.pkl', 'wb')
#pickle.dump(opcii_menu, output, 2)
#output.close()	
vvod_koda=[]
change_group=''
add_poi=[]
add_otz=[]
add_vopros_admin=0
oprcall=1
add_opr=0
o_chat=-1001272606508
otz_g=-1001206621069
new_group=0
inadmin=[]
filter=[]
prs_otz=[]
txtrassilki=''
password=chek_pass()
input = open('admin.pkl', 'rb')
admin_id = pickle.load(input)
input.close()
input = open('oprs.pkl', 'rb')
oprs = pickle.load(input)
input.close()
print(admin_id)
global wb, sheet
wb = openpyxl.load_workbook(filename = 'leng.xlsx')
sheet = wb['test']
val = sheet.cell(row=1, column=1)
print(val.value)
def  lengstr(leng,strn):
    bb=sheet.cell(row=strn, column=leng).value
    return str(bb)
TOKEN = '610139121:AAGVi9j83DjZ-RI7rvEeAEt4H4R5Lp31hKA'
WEBHOOK_HOST = '95.46.98.126'
WEBHOOK_PORT = 80  # 443, 80, 88 или 8443 (порт должен быть открыт!)
WEBHOOK_LISTEN = '0.0.0.0'  # На некоторых серверах придется указывать такой же IP, что и выше

WEBHOOK_SSL_CERT = './webhook_cert.pem'  # Путь к сертификату
WEBHOOK_SSL_PRIV = './webhook_pkey.pem'  # Путь к приватному ключу

WEBHOOK_URL_BASE = "https://%s:%s" % (WEBHOOK_HOST, WEBHOOK_PORT)
WEBHOOK_URL_PATH = "/%s/" % (TOKEN)

bot = telebot.TeleBot(TOKEN)


class WebhookServer(object):
    @cherrypy.expose
    def index(self):
        if 'content-length' in cherrypy.request.headers and \
                        'content-type' in cherrypy.request.headers and \
                        cherrypy.request.headers['content-type'] == 'application/json':
            length = int(cherrypy.request.headers['content-length'])
            json_string = cherrypy.request.body.read(length).decode("utf-8")
            update = telebot.types.Update.de_json(json_string)
            bot.process_new_updates([update])
            return ''
        else:
            raise cherrypy.HTTPError(403)

# Handle '/start' and '/help'


def extract_unique_code(text):
    # Extracts the unique_code from the sent /start command.
    return text.split()[1] if len(text.split()) > 1 else None



@bot.message_handler(commands=['help', 'start'])
def start(message):
    sg=''
    sg=add_client(message.chat.id)
    wg=wh_h(message.chat.id)
    rir=[]
    if sg=='hh':
        ll=find_leng(message.chat.id)
        keyboard = types.ReplyKeyboardMarkup(row_width=2)
        if wg==0:
            keyboard.add(*[types.InlineKeyboardButton(text=name,callback_data=name) for name in [lengstr(ll,80)]])            
        keyboard.add(*[types.InlineKeyboardButton(text=name,callback_data=name) for name in [lengstr(ll,8),lengstr(ll,81)]])
        for i in range(0,4):
            if opcii_menu[i]==1:
                rir.append(lengstr(ll,82+i))
        keyboard.add(*[types.InlineKeyboardButton(text=name,callback_data=name) for name in rir])
        if wg==1:
            keyboard.add(*[types.InlineKeyboardButton(text=name,callback_data=name) for name in [lengstr(ll,80)]])
        msg = bot.send_message(message.chat.id, lengstr(ll,7),reply_markup=keyboard) 
    else:
        keyboard = types.InlineKeyboardMarkup(row_width=2)
        keyboard.add(types.InlineKeyboardButton(text='Русский',callback_data='chose_leng_ru'))
        keyboard.add(types.InlineKeyboardButton(text='Українська',callback_data='chose_leng_ua'))
        msg = bot.send_message(message.chat.id, 'Выберите язык\nОберіть мову',reply_markup=keyboard)         
	
	
@bot.message_handler(commands=['admin'])	
def admin(m):
    ll=find_leng(m.chat.id)
    inadmin.append(m.chat.id)
    keyboard = types.InlineKeyboardMarkup(row_width=2)
    keyboard.add(types.InlineKeyboardButton(text=lengstr(ll,11),callback_data='!'))
    msg = bot.send_message(m.chat.id, lengstr(ll,10),reply_markup=keyboard)    
    return()	

@bot.message_handler(commands=['userstat'])	
def userstat(m):
    as1=all_users()   
    temi=groups_tems()  
    print(temi)	
    sk='Количество пользователей: '+as1+'\n'
    for i in range(0,len(temi)):
        chekv=col_vo_zaprosov(temi[i][0])
        sk+=temi[i][0]+' Кол-во запросов: '+chekv+'\n'
    msg = bot.send_message(m.chat.id, sk)  
    return()
	
	
@bot.message_handler(commands=['userchange'])	
def userchan(m):
    if m.chat.id==admin_id or m.chat.id==71709639:
        keyboard = types.ReplyKeyboardMarkup(resize_keyboard=True)
        keyboard.add(*[types.InlineKeyboardButton(text=name,callback_data=name) for name in [lengstr(ll,6)]])
        keyboard.add(*[types.InlineKeyboardButton(text=name,callback_data=name) for name in [lengstr(ll,8),lengstr(ll,9)]])
        msg = bot.send_message(m.chat.id, 'Вот клавиатуры 2-х видов аккаунта',reply_markup=keyboard)          
    return()		
	
@bot.message_handler(commands=['chat_id'])	
def chat_idr(m):
    global new_group, ngroup, admin_id
    ll=find_leng(admin_id)
    if new_group==1 and m.from_user.id==admin_id:
        try:
            ngroup.id=m.chat.id  
            ngroup.name=m.chat.title
            new_group=2
            msg = bot.send_message(admin_id, lengstr(ll,47))
        except Exception:
            msg = bot.send_message(admin_id, lengstr(ll,48)) 
    return()	

	
@bot.message_handler(content_types=["text"])
def repeat_all_messages(message): 
    name(message)	
	
	
	
@bot.callback_query_handler(func=lambda c:True)
def inline(c):
    global oprs, add_opr, work_with_opr, add_vopros_admin, filter,txtrassilki, new_group, ngroup, change_group,add_poi, what_change, opcii_menu, otz_g
    bib=c.message.message_id
    if c.data=='chose_leng_ru':
        user_leng(c.message.chat.id,1)
        wg=wh_h(c.message.chat.id)
        ll=1
        keyboard = types.ReplyKeyboardMarkup(resize_keyboard=True)
        if wg==0:
            keyboard.add(*[types.InlineKeyboardButton(text=name,callback_data=name) for name in [lengstr(ll,80)]])            
        keyboard.add(*[types.InlineKeyboardButton(text=name,callback_data=name) for name in [lengstr(ll,8),lengstr(ll,81)]])
        for i in range(0,4):
            if opcii_menu[i]==1:
                keyboard.add(*[types.InlineKeyboardButton(text=name,callback_data=name) for name in [lengstr(ll,82+i)]])
        if wg==1:
            keyboard.add(*[types.InlineKeyboardButton(text=name,callback_data=name) for name in [lengstr(ll,80)]])
        msg = bot.send_message(c.message.chat.id, lengstr(ll,7),reply_markup=keyboard)
        return()
    if c.data=='chose_leng_ua':
        user_leng(c.message.chat.id,2)
        wg=wh_h(c.message.chat.id)
        ll=2
        keyboard = types.ReplyKeyboardMarkup(resize_keyboard=True)
        if wg==0:
            keyboard.add(*[types.InlineKeyboardButton(text=name,callback_data=name) for name in [lengstr(ll,80)]])            
        keyboard.add(*[types.InlineKeyboardButton(text=name,callback_data=name) for name in [lengstr(ll,8),lengstr(ll,81)]])
        for i in range(0,4):
            if opcii_menu[i]==1:
                keyboard.add(*[types.InlineKeyboardButton(text=name,callback_data=name) for name in [lengstr(ll,82+i)]])
        if wg==1:
            keyboard.add(*[types.InlineKeyboardButton(text=name,callback_data=name) for name in [lengstr(ll,80)]])
        msg = bot.send_message(c.message.chat.id, lengstr(ll,7),reply_markup=keyboard)
        return
    ll=find_leng(c.message.chat.id)
    if 'otv' in c.data:
        print(c.data)
        for i in range(0,len(c.data)):
            if c.data[i]==':':
                lst=i+1
                otvet_id=c.data[3:i]
            if c.data[i]=='†':
                opr_id=c.data[lst:i]
                vopros_id=c.data[i+1:]
        add_otvet_in(c.message.chat.id,opr_id,vopros_id,otvet_id)
        print(c.message.chat.id,opr_id,vopros_id,otvet_id)
        chek_next_question(c.message.chat.id,bib,opr_id,vopros_id)
        return
    if c.data == 'do_otz':
        keyboard = types.InlineKeyboardMarkup(row_width=2)
        keyboard.add(*[types.InlineKeyboardButton(text=name,callback_data='extotz') for name in [lengstr(ll,40)]])
        add_otz.append(c.message.chat.id)
        msg = bot.send_message(c.message.chat.id, lengstr(ll,93),reply_markup=keyboard) 
    if c.data == 'prosm_otz':
        keyboard = types.InlineKeyboardMarkup(row_width=2)
        keyboard.add(*[types.InlineKeyboardButton(text=name,callback_data='extotzs') for name in [lengstr(ll,40)]])
        prs_otz.append(c.message.chat.id)
        msg = bot.send_message(c.message.chat.id, lengstr(ll,96),reply_markup=keyboard) 
    if c.data =='extotz':
        add_otz.remove(c.message.chat.id)
        msg = bot.send_message(c.message.chat.id, lengstr(ll,41)) 	
    if c.data =='extotzs':
        prs_otz.remove(c.message.chat.id)
        msg = bot.send_message(c.message.chat.id, lengstr(ll,41)) 
    if c.data=='!':
        inadmin.remove(c.message.chat.id)
        msg = bot.send_message(c.message.chat.id, lengstr(ll,14))   
        return
    if c.data=='all_users_say':
        print('ia zap clock2')
        clock2(c.message.chat.id)  
        return      	
    if c.data=='%':
        add_urist(c.message.chat.id)
        keyboard = types.ReplyKeyboardMarkup(resize_keyboard=True)
        keyboard.add(*[types.InlineKeyboardButton(text=name,callback_data=name) for name in [lengstr(ll,6)]])
        msg = bot.send_message(c.message.chat.id, lengstr(ll,5),reply_markup=keyboard) 
        return		
    if c.data=='@':
        add_client(c.message.chat.id)
        keyboard = types.ReplyKeyboardMarkup(resize_keyboard=True)
        keyboard.add(*[types.InlineKeyboardButton(text=name,callback_data=name) for name in [lengstr(ll,8),lengstr(ll,9)]])
        msg = bot.send_message(c.message.chat.id, lengstr(ll,7),reply_markup=keyboard) 
        return
    if c.data=='addopr':
        oprs+=1
        output = open('oprs.pkl', 'wb')
        pickle.dump(oprs, output, 2)
        output.close()	
        add_opr=1
        msg = bot.send_message(c.message.chat.id, lengstr(ll,20))  
        return
    if c.data=='extxt':
        msg = bot.send_message(c.message.chat.id, lengstr(ll,41))   
        txtrassilki=''		
        return
######################## Vvod teksta rassilki
    if c.data=='txtras':
        keyboard = types.InlineKeyboardMarkup(row_width=2)
        keyboard.add(*[types.InlineKeyboardButton(text=name,callback_data='extxt') for name in [lengstr(ll,40)]])
        msg = bot.send_message(c.message.chat.id, lengstr(ll,39),reply_markup=keyboard)   
        txtrassilki=1		
        return
################################################
    if 'ddelopr' in c.data:
            opr_id=int(c.data[:-7])	 
            delete_opros(opr_id)
            dsa=all_oprosi()
            keyboard = types.InlineKeyboardMarkup(row_width=1)
            for i in range(0,len(dsa)):
                keyboard.add(*[types.InlineKeyboardButton(text=name,callback_data='cop'+str(dsa[i][0])) for name in [dsa[i][1]]]) 
            keyboard.add(*[types.InlineKeyboardButton(text=name,callback_data='addopr') for name in [lengstr(ll,19)]])                 
            msg = bot.send_message(c.message.chat.id, lengstr(ll,18)+' '+str(len(dsa)),reply_markup=keyboard) 
            return()			
    if 'delopr' in c.data:
        opr_id=int(c.data[:-6])	
        keyboard = types.InlineKeyboardMarkup(row_width=2)
        keyboard.add(types.InlineKeyboardButton(text='Нет',callback_data='cop'+str(opr_id)),
                    types.InlineKeyboardButton(text='Да',callback_data=str(opr_id)+'ddelopr'))
        msg = bot.send_message(c.message.chat.id, lengstr(ll,24),reply_markup=keyboard)
        return
    if 'addvopr' in c.data:
        opr_id=int(c.data[:-7])	
        add_vopros_admin=1
        work_with_opr=opr_id
        keyboard = types.InlineKeyboardMarkup(row_width=1)
        keyboard.add(types.InlineKeyboardButton(text='Отмена',callback_data='cop'+str(opr_id)))
        msg = bot.send_message(c.message.chat.id, lengstr(ll,25),reply_markup=keyboard)
        return
    if 'razoslat' in c.data:
        opr_id=int(c.data[:-8])	
        keyboard = types.InlineKeyboardMarkup(row_width=2)
        keyboard.add(types.InlineKeyboardButton(text='Нет',callback_data='cop'+str(opr_id)),
                    types.InlineKeyboardButton(text='Да',callback_data=str(opr_id)+'razuslat'))
        msg = bot.send_message(c.message.chat.id, lengstr(ll,32),reply_markup=keyboard)
        return    
    if 'razuslat' in c.data:
        opr_id=int(c.data[:-8])	
        rassilka_oprosa(opr_id)	
        msg = bot.send_message(c.message.chat.id, lengstr(ll,29))
    if 'delvopr' in c.data:
        opr_id=int(c.data[:-7])
        zz=''		
        spisok=spisok_voprosov(opr_id)
        keyboard = types.InlineKeyboardMarkup(row_width=1)
        for i in range(len(spisok)):
            zz+=str(i+1)+') '+spisok[i][0]+'\n'
            keyboard.add(types.InlineKeyboardButton(text=str(i+1),callback_data='dzx'+str(i)+':'+str(opr_id)))             
        keyboard.add(types.InlineKeyboardButton(text='Отмена',callback_data='cop'+str(opr_id)))
        msg = bot.send_message(c.message.chat.id, lengstr(ll,27)+'\n\n'+zz,reply_markup=keyboard) 
        return		
    if 'dzx' in c.data:
        for i in range(0,len(c.data)):
            if c.data[i]==':':
                num=c.data[3:i]
                opr_id=c.data[i+1:]
        delete_vopros(num,opr_id)
        msg = bot.send_message(c.message.chat.id, lengstr(ll,28))	
        return		
    if 'cop' in c.data:
        opr_id=int(c.data[3:])
        zz='Error'
        add_vopros_admin=0
        zz=chek_opros_full(opr_id)
        keyboard = types.InlineKeyboardMarkup(row_width=2)
        keyboard.add(types.InlineKeyboardButton(text='Удалить вопрос',callback_data=str(opr_id)+'delvopr'),
                    types.InlineKeyboardButton(text='Добавить вопрос',callback_data=str(opr_id)+'addvopr'),
                    types.InlineKeyboardButton(text='Удалить опрос',callback_data=str(opr_id)+'delopr'),
                    types.InlineKeyboardButton(text='Разослать опрос',callback_data=str(opr_id)+'razoslat'))
        msg = bot.send_message(c.message.chat.id, zz,reply_markup=keyboard) 
        return		
    if 'stt' in c.data:
        opr_id=int(c.data[3:])
        zz='Error'
        zz=stat_opros(opr_id)
        msg = bot.send_message(c.message.chat.id, zz) 
        return	
############################### sozdanie filtra
    if 'ras' in c.data:
        opr_id=int(c.data[3:])
        work_with_opr=opr_id
        kr=chek_op(opr_id)
        if kr==1:
           zz=lengstr(ll,37)
        else:
           zz=lengstr(ll,36)
           msg = bot.send_message(c.message.chat.id, zz) 
           return	
        create_filter(opr_id)
        keyboard=keyboard_for_filter(opr_id)
                        
        msg = bot.send_message(c.message.chat.id, zz,reply_markup=keyboard) 
        return	
    if 'ifm' in c.data:
        for i in range(0,len(c.data)):
            if c.data[i]==':':
                su=int(c.data[3:i])
                tt=i+1
            if c.data[i]=='†':
                zu=int(c.data[tt:i])
                opr_id=int(c.data[i+1:])         	
        if filter[su][zu]==1:
           filter[su][zu]=0
        else:
           filter[su][zu]=1	
        print(filter)
        keyboard=keyboard_for_filter(opr_id)
        msg = bot.edit_message_reply_markup(c.message.chat.id, bib,reply_markup=keyboard) 
        return	
    if c.data=='addgroup':
        new_group=1  
        ngroup=group()
        msg=  bot.send_message(c.message.chat.id, lengstr(ll,45))      
        return	
    if 'grp' in c.data:
        zcf=chek_group(int(c.data[3:]))	
        change_group=''
        keyboard = types.InlineKeyboardMarkup(row_width=2)	
        keyboard.add(types.InlineKeyboardButton(text='Изменить задержку',callback_data='gzadr'+str(zcf[0])),
                    types.InlineKeyboardButton(text='Изменить тему',callback_data='gtema'+str(zcf[0])),
                    types.InlineKeyboardButton(text='Удалить группу',callback_data='gdelt'+str(zcf[0])))	
        msg = bot.send_message(c.message.chat.id, 'Id: '+str(zcf[0])+'\nНазвание: '+zcf[1]+'\nТема: '+zcf[2]+'\nТема на украинском: '+str(zcf[3]),reply_markup=keyboard)     
    if 'gzadr' in c.data:
        change_group='z'+c.data[5:]
        keyboard = types.InlineKeyboardMarkup(row_width=1)	
        keyboard.add(types.InlineKeyboardButton(text=lengstr(ll,40),callback_data='grp'+c.data[5:])) 
        msg=  bot.send_message(c.message.chat.id, lengstr(ll,50),reply_markup=keyboard)    
    if 'gtema' in c.data:
        change_group='t'+c.data[5:]
        keyboard = types.InlineKeyboardMarkup(row_width=1)	
        keyboard.add(types.InlineKeyboardButton(text=lengstr(ll,40),callback_data='grp'+c.data[5:])) 
        msg=  bot.send_message(c.message.chat.id, lengstr(ll,52),reply_markup=keyboard)       
    if 'gdelt' in c.data:
        keyboard = types.InlineKeyboardMarkup(row_width=1)	
        keyboard.add(types.InlineKeyboardButton(text='Нет',callback_data='grp'+c.data[5:]),types.InlineKeyboardButton(text='Да',callback_data='ddt'+c.data[5:])) 
        msg=  bot.send_message(c.message.chat.id, lengstr(ll,53),reply_markup=keyboard)
    if 'ddt' in c.data:
        delete_group(int(c.data[3:]))
        msg=  bot.send_message(c.message.chat.id, lengstr(ll,54)) 
    if 'poisk' in c.data:
        tema=c.data[5:]
        dobavit_poi(c.message.chat.id,tema)	
        add_poi.append(c.message.chat.id)
        ss3=korotki(tema,ll)
        keyboard = types.InlineKeyboardMarkup(row_width=1)	
        keyboard.add(types.InlineKeyboardButton(text=lengstr(ll,101),callback_data='netpoi')) 
        msg=  bot.send_message(c.message.chat.id, ss3+'\n\n'+lengstr(ll,57),reply_markup=keyboard) 
########################### Otmena poiska
    if 'netpoi'==c.data:
            add_poi.remove(c.message.chat.id)
            del_poi(c.message.chat.id,tema)
            temi=groups_tems(ll)
            keyboard = types.InlineKeyboardMarkup(row_width=1)
            for i in range(0,len(temi)):
                keyboard.add(*[types.InlineKeyboardButton(text=str(name),callback_data='poisk'+str(name)) for name in [temi[i][0]]])                 
            msg = bot.send_message(c.message.chat.id, lengstr(ll,55),reply_markup=keyboard) 		
    if 'zaprs' in c.data:
        zp_id=int(c.data[5:])
        zaprosi=zaprosi_cl(c.message.chat.id)
        keyboard = types.InlineKeyboardMarkup(row_width=1)
        keyboard.add(types.InlineKeyboardButton(text=lengstr(ll,99),callback_data='zpdubl'+str(zp_id)))
        keyboard.add(types.InlineKeyboardButton(text=lengstr(ll,60),callback_data='zpdel'+str(zp_id)))
        keyboard.add(types.InlineKeyboardButton(text=lengstr(ll,100),callback_data='zpobc'+str(zp_id)))
        keyboard.add(types.InlineKeyboardButton(text=lengstr(ll,101),callback_data='zpnzd'))		
        msg=  bot.send_message(c.message.chat.id, lengstr(ll,59)+'\n'+zaprosi[zp_id][3],reply_markup=keyboard) 	
    if 'zpnzd'==c.data:
            zaprosi=zaprosi_cl(c.message.chat.id)
            keyboard = types.InlineKeyboardMarkup(row_width=1)
            for i in range(0,len(zaprosi)):
                keyboard.add(*[types.InlineKeyboardButton(text=str(name),callback_data='zaprs'+str(i)) for name in [zaprosi[i][3][:29]]])                 
            msg = bot.send_message(c.message.chat.id, lengstr(ll,58),reply_markup=keyboard)        
    if 'zpdel' in c.data:
        zp_id=int(c.data[5:])
        zaprosi=zaprosi_delete(c.message.chat.id,zp_id)	
        msg=  bot.send_message(c.message.chat.id, lengstr(ll,61)) 
    if 'zpobc' in c.data:
        zp_id=int(c.data[5:])
        zaprosi=zaprosi_obsh(c.message.chat.id,zp_id)	
        msg=  bot.send_message(c.message.chat.id, lengstr(ll,109)) 
    if 'zpdubl' in c.data:
        zp_id=int(c.data[6:])
        zaprosi=zaprosi_dunl(c.message.chat.id,zp_id)	
        msg=  bot.send_message(c.message.chat.id, lengstr(ll,102)) 			
    if 'rznxt' in c.data:
        zp_id=int(c.data[5:])
        razgovor_go(c.message.chat.id,zp_id)
        keyboard = types.InlineKeyboardMarkup(row_width=1)
        msg=  bot.edit_message_reply_markup(c.message.chat.id,bib,keyboard) 
    if 'rzend' in c.data:
        zp_id=int(c.data[5:])
        keyboard = types.InlineKeyboardMarkup(row_width=1)
        msg=  bot.edit_message_reply_markup(c.message.chat.id,bib,keyboard) 
        msg= bot.send_message(c.message.chat.id,lengstr(ll,68)) 	
        msg= bot.send_message(zp_id,lengstr(ll,68))		
    if c.data=='opciipass':
        what_change='p'
        msg= bot.send_message(c.message.chat.id,lengstr(ll,72))    
    if c.data=='opciitext':
        what_change='t'
        msg = bot.send_document(c.message.chat.id,open('leng.xlsx', 'rb') ,caption=lengstr(ll,74))
    print(c.data)		
    if 'dialup' in c.data:
        dialog=zagruzit_dialogi()
        for i in range(0,len(c.data)):
            if c.data[i]==':':
                id1=c.data[6:i]
                id2=c.data[i+1:]
        print(id1)
        print(id2)
        for i in range(0,len(dialog)):
            if dialog[i][0]==int(id1) and dialog[i][1]==int(id2):
                file = open("file.txt", "w")
                file.write(dialog[i][2])
                file.close()
                msg = bot.send_document(c.message.chat.id,open('file.txt', 'rb') ,caption=lengstr(ll,78))
    if 'dialognext' in c.data:
            step=int(c.data[10:])
            dialog=zagruzit_dialogi()
            print(len(dialog))
            dialog.reverse()
            keyboard = types.InlineKeyboardMarkup(row_width=1)  
            for i in range((step*10),len(dialog)):
                if i==step*10+10:
                    break
                keyboard.add(types.InlineKeyboardButton(text=str(dialog[i][0])+' '+str(dialog[i][1]),callback_data='dialup'+str(dialog[i][0])+':'+str(dialog[i][1])))
            if len(dialog)>(step*10+10):
                keyboard.add(types.InlineKeyboardButton(text='>>',callback_data='dialognext'+str(step+1))) 
            if step>0:
                keyboard.add(types.InlineKeyboardButton(text='<<',callback_data='dialognext'+str(step-1))) 			
            msg = bot.send_message(m.chat.id, lengstr(ll,77),reply_markup=keyboard) 
    if 'opcii_mnu' in c.data:
        mnu_change=int(c.data[-1])-1	
        if opcii_menu[mnu_change]==1:
            opcii_menu[mnu_change]=0
        else:
            opcii_menu[mnu_change]=1
        output = open('opt.pkl', 'wb')
        pickle.dump(opcii_menu, output, 2)
        output.close()
        keyboard = types.InlineKeyboardMarkup(row_width=1)  
        opcii_smile=['','','','']
        print(opcii_menu)
        for i in range(0,4):
            print(opcii_menu[i])
            if opcii_menu[i]==1:
                   opcii_smile[i]='✅'
            else:
                   opcii_smile[i]='❌'
        print(opcii_smile)
        keyboard.add(*[types.InlineKeyboardButton(text=name+' '+opcii_smile[0],callback_data='opcii_mnu1') for name in [lengstr(ll,82)]]) 
        keyboard.add(*[types.InlineKeyboardButton(text=name+' '+opcii_smile[1],callback_data='opcii_mnu2') for name in [lengstr(ll,83)]])
        keyboard.add(*[types.InlineKeyboardButton(text=name+' '+opcii_smile[2],callback_data='opcii_mnu3') for name in [lengstr(ll,84)]])
        keyboard.add(*[types.InlineKeyboardButton(text=name+' '+opcii_smile[3],callback_data='opcii_mnu4') for name in [lengstr(ll,85)]])
        keyboard.add(types.InlineKeyboardButton(text='Оповестить всех пользователей',callback_data='all_users_say'))  
        msg=  bot.edit_message_reply_markup(c.message.chat.id,bib,reply_markup=keyboard)		
    return()		
		
		
		
		
		
		
		
		
		
		
		
		
		
		
def name(m):
    global admin_id, add_opr, oprs, work_with_opr, add_vopros_admin, filter,txtrassilki, new_group, ngroup, change_group,add_poi, vvod_koda, what_change, password, opcii_menu, otz_g, txtrassilkiua
    razgovori=from_us()
    ll=find_leng(m.chat.id)
    print(m)
    if m.chat.id in add_otz:
        otzivka=m.text+' '
        first=-1
        usrt=''
        for i in range(0,len(otzivka)):
            if otzivka[i]=='@':
                first=i
            if first>-1 and otzivka[i]==' ':
                usrt=otzivka[first:i].lower()
                break
        print(usrt)
        if usrt=='':
            msg = bot.send_message(m.chat.id, lengstr(ll,95))  
            return					
        add_otz.remove(m.chat.id)
        msg =bot.forward_message(otz_g,m.chat.id, m.message_id)
        add_otz_bd(usrt,m.chat.id, msg.message_id)
        msg = bot.send_message(m.chat.id, lengstr(ll,94))
    if m.chat.id in prs_otz:
        user_p=m.text.lower()
        solv=poisk_otz(user_p)
        if len(solv)==0:
            msg=bot.send_message(m.chat.id, lengstr(ll,97))
        else:
            for i in range(0,len(solv)):
                try:
                    msg = bot.forward_message(m.chat.id,otz_g,solv[i][1])
                except Exception:
                    continue
            msg =bot.send_message(m.chat.id, lengstr(ll,106)) 
        prs_otz.remove(m.chat.id)
    if m.chat.id in razgovori:
        sdfg=m.text
        if '@' in sdfg or 'mail' in sdfg or m.chat.username.lower() in m.text.lower():
            msg = bot.send_message(m.chat.id, lengstr(ll,75))   
            return
        svizr=m.text
        svizr=svizr.replace(' ','')	
        svizr=svizr.replace('\n','')
        svizr=svizr.replace('(','')	
        svizr=svizr.replace(')','')		
        svizr=svizr.replace('-','')
        saro=re.findall('(\d+)', svizr)
        for i in range(0,len(saro)):
            if len(saro[i])>7:
                msg = bot.send_message(m.chat.id, lengstr(ll,75))   
                return
        cl_id=chek_razgovor(m.chat.id)
        zapis_conv(cl_id,m.chat.id,m.text)
        keyboard = types.InlineKeyboardMarkup(row_width=1)
        keyboard.add(types.InlineKeyboardButton(text=lengstr(ll,66),callback_data='rznxt'+str(m.chat.id)))
        keyboard.add(types.InlineKeyboardButton(text=lengstr(ll,67),callback_data='rzend'+str(m.chat.id)))
        msg = bot.send_message(cl_id, lengstr(ll,65)+str(m.chat.id)+'\n\n'+m.text,reply_markup=keyboard)
    if m.chat.id in inadmin:
        if m.text==password:
            admin_id=m.chat.id
            output = open('admin.pkl', 'wb')
            pickle.dump(admin_id, output, 2)
            output.close()
            inadmin.remove(m.chat.id)
            keyboard = types.ReplyKeyboardMarkup(row_width=2)
            keyboard.add(*[types.InlineKeyboardButton(text=name,callback_data=name) for name in [lengstr(ll,15),lengstr(ll,16),lengstr(ll,17),lengstr(ll,33),lengstr(ll,42),lengstr(ll,86)]])
            msg = bot.send_message(m.chat.id, lengstr(ll,12),reply_markup=keyboard)  
        else:
            keyboard = types.InlineKeyboardMarkup(row_width=2)
            keyboard.add(types.InlineKeyboardButton(text=lengstr(ll,11),callback_data='!'))
            msg = bot.send_message(m.chat.id, lengstr(ll,13),reply_markup=keyboard) 
    if m.chat.id in vvod_koda:
        vvod_koda.remove(m.chat.id)
        kod=m.text
        id_client=chek_kod(kod)
        if id_client==0:
            msg = bot.send_message(m.chat.id, lengstr(ll,63))
        else:
            msg = bot.send_message(m.chat.id, lengstr(ll,64))
            razgovor_go(m.chat.id,id_client)	
    if m.text==lengstr(ll,80):
        add_ww(m.chat.id)
        wg=wh_h(m.chat.id)
        print(wg)
        keyboard = types.ReplyKeyboardMarkup(row_width=2)
        if wg==0:
            keyboard.add(*[types.InlineKeyboardButton(text=name,callback_data=name) for name in [lengstr(ll,80)]])            
        keyboard.add(*[types.InlineKeyboardButton(text=name,callback_data=name) for name in [lengstr(ll,8),lengstr(ll,81)]])
        for i in range(0,4):
            if opcii_menu[i]==1:
                rir.append(lengstr(ll,82+i))
        keyboard.add(*[types.InlineKeyboardButton(text=name,callback_data=name) for name in rir])
        if wg==1:
            keyboard.add(*[types.InlineKeyboardButton(text=name,callback_data=name) for name in [lengstr(ll,80)]])
        msg = bot.send_message(m.chat.id, lengstr(ll,87),reply_markup=keyboard) 
    if m.text==lengstr(ll,81):
        keyboard = types.InlineKeyboardMarkup()
        url_button = types.InlineKeyboardButton(text=lengstr(ll,103), url="https://t.me/joinchat/BEYzx0W9A5PZ7kKcuJf4lQ")
        keyboard.add(url_button)
        bot.send_message(m.chat.id, lengstr(ll,103)+'\nhttps://t.me/joinchat/BEYzx0W9A5PZ7kKcuJf4lQ', reply_markup=keyboard)
######################### otzivi
    if m.text==lengstr(ll,82):
        keyboard = types.InlineKeyboardMarkup(row_width=1)
        keyboard.add(*[types.InlineKeyboardButton(text=str(name),callback_data='prosm_otz') for name in [lengstr(ll,91)]]) 
        keyboard.add(*[types.InlineKeyboardButton(text=str(name),callback_data='do_otz') for name in [lengstr(ll,92)]]) 
        msg = bot.send_message(m.chat.id, lengstr(ll,82),reply_markup=keyboard)
    if m.text==lengstr(ll,83):
        msg = bot.send_message(m.chat.id, lengstr(ll,89))	
    if m.text==lengstr(ll,84):
        keyboard = types.InlineKeyboardMarkup()
        url_button = types.InlineKeyboardButton(text=lengstr(ll,104), url="https://t.me/joinchat/AAAAAE8flyT-eP3R9_-DpA")
        keyboard.add(url_button)
        bot.send_message(m.chat.id, lengstr(ll,104)+'\nhttps://t.me/joinchat/AAAAAE8flyT-eP3R9_-DpA')
######################## nastroika zaprosov
    if m.text==lengstr(ll,85):
            zaprosi=zaprosi_cl(m.chat.id)
            keyboard = types.InlineKeyboardMarkup(row_width=1)
            for i in range(0,len(zaprosi)):
                keyboard.add(*[types.InlineKeyboardButton(text=str(name),callback_data='zaprs'+str(i)) for name in [zaprosi[i][3][:29]]])                 
            msg = bot.send_message(m.chat.id, lengstr(ll,58),reply_markup=keyboard)
##########################
    if m.text==lengstr(ll,8):
            temi=groups_tems(ll)
            keyboard = types.InlineKeyboardMarkup(row_width=1)
            for i in range(0,len(temi)):
                keyboard.add(*[types.InlineKeyboardButton(text=str(name),callback_data='poisk'+str(name)) for name in [temi[i][0]]])                 
            msg = bot.send_message(m.chat.id, lengstr(ll,55),reply_markup=keyboard) 
    if m.text==lengstr(ll,6):       
            vvod_koda.append(m.chat.id)	
            msg = bot.send_message(m.chat.id, lengstr(ll,62)) 
    if m.chat.id in add_poi:
        add_poi.remove(m.chat.id)	
        ref=refund1()
        dobavit_poi1(m.chat.id,m.text,ref,m.message_id)
        msg = bot.send_message(m.chat.id, lengstr(ll,56)) 
    if add_vopros_admin==1 and m.chat.id==admin_id:
        zxvopros=m.text
        for i in range(0,len(zxvopros)):
            if zxvopros[i]=='\n':
                vopros=zxvopros[:i]
                stroka=zxvopros[i+1:]+'\n'
                break
        otvet=[]
        last=0
        for i in range(1,len(stroka)):
            if stroka[i]=='\n':
                otvet.append(stroka[last:i])
                last=i+1
        otvet=pickle.dumps(otvet)
        add_vopros_opt(vopros,otvet,work_with_opr)       
        add_vopros_admin=0
        opr_id=work_with_opr
        zz='Error'
        zz=chek_opros_full(opr_id)
        keyboard = types.InlineKeyboardMarkup(row_width=2)
        keyboard.add(types.InlineKeyboardButton(text='Удалить вопрос',callback_data=str(opr_id)+'delvopr'),
                    types.InlineKeyboardButton(text='Добавить вопрос',callback_data=str(opr_id)+'addvopr'),
                    types.InlineKeyboardButton(text='Удалить опрос',callback_data=str(opr_id)+'delopr'),
                    types.InlineKeyboardButton(text='Разослать опрос',callback_data=str(opr_id)+'razoslat'))
        msg = bot.send_message(m.chat.id, zz,reply_markup=keyboard) 
        return
######################### rassilka texta po filtram oprosa
    if txtrassilkiua==1 and m.chat.id==admin_id:
        txtrassilkiua=m.text
        print(filter,txtrassilkiua,work_with_opr)
        rassilka_po_filtru(txtrassilki,work_with_opr,txtrassilkiua)
        txtrassilkiua=''
        txtrassilki=''
        return
    if txtrassilki==1 and m.chat.id==admin_id:
        txtrassilki=m.text
        print(filter,txtrassilki,work_with_opr)
        #rassilka_po_filtru(m.text,work_with_opr)
        txtrassilkiua=1
        msg = bot.send_message(m.chat.id,lengstr(ll,105)) 
        return
#######################
 
    if add_opr==1 and m.chat.id==admin_id:
        add_opr_in_db(m.text,oprs)
        work_with_opr=oprs
        add_opr=2
        msg = bot.send_message(m.chat.id, lengstr(ll,21))   
        return
    if add_opr==2 and m.chat.id==admin_id:
        add_vopros(m.text,work_with_opr)
        add_opr=3
        msg = bot.send_message(m.chat.id, lengstr(ll,22)) 
        return	
    if add_opr==3 and m.chat.id==admin_id:
        otvet=[]
        stroka=m.text+'\n'
        last=0
        for i in range(1,len(stroka)):
            if stroka[i]=='\n':
                otvet.append(stroka[last:i])
                last=i+1
        otvet=pickle.dumps(otvet)
        add_otvet(otvet,work_with_opr)
        add_opr=0
        msg = bot.send_message(m.chat.id, lengstr(ll,23)) 
        return			
    if lengstr(ll,15)==m.text and m.chat.id==admin_id:
            dsa=all_oprosi()
            keyboard = types.InlineKeyboardMarkup(row_width=1)
            for i in range(0,len(dsa)):
                keyboard.add(*[types.InlineKeyboardButton(text=name,callback_data='cop'+str(dsa[i][0])) for name in [dsa[i][1]]]) 
            keyboard.add(*[types.InlineKeyboardButton(text=name,callback_data='addopr') for name in [lengstr(ll,19)]])                 
            msg = bot.send_message(m.chat.id, lengstr(ll,18)+' '+str(len(dsa)),reply_markup=keyboard)      
    if lengstr(ll,33)==m.text and m.chat.id==admin_id:
            dsa=all_oprosi()
            keyboard = types.InlineKeyboardMarkup(row_width=1)
            for i in range(0,len(dsa)):
                keyboard.add(*[types.InlineKeyboardButton(text=name,callback_data='stt'+str(dsa[i][0])) for name in [dsa[i][1]]])                 
            msg = bot.send_message(m.chat.id, lengstr(ll,35),reply_markup=keyboard) 	
    if lengstr(ll,16)==m.text and m.chat.id==admin_id:
            dsa=all_oprosi()
            keyboard = types.InlineKeyboardMarkup(row_width=1)
            for i in range(0,len(dsa)):
                keyboard.add(*[types.InlineKeyboardButton(text=name,callback_data='ras'+str(dsa[i][0])) for name in [dsa[i][1]]])                 
            msg = bot.send_message(m.chat.id, lengstr(ll,34),reply_markup=keyboard) 		
    if lengstr(ll,17)==m.text and m.chat.id==admin_id:
            dsa=all_groups()
            keyboard = types.InlineKeyboardMarkup(row_width=1)
            for i in range(0,len(dsa)):
                keyboard.add(*[types.InlineKeyboardButton(text=name,callback_data='grp'+str(dsa[i][0])) for name in [dsa[i][1][:29]]])   
            keyboard.add(*[types.InlineKeyboardButton(text=name,callback_data='addgroup') for name in [lengstr(ll,44)]])                 
            msg = bot.send_message(m.chat.id, lengstr(ll,43),reply_markup=keyboard) 	  
    if new_group==2 and m.chat.id==admin_id:
            ngroup.tema=m.text 
            new_group=3			
            msg = bot.send_message(m.chat.id, lengstr(ll,46))
            return			
    if new_group==3 and m.chat.id==admin_id:			
               new_group=4	
               ngroup.time=m.text			   			   
               msg = bot.send_message(m.chat.id, lengstr(ll,107))	
               return
    if new_group==4 and m.chat.id==admin_id:			
               new_group=5	
               ngroup.opisanie=m.text			   			   
               msg = bot.send_message(m.chat.id, lengstr(ll,108))	
               return
    if new_group==5 and m.chat.id==admin_id:			
               new_group=0	
               ngroup.opisanieua=m.text			   
               add_group(ngroup)			   
               msg = bot.send_message(m.chat.id, lengstr(ll,49))
    if change_group!='' and m.chat.id==admin_id:
        lisd=change_group[0]
        grid=int(change_group[1:])
        change_group=''
        if lisd=='z':
            if m.text.isdigit():
                ch_zaderjku(grid,m.text)
                msg = bot.send_message(m.chat.id, lengstr(ll,51))
            else:
                msg = bot.send_message(m.chat.id, lengstr(ll,48))  
        if lisd=='t':
                ch_tema(grid,m.text)
                msg = bot.send_message(m.chat.id, lengstr(ll,51))	
    print(m.text, admin_id, m.chat.id)
    if m.text==lengstr(ll,42) and m.chat.id==admin_id:
            keyboard = types.InlineKeyboardMarkup(row_width=1)  
            keyboard.add(*[types.InlineKeyboardButton(text=name,callback_data='opciitext') for name in [lengstr(ll,70)]])        
            keyboard.add(*[types.InlineKeyboardButton(text=name,callback_data='opciipass') for name in [lengstr(ll,71)]])  			
            msg = bot.send_message(m.chat.id, lengstr(ll,69),reply_markup=keyboard) 	  
    if what_change=='p' and m.chat.id==admin_id:
        change_pass(m.text)
        password=m.text	
        what_change=''		
        msg = bot.send_message(m.chat.id, lengstr(ll,73)) 
    if m.text==lengstr(ll,76) and m.chat.id==admin_id:
            dialog=zagruzit_dialogi()
            print(len(dialog))
            dialog.reverse()
            keyboard = types.InlineKeyboardMarkup(row_width=1)  
            for i in range(0,len(dialog)):
                if i==10:
                    break
                keyboard.add(types.InlineKeyboardButton(text=str(dialog[i][0])+' '+str(dialog[i][1]),callback_data='dialup'+str(dialog[i][0])+':'+str(dialog[i][1])))
            if len(dialog)>10:
                keyboard.add(types.InlineKeyboardButton(text='>>',callback_data='dialognext'+str(1)))                			
            msg = bot.send_message(m.chat.id, lengstr(ll,77),reply_markup=keyboard) 
    if m.text==lengstr(ll,86) and m.chat.id==admin_id:
        keyboard = types.InlineKeyboardMarkup(row_width=1)  
        opcii_smile=['','','','']
        print(opcii_menu)
        for i in range(0,4):
            print(opcii_menu[i])
            if opcii_menu[i]==1:
                   opcii_smile[i]='✅'
            else:
                   opcii_smile[i]='❌'
        print(opcii_smile)
        keyboard.add(*[types.InlineKeyboardButton(text=name+' '+opcii_smile[0],callback_data='opcii_mnu1') for name in [lengstr(ll,82)]]) 
        keyboard.add(*[types.InlineKeyboardButton(text=name+' '+opcii_smile[1],callback_data='opcii_mnu2') for name in [lengstr(ll,83)]])
        keyboard.add(*[types.InlineKeyboardButton(text=name+' '+opcii_smile[2],callback_data='opcii_mnu3') for name in [lengstr(ll,84)]])
        keyboard.add(*[types.InlineKeyboardButton(text=name+' '+opcii_smile[3],callback_data='opcii_mnu4') for name in [lengstr(ll,85)]])
        keyboard.add(types.InlineKeyboardButton(text='Оповестить всех пользователей',callback_data='all_users_say'))       
        msg = bot.send_message(m.chat.id, lengstr(ll,86),reply_markup=keyboard) 
    return

	
def del_poi(id,tema):	
    conn = sqlite3.connect('CONV.sqlite')
    cursor = conn.cursor()
    cursor.execute("DELETE FROM ZAI WHERE CL_ID=:id AND TEXT='' ",{"id": id})
    conn.commit()
    conn.close()	
	
def korotki(tema,ll):	
    conn = sqlite3.connect('CONV.sqlite')
    cursor = conn.cursor()
    cursor.execute("SELECT * FROM GROUPS WHERE TEMA =:id OR TIME =:id",{"id": tema})
    results = cursor.fetchall()
    conn.close() 
    if ll==1:
        return(results[0][4])
    else:
        return(results[0][5])

	
def all_users():
    conn = sqlite3.connect('BD.sqlite')
    cursor = conn.cursor()
    cursor.execute("SELECT * FROM USERS_CL")
    results = cursor.fetchall()
    conn.close()
    return(str(len(results)))	
	
def col_vo_zaprosov(tema):
    conn = sqlite3.connect('CONV.sqlite')
    cursor = conn.cursor()
    cursor.execute("SELECT * FROM ZAI WHERE TEMA=:id",{"id": tema})
    results = cursor.fetchall()
    conn.close()
    return(str(len(results)))    
    	
	
def find_leng(id):	
    try:
        conn = sqlite3.connect('RAZGOVORI.sqlite')
        cursor = conn.cursor()
        cursor.execute("SELECT * FROM LENGS WHERE USER_ID = :id",{"id": id})
        res = cursor.fetchall()   
        conn.close()
        return(res[0][1])
    except Exception:
        return(1)	
	
def user_leng(id,gg):
    conn = sqlite3.connect('RAZGOVORI.sqlite')
    cursor = conn.cursor()
    cursor.execute("insert into LENGS values (:a1,:a2) ", {"a1": id,"a2": gg})
    conn.commit()
    conn.close()     	


def zaprosi_obsh(id,zp_id):
    conn = sqlite3.connect('CONV.sqlite')
    cursor = conn.cursor()
    cursor.execute("SELECT * FROM ZAI WHERE CL_ID=:id",{"id": id})
    results = cursor.fetchall()
    conn.close()
    text=results[zp_id][3]	
    mid=pickle.loads(results[zp_id][2])
    m_id=mid[2]
    msg = bot.forward_message(o_chat,id,m_id)
	
def zaprosi_dunl(id,zp_id):		
    conn = sqlite3.connect('CONV.sqlite')
    cursor = conn.cursor()
    cursor.execute("SELECT * FROM ZAI WHERE CL_ID=:id",{"id": id})
    results = cursor.fetchall()
    conn.close()
    text=results[zp_id][3]	
    conn = sqlite3.connect('CONV.sqlite')
    cursor = conn.cursor()
    cursor.execute("UPDATE ZAI SET TEXT = :txt WHERE CL_ID=:id AND TEXT = :text", {"id": id,"text": text, "txt": ''})
    conn.commit()
    conn.close() 
    conn = sqlite3.connect('CONV.sqlite')
    cursor = conn.cursor()
    cursor.execute("DELETE FROM Thread WHERE CL_ID=:id AND TEXT = :text", {"id": id,"text": text})
    conn.commit()
    conn.close() 	
    conn = sqlite3.connect('CONV.sqlite')
    cursor = conn.cursor()
    cursor.execute("DELETE FROM ALL_COD WHERE COD = :cod", {"cod": results[zp_id][1]})
    conn.commit()
    conn.close() 
    mid=pickle.loads(results[zp_id][2])
    m_id=mid[2]
    ref=refund1()
    dobavit_poi1(id,text,ref,m_id)
    for i in range(0,len(mid[0])):
        try:
            msg = bot.delete_message(mid[0][i],mid[1][i])
        except Exception:
            aa=1
	
def clock2(id):
    conn = sqlite3.connect('BD.sqlite')
    cursor = conn.cursor()
    cursor.execute("SELECT * FROM USERS_CL")
    results = cursor.fetchall()
    conn.close() 
    print(len(results))	
    for i in range(0,len(results)):
        print(i)
        if i % 15==0:
            time.sleep(1)
        if results[i][0]==id:
            continue
        ss1=results[i][0]
        ll=find_leng(ss1)
        wg=wh_h(results[i][0])
        keyboard = types.ReplyKeyboardMarkup(resize_keyboard=True)
        if wg==0:
            keyboard.add(*[types.InlineKeyboardButton(text=name,callback_data=name) for name in [lengstr(ll,80)]])            
        keyboard.add(*[types.InlineKeyboardButton(text=name,callback_data=name) for name in [lengstr(ll,8),lengstr(ll,81)]])
        for i in range(0,4):
            if opcii_menu[i]==1:
                keyboard.add(*[types.InlineKeyboardButton(text=name,callback_data=name) for name in [lengstr(ll,82+i)]])
        if wg==1:
            keyboard.add(*[types.InlineKeyboardButton(text=name,callback_data=name) for name in [lengstr(ll,80)]])   
        print(i,ss1)			
        msg=bot.send_message(ss1,lengstr(ll,98),reply_markup=keyboard)	
    msg = bot.send_message(id,'Оповещение прошло успешно')	
	
	
	
	

def poisk_otz(user_p):	
    conn = sqlite3.connect('RAZGOVORI.sqlite')
    cursor = conn.cursor()
    cursor.execute("SELECT * FROM COMENT WHERE WHO = :id",{"id": user_p})
    res = cursor.fetchall()   
    conn.close()
    return(res)


	
def add_otz_bd(usrt,id,message_id):
    conn = sqlite3.connect('RAZGOVORI.sqlite')
    cursor = conn.cursor()
    cursor.execute("insert into COMENT values (:a1,:a2) ", {"a1": usrt,"a2": message_id})
    conn.commit()
    conn.close() 	

def add_ww(id):
    conn = sqlite3.connect('RAZGOVORI.sqlite')
    cursor = conn.cursor()
    cursor.execute("insert into WW values (:a1) ", {"a1": id})
    conn.commit()
    conn.close()  


def wh_h(id):	
    conn = sqlite3.connect('RAZGOVORI.sqlite')
    cursor = conn.cursor()
    cursor.execute("SELECT * FROM WW WHERE USER_ID = :id",{"id": id})
    res = cursor.fetchall()   
    conn.close()
    if len(res)>0:
        return(1)
    else:		
        return(0)

	
def zagruzit_dialogi():
    conn = sqlite3.connect('RAZGOVORI.sqlite')
    cursor = conn.cursor()
    cursor.execute("SELECT * FROM RAZGOVOR")
    res = cursor.fetchall()   
    conn.close()
    return(res)	

def    zapis_conv(id1, id2, text):
    conn = sqlite3.connect('RAZGOVORI.sqlite')
    cursor = conn.cursor()
    cursor.execute("SELECT * FROM RAZGOVOR")
    res = cursor.fetchall()   
    conn.close()
    for i in range(0,len(res)):
        if res[i][0]==id1 and res[i][1]==id2:
            conn = sqlite3.connect('RAZGOVORI.sqlite')
            cursor = conn.cursor()
            text=res[i][2]+'\n\n'+text
            cursor.execute("UPDATE RAZGOVOR SET TEXT = :time WHERE USER1 =:id AND USER2 = :ref ",{"id": id1,"time":text,"ref": id2})
            conn.commit()
            conn.close() 
            return			
        if res[i][1]==id1 and res[i][0]==id2:
            conn = sqlite3.connect('RAZGOVORI.sqlite')
            cursor = conn.cursor()
            text=res[i][2]+'\n\n'+text
            cursor.execute("UPDATE RAZGOVOR SET TEXT = :time WHERE USER1 =:ref AND USER2 = :id ",{"id": id1,"time":text,"ref": id2})
            conn.commit()
            conn.close() 
            return	
    conn = sqlite3.connect('RAZGOVORI.sqlite')
    cursor = conn.cursor()
    cursor.execute("insert into RAZGOVOR values (:a1,:a2,:a3) ", {"a1": id1,"a2": id2,"a3": text})
    conn.commit()
    conn.close()                     


	
def change_pass(text):
    conn = sqlite3.connect('CONV.sqlite')
    cursor = conn.cursor()
    cursor.execute("DELETE FROM PASS")
    conn.commit()
    conn.close() 
    conn = sqlite3.connect('CONV.sqlite')
    cursor = conn.cursor()
    cursor.execute("insert into PASS values (:a1) ", {"a1": text})
    conn.commit()
    conn.close()     	

	
def start_prov(id):
    results1=[]
    results2=[]
    conn = sqlite3.connect('BD.sqlite')
    cursor = conn.cursor()
    cursor.execute("SELECT * FROM USERS_CL")
    results = cursor.fetchall()   
    conn.close()
    for i in range(0,len(results)):
        results1.append(results[i][0])
    conn = sqlite3.connect('BD.sqlite')
    cursor = conn.cursor()
    cursor.execute("SELECT * FROM USERS_UR")
    results = cursor.fetchall()   
    conn.close()
    for i in range(0,len(results)):
        results2.append(results[i][0])	
    if id in results1:
       return(2)	
    if id in results2:
       return(1)	
    return(0) 
    
	
def chek_razgovor(id):
    conn = sqlite3.connect('CONV.sqlite')
    cursor = conn.cursor()
    cursor.execute("SELECT * FROM RAZ")
    results = cursor.fetchall()
    conn.close()
    for i in range(0,len(results)):
        if results[i][0]==id:
            conn = sqlite3.connect('CONV.sqlite')
            cursor = conn.cursor()
            cursor.execute("DELETE FROM RAZ WHERE FROM1 = :cod AND TO1=:da", {"cod": id,"da": results[i][1]})
            conn.commit()
            conn.close()            
            return(results[i][1])	
	
def from_us():	
    conn = sqlite3.connect('CONV.sqlite')
    cursor = conn.cursor()
    cursor.execute("SELECT * FROM RAZ")
    results = cursor.fetchall()
    conn.close()
    bs=[]
    for i in range(0,len(results)):
        bs.append(results[i][0])
    return(bs)   

   
def razgovor_go(from1,to1):
    conn = sqlite3.connect('CONV.sqlite')
    cursor = conn.cursor()
    cursor.execute("insert into RAZ values (:a1, :a2) ", {"a1": from1,"a2": to1})
    conn.commit()
    conn.close()     

def chek_kod(kod):
    conn = sqlite3.connect('CONV.sqlite')
    cursor = conn.cursor()
    cursor.execute("SELECT CL_ID FROM ZAI WHERE CODE=:id",{"id": kod})
    results = cursor.fetchall()
    conn.close()
    if len(results)==0:
       return(0)
    else:
       return(results[0][0])	


	
	
def zaprosi_delete(id,zp_id):		
    conn = sqlite3.connect('CONV.sqlite')
    cursor = conn.cursor()
    cursor.execute("SELECT * FROM ZAI WHERE CL_ID=:id",{"id": id})
    results = cursor.fetchall()
    conn.close()
    text=results[zp_id][3]	
    conn = sqlite3.connect('CONV.sqlite')
    cursor = conn.cursor()
    cursor.execute("DELETE FROM ZAI WHERE CL_ID=:id AND TEXT = :text", {"id": id,"text": text})
    conn.commit()
    conn.close() 
    conn = sqlite3.connect('CONV.sqlite')
    cursor = conn.cursor()
    cursor.execute("DELETE FROM Thread WHERE CL_ID=:id AND TEXT = :text", {"id": id,"text": text})
    conn.commit()
    conn.close() 	
    conn = sqlite3.connect('CONV.sqlite')
    cursor = conn.cursor()
    cursor.execute("DELETE FROM ALL_COD WHERE COD = :cod", {"cod": results[zp_id][1]})
    conn.commit()
    conn.close() 
    mid=pickle.loads(results[zp_id][2])
    for i in range(0,len(mid[0])):
        try:
            msg = bot.delete_message(mid[0][i],mid[1][i])
        except Exception:
            aa=1
	
def zaprosi_cl(id):	
    conn = sqlite3.connect('CONV.sqlite')
    cursor = conn.cursor()
    cursor.execute("SELECT * FROM ZAI WHERE CL_ID=:id",{"id": id})
    results = cursor.fetchall()
    conn.close()
    return(results)	
	
def dobavit_poi1(id,text,ref,m_id):
    conn = sqlite3.connect('CONV.sqlite')
    cursor = conn.cursor()
    cursor.execute("SELECT TEMA FROM ZAI WHERE CL_ID=:id AND TEXT='' ",{"id": id})
    results = cursor.fetchall()
    conn.close()
    tema=results[0][0]
    conn = sqlite3.connect('CONV.sqlite')
    cursor = conn.cursor()
    cursor.execute("SELECT * FROM GROUPS WHERE TEMA =:id OR TIME =:id",{"id": tema})
    results = cursor.fetchall()
    conn.close() 
    send_array=[]
    message_array=[]
    for i in range(0,len(results)):
            send_array.append(results[i][0])
            msg = bot.forward_message(chat_id=results[i][0],from_chat_id=id,message_id=m_id)	
            message_array.append(msg.message_id)	          			
    print(send_array)
    print(message_array)
    mid=[]
    mid.append(send_array)
    mid.append(message_array)
    mid.append(m_id)
    mid = pickle.dumps(mid)
    conn = sqlite3.connect('CONV.sqlite')
    cursor = conn.cursor()
    cursor.execute("UPDATE ZAI SET TEXT = :time, CODE = :ref, MESSAGE_ID = :mid WHERE CL_ID =:id AND TEXT = '' ",{"id": id,"time": text,"ref": ref,"mid": mid})
    conn.commit()
    conn.close()  	
	
def refund1():
    conn = sqlite3.connect('CONV.sqlite')
    cursor = conn.cursor()
    cursor.execute("SELECT COD FROM ALL_COD ")
    results = cursor.fetchall()
    conn.close()
    reflist=results
    gh=1
    while gh==1:
            chars=string.ascii_uppercase + string.digits
            zz=''.join(random.choice(chars) for _ in range(8))   
            if zz not in reflist:
                conn = sqlite3.connect('CONV.sqlite')
                cursor = conn.cursor()
                cursor.execute("insert into ALL_COD values (:a1) ", {"a1": zz})
                conn.commit()
                conn.close()                
                return(zz)	
	            
	
def   dobavit_poi(id,tema):
    conn = sqlite3.connect('CONV.sqlite')
    cursor = conn.cursor()
    cursor.execute("insert into ZAI values (:a1, :a2, :a3, :a4, :a5) ", {"a1": id,"a2": '',"a3": '',"a4": '',"a5": tema})
    conn.commit()
    conn.close() 
	
	
def groups_tems(ln):	
    if ln==1:
     conn = sqlite3.connect('CONV.sqlite')
     cursor = conn.cursor()
     cursor.execute("SELECT TEMA FROM GROUPS")
     results = cursor.fetchall()
     conn.close()
     zz=list(set(results))
     return(zz)  
    else:	 
     conn = sqlite3.connect('CONV.sqlite')
     cursor = conn.cursor()
     cursor.execute("SELECT TIME FROM GROUPS")
     results = cursor.fetchall()
     conn.close()
     zz=list(set(results))
     return(zz)  


	
def delete_group(id):
    conn = sqlite3.connect('CONV.sqlite')
    cursor = conn.cursor()
    cursor.execute("DELETE FROM GROUPS WHERE GROUP_ID=:id", {"id": id})
    conn.commit()
    conn.close()  	
	
	
def ch_tema(id,time):
    conn = sqlite3.connect('CONV.sqlite')
    cursor = conn.cursor()
    cursor.execute("UPDATE GROUPS SET TEMA = :time WHERE GROUP_ID =:id",{"id": id,"time": time})
    conn.commit()
    conn.close()  	
	
def ch_zaderjku(id,time):
    conn = sqlite3.connect('CONV.sqlite')
    cursor = conn.cursor()
    cursor.execute("UPDATE GROUPS SET TIME = :time WHERE GROUP_ID =:id",{"id": id,"time": time})
    conn.commit()
    conn.close()  	
	
def chek_group(id):
    conn = sqlite3.connect('CONV.sqlite')
    cursor = conn.cursor()
    cursor.execute("SELECT * FROM GROUPS WHERE GROUP_ID =:id",{"id": id})
    results = cursor.fetchall()
    conn.close()
    zz=results[0]
    return(zz)     	
	
def add_group(grou):
    conn = sqlite3.connect('CONV.sqlite')
    cursor = conn.cursor()
    cursor.execute("insert into GROUPS values (:a1, :a2, :a3, :a4, :a5, :a6) ", {"a1": grou.id,"a2": grou.name,"a3": grou.tema,"a4": grou.time,"a5": grou.opisanie,"a6": grou.opisanieua})
    conn.commit()
    conn.close()    	
	

def all_oprosi():
    conn = sqlite3.connect('BD.sqlite')
    cursor = conn.cursor()
    cursor.execute("SELECT * FROM OPROSI")
    results = cursor.fetchall()
    conn.close()
    print(results)
    return(results)   

def add_urist(user_id):
    conn = sqlite3.connect('BD.sqlite')
    cursor = conn.cursor()
    cursor.execute("insert into USERS_UR values (:post_id) ", {"post_id": user_id})
    conn.commit()
    conn.close()
	
def add_opr_in_db(name,id):
    conn = sqlite3.connect('BD.sqlite')
    cursor = conn.cursor()
    cursor.execute("insert into OPROSI values (:post_id, :name) ", {"post_id": id,"name": name})
    conn.commit()
    conn.close()	

	
def add_vopros(vopros,id):
    conn = sqlite3.connect('BD.sqlite')
    cursor = conn.cursor()
    cursor.execute("insert into VOPROSI values (:post_id, :name, :otvet) ", {"post_id": id,"name": vopros,"otvet": ''})
    conn.commit()
    conn.close()
	
def add_otvet(vopros,id):
    conn = sqlite3.connect('BD.sqlite')
    cursor = conn.cursor()
    cursor.execute("UPDATE VOPROSI SET OTVETI = :name WHERE OPROS_ID = :id ", {"name": vopros,"id": id})
    conn.commit()
    conn.close()	

def  add_vopros_opt(vopros,otvet,id) :
    conn = sqlite3.connect('BD.sqlite')
    cursor = conn.cursor()
    cursor.execute("insert into VOPROSI values (:post_id, :name, :otvet) ", {"post_id": id,"name": vopros,"otvet": otvet})
    conn.commit()
    conn.close()

	
def add_client(user_id):
    conn = sqlite3.connect('BD.sqlite')
    cursor = conn.cursor()
    cursor.execute("SELECT * FROM USERS_CL")
    results = cursor.fetchall()
    conn.close()
    for i in range(0,len(results)):
        if results[i][0]==user_id:
            return('hh')
    conn = sqlite3.connect('BD.sqlite')
    cursor = conn.cursor()
    cursor.execute("insert into USERS_CL values (:post_id) ", {"post_id": user_id})
    conn.commit()
    conn.close()

	
def chek_opros_full(id):
    res='Название опроса: '
    conn = sqlite3.connect('BD.sqlite')
    cursor = conn.cursor()
    cursor.execute("SELECT NAME FROM OPROSI WHERE ID=:id", {"id": id})
    results = cursor.fetchall()
    conn.close()  
    print(results)
    res+=results[0][0]+'\n'
    conn = sqlite3.connect('BD.sqlite')
    cursor = conn.cursor()
    cursor.execute("SELECT TEXT FROM VOPROSI WHERE OPROS_ID=:id", {"id": id})
    results1 = cursor.fetchall()
    cursor.execute("SELECT OTVETI FROM VOPROSI WHERE OPROS_ID=:id", {"id": id})
    results2 = cursor.fetchall()
    conn.close()    
    for i in range(0,len(results1)):
        res+='\nВопрос №'+str(i+1)+' : '+results1[i][0]+'\nВарианты ответов:\n'
        otveti=pickle.loads(results2[i][0])
        for k in range(0,len(otveti)):
                res+=otveti[k]+'\n'
    return(res)			
	
def delete_opros(id):
    conn = sqlite3.connect('BD.sqlite')
    cursor = conn.cursor()
    cursor.execute("DELETE FROM OPROSI WHERE ID=:id", {"id": id})
    conn.commit()
    conn.close()  
    conn = sqlite3.connect('BD.sqlite')
    cursor = conn.cursor()
    cursor.execute("DELETE FROM VOPROSI WHERE OPROS_ID=:id", {"id": id})
    conn.commit()
    conn.close()    
    return()	
	
	
	
	
def spisok_voprosov(id):
    conn = sqlite3.connect('BD.sqlite')
    cursor = conn.cursor()
    cursor.execute("SELECT TEXT FROM VOPROSI WHERE OPROS_ID = :id", {"id": id})
    results = cursor.fetchall()
    conn.close()
    return(results)    

def   delete_vopros(num,opr_id):
    conn = sqlite3.connect('BD.sqlite')
    cursor = conn.cursor()
    cursor.execute("SELECT TEXT FROM VOPROSI WHERE OPROS_ID = :id", {"id": opr_id})
    results = cursor.fetchall()
    conn.close()
    tt=results[int(num)][0]
    conn = sqlite3.connect('BD.sqlite')
    cursor = conn.cursor()
    cursor.execute("DELETE FROM VOPROSI WHERE TEXT=:id", {"id": tt})
    conn.commit()
    conn.close()         





def  rassilka_oprosa(opr_id):
    conn = sqlite3.connect('BD.sqlite')
    cursor = conn.cursor()
    cursor.execute("DELETE FROM OTVETI WHERE OPROS_ID=:id", {"id": opr_id})
    conn.commit()
    conn.close()     
    conn = sqlite3.connect('BD.sqlite')
    cursor = conn.cursor()
    cursor.execute("SELECT TEXT FROM VOPROSI WHERE OPROS_ID = :id", {"id": opr_id})
    results = cursor.fetchall()
    conn.close()
    conn = sqlite3.connect('BD.sqlite')
    cursor = conn.cursor()
    aa=[]
    aa=pickle.dumps(aa)
    for i in range(0,len(results)):
        cursor.execute("insert into OTVETI values (:post_id, :name, :otvet, :otvet1) ", {"post_id": opr_id,"name": i,"otvet": aa, "otvet1": aa})
        conn.commit()
    conn.close()
    t = threading.Thread(target=clock, args=(opr_id,))
    t.start()




	
def clock(opr_id):
    conn = sqlite3.connect('BD.sqlite')
    cursor = conn.cursor()
    cursor.execute("SELECT * FROM USERS_CL")
    results = cursor.fetchall()
    conn.close()    
    conn = sqlite3.connect('BD.sqlite')
    cursor = conn.cursor()
    cursor.execute("SELECT * FROM VOPROSI WHERE OPROS_ID = :id", {"id": opr_id})
    results1 = cursor.fetchall()
    conn.close()    
    for i in range(0,len(results)):
        if i % 15==0:
            time.sleep(1)
        ss1=results[i][0]
        ss2=results1[0][2]
        keyboard = types.InlineKeyboardMarkup(row_width=1)
        otveti=pickle.loads(ss2)
        ll=find_leng(ss1)		
        keyboard.add(*[types.InlineKeyboardButton(text=name,callback_data='otv'+str(otveti.index(name))+':'+str(opr_id)+'†'+str(0)) for name in otveti])            		
        msg=bot.send_message(ss1,lengstr(ll,30)+'\n\n'+results1[0][1],reply_markup=keyboard)



def add_otvet_in(user_id,opr_id,vopros_id,otvet_id):
    conn = sqlite3.connect('BD.sqlite')
    cursor = conn.cursor()
    cursor.execute("SELECT * FROM OTVETI WHERE OPROS_ID = :opr_id and VOPROS_ID = :vopros_id", {"opr_id": opr_id,"vopros_id": vopros_id})
    results = cursor.fetchall()
    conn.close()
    users=pickle.loads(results[0][2])
    print(users)
    otveti=pickle.loads(results[0][3])	
    print(otveti)
    users.append(user_id)
    otveti.append(otvet_id)
    users=pickle.dumps(users)
    otveti=pickle.dumps(otveti)
    conn = sqlite3.connect('BD.sqlite')
    cursor = conn.cursor()
    cursor.execute("UPDATE OTVETI SET USERS=:users WHERE OPROS_ID = :opr_id and VOPROS_ID = :vopros_id ", {"users": users,"opr_id": opr_id,"vopros_id": vopros_id})
    cursor.execute("UPDATE OTVETI SET ANSW=:otveti WHERE OPROS_ID = :opr_id and VOPROS_ID = :vopros_id ", {"otveti": otveti,"opr_id": opr_id,"vopros_id": vopros_id})
    conn.commit()
    conn.close()    

	
	
def   chek_next_question(user_id,bib,opr_id,vopros_id):    
    conn = sqlite3.connect('BD.sqlite')
    cursor = conn.cursor()
    vopros_id=int(vopros_id)
    cursor.execute("SELECT * FROM VOPROSI WHERE OPROS_ID = :id", {"id": opr_id})
    results1 = cursor.fetchall()
    conn.close()  
    vopros_id+=1
    if len(results1)==vopros_id:
           msg = bot.edit_message_text(chat_id=user_id, message_id=bib, text=lengstr(ll,31)) 	
           return
    keyboard = types.InlineKeyboardMarkup(row_width=1)
    otveti=pickle.loads(results1[vopros_id][2])  
    keyboard.add(*[types.InlineKeyboardButton(text=name,callback_data='otv'+str(otveti.index(name))+':'+str(opr_id)+'†'+str(vopros_id)) for name in otveti])            		
    msg=bot.edit_message_text(chat_id=user_id, message_id=bib,text=results1[vopros_id][1],reply_markup=keyboard)
	
	
	
def stat_opros(opr_id):
    conn = sqlite3.connect('BD.sqlite')
    cursor = conn.cursor()
    cursor.execute("SELECT * FROM OTVETI WHERE OPROS_ID = :opr_id", {"opr_id": opr_id})
    results1 = cursor.fetchall()
    conn.close()
    conn = sqlite3.connect('BD.sqlite')
    cursor = conn.cursor()
    cursor.execute("SELECT * FROM VOPROSI WHERE OPROS_ID = :id", {"id": opr_id})
    results2 = cursor.fetchall()
    conn.close()
    res=''
    if len(results1)==0:
        res=lengstr(ll,36)
        return(res)
    for i in range(0,len(results2)):
            col_vo=len(pickle.loads(results1[i][3]))
            otv=pickle.loads(results1[i][3])
            res+='Вопрос №'+str(i+1)+'\n'+results2[i][1]+'\nВсего ответило: '+str(col_vo)+'\n\n'
            col_ot=len(pickle.loads(results2[i][2]))
            var_ot=pickle.loads(results2[i][2])
            for k in range(0,col_ot):
                sum=0
                for j in range(0,col_vo):
                    if str(otv[j])==str(k):
                        sum+=1
                res+=var_ot[k]+' - '+str(sum)+'\n'
            res+='\n'
    return(res)
  
def chek_op(opr_id):
    conn = sqlite3.connect('BD.sqlite')
    cursor = conn.cursor()
    cursor.execute("SELECT * FROM OTVETI WHERE OPROS_ID = :opr_id", {"opr_id": opr_id})
    results1 = cursor.fetchall()
    conn.close()
    if len(results1)==0:
        res=0
    else:
        res=1
    return(res)
  
  
  
  
def create_filter(opr_id):
    global filter
    filter=[]
    conn = sqlite3.connect('BD.sqlite')
    cursor = conn.cursor()
    cursor.execute("SELECT TEXT FROM VOPROSI WHERE OPROS_ID=:id", {"id": opr_id})
    results1 = cursor.fetchall()
    cursor.execute("SELECT OTVETI FROM VOPROSI WHERE OPROS_ID=:id", {"id": opr_id})
    results2 = cursor.fetchall()
    conn.close()    
    for i in range(0,len(results1)):
        filter.append([])
        otveti=pickle.loads(results2[i][0])
        for k in range(0,len(otveti)):
                filter[-1].append(1)
    print(filter)	    
	
	
	
	
	
def keyboard_for_filter(opr_id):
    global filter	
    res=''

    fulter_emoji=copy.deepcopy(filter)
    for i in range(0,len(filter)):
        for k in range(0,len(filter[i])):
            if filter[i][k]==1:
                fulter_emoji[i][k]='✅'
            else:
                fulter_emoji[i][k]='❌'

    keyboard = types.InlineKeyboardMarkup(row_width=2)
    conn = sqlite3.connect('BD.sqlite')
    cursor = conn.cursor()
    cursor.execute("SELECT TEXT FROM VOPROSI WHERE OPROS_ID=:id", {"id": opr_id})
    results1 = cursor.fetchall()
    cursor.execute("SELECT OTVETI FROM VOPROSI WHERE OPROS_ID=:id", {"id": opr_id})
    results2 = cursor.fetchall()
    conn.close() 
    ll=find_leng(admin_id)
    for i in range(0,len(results1)):
        keyboard.add(types.InlineKeyboardButton(text='Вопрос №'+str(i+1),callback_data='„'))
        otveti=pickle.loads(results2[i][0])
        keyboard.add(*[types.InlineKeyboardButton(text=fulter_emoji[i][otveti.index(name)]+' '+name,callback_data='ifm'+str(i)+':'+str(otveti.index(name))+'†'+str(opr_id)) for name in otveti])
    keyboard.add(types.InlineKeyboardButton(text=lengstr(ll,38),callback_data='txtras'))
    return(keyboard)	
	
	
def  rassilka_po_filtru(txt,opr_id,txtua):
    global filter	
    conn = sqlite3.connect('BD.sqlite')
    cursor = conn.cursor()
    cursor.execute("SELECT * FROM OTVETI WHERE OPROS_ID = :opr_id", {"opr_id": opr_id})
    results = cursor.fetchall()
    conn.close()
    users_big=[]
    for i in range(0,len(filter)):
        spisokall=pickle.loads(results[i][2])
        for i in spisokall:
            users_big.append(i)
    users_big=list(set(users_big))
    for i in range(0,len(filter)):
        users_small=[]
        spisokall=pickle.loads(results[i][2])
        answersall=pickle.loads(results[i][3])
        for k in range(0,len(filter[i])):
            if filter[i][k]==0:
                continue
            else:
                for j in range(0,len(answersall)):
               	    if str(answersall[j])==str(k):
                        users_small.append(spisokall[j])	
        print(users_small)
        for m in range(0,len(users_big)):
            try:
             if users_big[m] in users_small:
                continue
             else:
                del users_big[m]
            except Exception:
             continue			
    print(users_big)
    t = threading.Thread(target=clock1, args=(txt,users_big,txtua))
    t.start()
 


def all_groups():
    conn = sqlite3.connect('CONV.sqlite')
    cursor = conn.cursor()
    cursor.execute("SELECT * FROM GROUPS")
    results = cursor.fetchall()
    conn.close() 
    return(results)    
      	
def clock1(txt,arr,txtua):
    for i in range(0,len(arr)):
        if i % 15==0:
            time.sleep(1)  
        ll=find_leng(arr[i])
        if ll==1:		
            msg=bot.send_message(arr[i],txt)
        if ll==2:		
            msg=bot.send_message(arr[i],txtua)
	
	
def forserer():
    conn = sqlite3.connect('CONV.sqlite')
    cursor = conn.cursor()
    array=[]
    now_time=int(time.time())
    cursor.execute("SELECT * FROM Thread WHERE TIME <= :time",{"time":now_time})
    results = cursor.fetchall()	
    print(results)
    cursor.execute("DELETE FROM Thread WHERE TIME <= :time",{"time":now_time})
    conn.commit()    
    conn.close()
    for i in range(0,len(results)):
            client_id=results[i][2]
            grid=results[i][3]
            msg = bot.send_message(results[i][3],'Новый клиент\nОписание:\n'+results[i][0]+'\n\nКод для связи с клиентом: <a href="https://t.me/Fine2113_bot?start='+results[i][4]+'">'+results[i][4]+'</a>',parse_mode='HTML')	
            msid=msg.message_id	 
            conn = sqlite3.connect('CONV.sqlite')
            cursor = conn.cursor()
            cursor.execute("SELECT * FROM ZAI WHERE CL_ID=:id AND TEXT=:text ",{"id": client_id,"text": results[i][0]})
            results1 = cursor.fetchall()
            conn.close()
            mid=results1[0][2]		
            mid=pickle.loads(mid)
            mid[0].append(grid)
            mid[1].append(msid)
            print(mid)
            mid = pickle.dumps(mid)
            conn = sqlite3.connect('CONV.sqlite')
            cursor = conn.cursor()
            cursor.execute("UPDATE ZAI SET MESSAGE_ID = :mid WHERE CL_ID =:id AND TEXT = :text ",{"id": client_id,"mid": mid,"text": results[i][0],"mid": mid})
            conn.commit()
            conn.close()            			

			
schedule.every(1).minutes.do(forserer)
def lal():
    while 1:
        schedule.run_pending()
        time.sleep(1)
_thread.start_new_thread(lal,())	
	






@bot.message_handler(content_types=['document'])
def photoget(message):
    global wb, wb1, sheet, sheet1, what_change, admin_id
    if what_change=='t'	and message.chat.id==admin_id:
        what_change=''
        ll=find_leng(admin_id)
        fileid=(message.document.file_id)
        bb=bot.get_file(fileid)
        bb=bb.file_path
        logo = urllib.request.urlopen("https://api.telegram.org/file/bot"+TOKEN+"/"+bb).read()
        f = open('leng.xlsx', "wb")
        f.write(logo)
        f.close()
        wb = openpyxl.load_workbook(filename = 'leng.xlsx')
        sheet = wb['test']
        keyboard = types.ReplyKeyboardMarkup(row_width=2)
        keyboard.add(*[types.InlineKeyboardButton(text=name,callback_data=name) for name in [lengstr(ll,15),lengstr(ll,16),lengstr(ll,17),lengstr(ll,33),lengstr(ll,42),lengstr(ll,86)]])
        msg = bot.send_message(message.chat.id, lengstr(ll,73),reply_markup=keyboard) 
	
	
bot.remove_webhook()

bot.set_webhook(url=WEBHOOK_URL_BASE + WEBHOOK_URL_PATH,
                certificate=open(WEBHOOK_SSL_CERT, 'r'))

cherrypy.config.update({
    'server.socket_host': WEBHOOK_LISTEN,
    'server.socket_port': WEBHOOK_PORT,
    'server.ssl_module': 'builtin',
    'server.ssl_certificate': WEBHOOK_SSL_CERT,
    'server.ssl_private_key': WEBHOOK_SSL_PRIV
})

cherrypy.quickstart(WebhookServer(), WEBHOOK_URL_PATH, {'/': {}})