import openpyxl
import time
import telebot
from telebot import types
import _pickle as pickle
import requests
import json
import cherrypy
TOKEN = '527514328:AAHpk306W5jtpeGES4Gn6Vy2kt94oMlgErg'
WEBHOOK_HOST = '95.46.98.126'
WEBHOOK_PORT = 8443  # 443, 80, 88 или 8443 (порт должен быть открыт!)
WEBHOOK_LISTEN = '0.0.0.0'  # На некоторых серверах придется указывать такой же IP, что и выше

WEBHOOK_SSL_CERT = './webhook_cert.pem'  # Путь к сертификату
WEBHOOK_SSL_PRIV = './webhook_pkey.pem'  # Путь к приватному ключу

WEBHOOK_URL_BASE = "https://%s:%s" % (WEBHOOK_HOST, WEBHOOK_PORT)
WEBHOOK_URL_PATH = "/%s/" % (TOKEN)
bot = telebot.TeleBot(TOKEN)
wb = openpyxl.load_workbook(filename = 'leng.xlsx')
sheet = wb['test']
val = sheet.cell(row=1, column=1)
print(val.value)
def  lengstr(leng,strn):
    bb=sheet.cell(row=strn, column=leng).value
    return str(bb)

class WebhookServer(object):
    @cherrypy.expose
    def index(self):
        if 'content-length' in cherrypy.request.headers and \
                        'content-type' in cherrypy.request.headers and \
                        cherrypy.request.headers['content-type'] == 'application/json':
            length = int(cherrypy.request.headers['content-length'])
            json_string = cherrypy.request.body.read(length).decode("utf-8")
            update = telebot.types.Update.de_json(json_string)
            bot.process_new_updates([update])
            return ''
        else:
            raise cherrypy.HTTPError(403)
class userobj():
    id=0
    leng=0
    sleng=1
    erp=0
    erpm=0
    passw=0
    erpkod=0
    session=''
    login=0
    email=''
    change=0
    addadd=0
global bdpol	
bdpol=[]
def nomer(b):
    global bdpol
    z=0
    k=-1
    for i in bdpol:
        k+=1
        if bdpol[k].id==b:
            z=k
            return(z)
# Handle '/start' and '/help'
@bot.message_handler(commands=['help', 'start'])
def start(message):
    global bdpol
    hesh=message.chat.id
    z=0
    k=-1
    for i in bdpol:
        k+=1
        if bdpol[k].id==hesh:
            z=1
            return
    if z==0:
        bdpol.append(userobj())
        print(bdpol[-1].id)
        bdpol[-1].id=hesh
        bdpol[-1].start=1
    keyboard = types.InlineKeyboardMarkup(row_width=1)
    for i in range(1,20):
       if lengstr(i,1) !='None':
          keyboard.add(*[types.InlineKeyboardButton(text=name,callback_data=str(i)) for name in [lengstr(i,1)]])
    msg = bot.send_message(message.chat.id, 'Select the language',
    reply_markup=keyboard)


@bot.message_handler(content_types=["text"])
def repeat_all_messages(message): 
    name(message)	
	
	
	
@bot.callback_query_handler(func=lambda c:True)
def inline(c):
    global bdpol
    k=nomer(c.message.chat.id)
    ll=bdpol[k].leng
    if bdpol[k].sleng ==1: 
        bdpol[k].leng =int(c.data)
        bdpol[k].sleng =0
        ll=bdpol[k].leng
        keyboard = types.InlineKeyboardMarkup(row_width=1)
        keyboard.add(*[types.InlineKeyboardButton(text=name,callback_data='reg') for name in [lengstr(ll,2)]])
        keyboard.add(*[types.InlineKeyboardButton(text=name,callback_data='login') for name in [lengstr(ll,61)]])
        msg = bot.send_message(c.message.chat.id,lengstr(ll,10) ,reply_markup=keyboard)		
    if c.data=='login':
        bdpol[k].login=1
        msg = bot.send_message(c.message.chat.id,lengstr(ll,11) )	 
    ############################# opcii		
    if c.data=='optld':
        keyboard = types.InlineKeyboardMarkup(row_width=1)
        keyboard.add(*[types.InlineKeyboardButton(text=name,callback_data='optldpass') for name in [lengstr(ll,59)]])
        keyboard.add(*[types.InlineKeyboardButton(text=name,callback_data='optldinfo') for name in [lengstr(ll,60)]])
        msg = bot.send_message(c.message.chat.id,lengstr(ll,9)+': '+lengstr(ll,55),reply_markup=keyboard)
    if c.data=='optcard':
        keyboard = types.InlineKeyboardMarkup(row_width=1)
        keyboard.add(*[types.InlineKeyboardButton(text=name,callback_data='optcardadd') for name in [lengstr(ll,18)]])
        keyboard.add(*[types.InlineKeyboardButton(text=name,callback_data='optcarddel') for name in [lengstr(ll,56)]])
        keyboard.add(*[types.InlineKeyboardButton(text=name,callback_data='optcardch') for name in [lengstr(ll,58)]])
        msg = bot.send_message(c.message.chat.id,lengstr(ll,9)+': '+lengstr(ll,54),reply_markup=keyboard)	
    if c.data=='optkosh':
        keyboard = types.InlineKeyboardMarkup(row_width=1)
        keyboard.add(*[types.InlineKeyboardButton(text=name,callback_data='optkoshadd') for name in [lengstr(ll,18)]])
        keyboard.add(*[types.InlineKeyboardButton(text=name,callback_data='optkoshdel') for name in [lengstr(ll,56)]])
        keyboard.add(*[types.InlineKeyboardButton(text=name,callback_data='optkoshch') for name in [lengstr(ll,58)]])
        msg = bot.send_message(c.message.chat.id,lengstr(ll,9)+': '+lengstr(ll,57),reply_markup=keyboard)
####################################### obrabotka vseh opci 
####################################### Настройки пользователя которых нет
    if c.data=='optldpass':
        msg = bot.send_message(c.message.chat.id,'Простите, этого метода я так и не дождался.')   
    if c.data=='optldinfo':
        msg = bot.send_message(c.message.chat.id,'Простите этот метод есть, но работает он не правильно. Может я подожду месяцев 6 и все будет готово.')  
    if c.data=='optcarddel':
        msg = bot.send_message(c.message.chat.id,'Простите, этого метода я так и не дождался.')  
    if c.data=='optcardch':
        msg = bot.send_message(c.message.chat.id,'Простите, этого метода я так и не дождался.')  
    if c.data=='optkoshdel':
        msg = bot.send_message(c.message.chat.id,'Простите, этого метода я так и не дождался.')  
    if c.data=='optcoshch':
        msg = bot.send_message(c.message.chat.id,'Простите, этого метода я так и не дождался.')  
####################################### Vnesenie novogo scheta
    if c.data=='optcardadd': 
        bdpol[k].change='card'
        payload = {'token': 'iQPU9x7HuK3wTPGFeZeTFRqlX8JCaMiqY2VCNoUdaf0ZtPBM1u'}
        r = requests.get('http://gate.paytochina.ru/rest/help/currencyList', params=payload)
        r=r.json()
        print(r)
        i=0
        dictval=[]
        while True:
            try:
                dictval.append(r['data'][i]['currency'])			
                i+=1
            except Exception:
                break
        dictval=list(set(dictval))
        dictval.remove('')
        print(dictval)
        dictval.sort()
        keyboard = types.InlineKeyboardMarkup(row_width=3)
        keyboard.add(*[types.InlineKeyboardButton(text=name,callback_data=name) for name in dictval])
        msg = bot.send_message(c.message.chat.id, lengstr(ll,73),
        reply_markup=keyboard)		
        bdpol[k].addadd=1	
        return()		
    if c.data=='optkoshadd': 
        bdpol[k].change='account'
        payload = {'token': 'iQPU9x7HuK3wTPGFeZeTFRqlX8JCaMiqY2VCNoUdaf0ZtPBM1u'}
        r = requests.get('http://gate.paytochina.ru/rest/help/currencyList', params=payload)
        r=r.json()
        i=0
        dictval=[]
        while True:
            try:
                dictval.append(r['data'][i]['currency'])			
                i+=1
            except Exception:
                break
        dictval=list(set(dictval))
        dictval.remove('')
        print(dictval)
        dictval.sort()
        keyboard = types.InlineKeyboardMarkup(row_width=3)
        keyboard.add(*[types.InlineKeyboardButton(text=name,callback_data=name) for name in dictval])
        msg = bot.send_message(c.message.chat.id, lengstr(ll,73),
        reply_markup=keyboard)		
        bdpol[k].addadd=1
        return()
    if bdpol[k].addadd==1:  
        bdpol[k].addadd=2
        bdpol[k].newtype=c.data
        payload = {'token': 'iQPU9x7HuK3wTPGFeZeTFRqlX8JCaMiqY2VCNoUdaf0ZtPBM1u'}
        r = requests.get('http://gate.paytochina.ru/rest/help/currencyList', params=payload)
        r=r.json()
        i=0
        dictval=[]
        while True:
            try:
                if r['data'][i]['currency']==bdpol[k].newtype:
                    dictval.append(r['data'][i]['title_en'])			
                i+=1
            except Exception:
                break
        dictval=list(set(dictval))
        print(dictval)
        dictval.sort()
        keyboard = types.InlineKeyboardMarkup(row_width=3)
        keyboard.add(*[types.InlineKeyboardButton(text=name,callback_data=name) for name in dictval])
        msg = bot.send_message(c.message.chat.id, lengstr(ll,75),
        reply_markup=keyboard)		
        return()   
    if bdpol[k].addadd==2:  
        bdpol[k].addadd=3
        payload = {'token': 'iQPU9x7HuK3wTPGFeZeTFRqlX8JCaMiqY2VCNoUdaf0ZtPBM1u'}
        r = requests.get('http://gate.paytochina.ru/rest/help/currencyList', params=payload)
        r=r.json()
        i=-1
        print(c.data)
        print(bdpol[k].newtype)
        dictval=''
        while True:
                i+=1
                if r['data'][i]['title_en']==c.data and r['data'][i]['currency']==bdpol[k].newtype:			
                    dictval=r['data'][i]['id']
                    print(dictval)
                    break
        bdpol[k].newtype=dictval          
        if bdpol[k].change == 'account':
            msg = bot.send_message(c.message.chat.id, lengstr(ll,76))
            return()
        if bdpol[k].change == 'card':
            msg = bot.send_message(c.message.chat.id, lengstr(ll,77))	
            return() 		
	############################		
    if c.data=='instr':
        msg = bot.send_message(c.message.chat.id,lengstr(ll,53))
    if c.data=='reg':	
        msg = bot.send_message(c.message.chat.id,lengstr(ll,11))
        bdpol[k].erpm=1	
    if c.data=='limits':
        payload = {'token': 'iQPU9x7HuK3wTPGFeZeTFRqlX8JCaMiqY2VCNoUdaf0ZtPBM1u'}
        r = requests.get('http://gate.paytochina.ru/rest/help/currencyList', params=payload)
        r=r.json()
        i=0
        dictval=[]
        while True:
            try:
                dictval.append(r['data'][i]['currency'])			
                i+=1
            except Exception:
                break
        dictval=list(set(dictval))
        dictval.remove('')
        print(dictval)
        dictval.sort()
        keyboard = types.InlineKeyboardMarkup(row_width=3)
        keyboard.add(*[types.InlineKeyboardButton(text=name,callback_data=name+'lim') for name in dictval])	
        msg = bot.send_message(c.message.chat.id,lengstr(ll,66),reply_markup=keyboard)
    if c.data[-3:]=='lim':
        payload = {'token': 'iQPU9x7HuK3wTPGFeZeTFRqlX8JCaMiqY2VCNoUdaf0ZtPBM1u'}
        r = requests.get('http://gate.paytochina.ru/rest/help/currencyList', params=payload)
        r=r.json()
        i=0
        dictval=''
        while True:
            try:
                if r['data'][i]['currency']==c.data[:-3]:
                    zzz=r['data'][i]['title_en']
                    if len(zzz)>30:
                        zzz=zzz[0:30]+'...'
                    else:
                        dictval+='<b>'+r['data'][i]['title_en']+'</b>:\n'+r['data'][i]['reserv']+'\n'				
                i+=1
            except Exception:
                break	
        msg = bot.send_message(c.message.chat.id,lengstr(ll,63)+'\n'+dictval,parse_mode='HTML')       	
    if c.data=='srok': 
        payload = {'token': 'iQPU9x7HuK3wTPGFeZeTFRqlX8JCaMiqY2VCNoUdaf0ZtPBM1u'}
        r = requests.get('http://gate.paytochina.ru/rest/help/Term', params=payload)
        r=r.json()
        z=list(r['data'].keys())
        print(z[0])
        i=0
        dictval=[]
        while True:
            try:
                dictval.append(z[i])			
                i+=1
            except Exception:
                break
        dictval=list(set(dictval))
        dictval.sort()
        keyboard = types.InlineKeyboardMarkup(row_width=3)
        keyboard.add(*[types.InlineKeyboardButton(text=name,callback_data=name+'sro') for name in dictval])	
        msg = bot.send_message(c.message.chat.id,lengstr(ll,66),reply_markup=keyboard) 
    if c.data[-3:]=='sro':
        payload = {'token': 'iQPU9x7HuK3wTPGFeZeTFRqlX8JCaMiqY2VCNoUdaf0ZtPBM1u'}
        r = requests.get('http://gate.paytochina.ru/rest/help/Term', params=payload)
        r=r.json()
        i=0
        z=list(r['data'][c.data[:-3]].keys())
        dictval=''
        print(r['data'][c.data[:-3]][z[0]])
        while True:
            try:
                dictval+='<b>'+z[i]+'</b>:'+r['data'][c.data[:-3]][z[i]]+'\n'				
                i+=1
            except Exception:
                break	
        msg = bot.send_message(c.message.chat.id,lengstr(ll,67)+'\n'+dictval,parse_mode='HTML')  
    if c.data=='komis': 
        payload = {'token': 'iQPU9x7HuK3wTPGFeZeTFRqlX8JCaMiqY2VCNoUdaf0ZtPBM1u'}
        r = requests.get('http://gate.paytochina.ru/rest/help/Term', params=payload)
        r=r.json()
        z=list(r['data'].keys())
        print(z[0])
        i=0
        dictval=[]
        while True:
            try:
                dictval.append(z[i])			
                i+=1
            except Exception:
                break
        dictval=list(set(dictval))
        dictval.sort()
        keyboard = types.InlineKeyboardMarkup(row_width=3)
        keyboard.add(*[types.InlineKeyboardButton(text=name,callback_data=name+'kom') for name in dictval])	
        msg = bot.send_message(c.message.chat.id,lengstr(ll,66),reply_markup=keyboard) 
    if c.data[-3:]=='kom':
        payload = {'token': 'iQPU9x7HuK3wTPGFeZeTFRqlX8JCaMiqY2VCNoUdaf0ZtPBM1u'}
        r = requests.get('http://gate.paytochina.ru/rest/help/Fee', params=payload)
        r=r.json()
        i=0
        z=list(r['data'][c.data[:-3]].keys())
        dictval=''
        print(r['data'][c.data[:-3]][z[0]])
        while True:
            try:
                dictval+='<b>'+z[i]+'</b>:'+r['data'][c.data[:-3]][z[i]]+'\n'				
                i+=1
            except Exception:
                break	
        msg = bot.send_message(c.message.chat.id,lengstr(ll,64)+'\n'+dictval,parse_mode='HTML') 		
		
		
		
		
		
def name(m):
    global bdpol
    k=nomer(m.chat.id)
    ll=bdpol[k].leng
    if m.text ==lengstr(ll,9):
        keyboard = types.InlineKeyboardMarkup(row_width=1)
        keyboard.add(*[types.InlineKeyboardButton(text=name,callback_data='optcard') for name in [lengstr(ll,54)]])
        keyboard.add(*[types.InlineKeyboardButton(text=name,callback_data='optkosh') for name in [lengstr(ll,57)]])
        keyboard.add(*[types.InlineKeyboardButton(text=name,callback_data='optld') for name in [lengstr(ll,55)]])
        msg = bot.send_message(m.chat.id,lengstr(ll,9),reply_markup=keyboard)
################## Dobavlenie schet/carti
    if    bdpol[k].addadd==3:
                print(bdpol[k].newtype,bdpol[k].session,m.text)
                payload = (('token', 'iQPU9x7HuK3wTPGFeZeTFRqlX8JCaMiqY2VCNoUdaf0ZtPBM1u'), ('id',int(bdpol[k].newtype)),('session',bdpol[k].session),('type','account'),('number',m.text))
                r = requests.post('http://gate.paytochina.ru/rest/user/SignIn', data=payload)
                r=r.json()    
                print(r)
                bdpol[k].addadd=0
                msg = bot.send_message(m.chat.id,lengstr(ll,78))     
                return()                 				
#### ############## Login
    if bdpol[k].login==1:
        bdpol[k].email=m.text
        msg = bot.send_message(m.chat.id,lengstr(ll,13)) 
        bdpol[k].login=2
        return	
    if bdpol[k].login==2:
        bdpol[k].passw=m.text
        payload = (('token', 'iQPU9x7HuK3wTPGFeZeTFRqlX8JCaMiqY2VCNoUdaf0ZtPBM1u'), ('login',bdpol[k].email),('password',bdpol[k].passw))
        r = requests.post('http://gate.paytochina.ru/rest/user/SignIn', data=payload)
        r=r.json()
        print(r)
        try:
            bdpol[k].session=r['data']['session']
        except Exception:
            keyboard = types.InlineKeyboardMarkup(row_width=1)
            keyboard.add(*[types.InlineKeyboardButton(text=name,callback_data='reg') for name in [lengstr(ll,2)]])
            keyboard.add(*[types.InlineKeyboardButton(text=name,callback_data='login') for name in [lengstr(ll,61)]])
            msg = bot.send_message(m.chat.id,lengstr(ll,74) ,reply_markup=keyboard)	     
            return()			
        print(bdpol[k].session)		
        bdpol[k].login=0
        keyboard = types.ReplyKeyboardMarkup(resize_keyboard=True)
        keyboard.add(*[types.InlineKeyboardButton(text=name,callback_data=name) for name in [lengstr(ll,3),lengstr(ll,4),lengstr(ll,5),lengstr(ll,6),lengstr(ll,7),lengstr(ll,8),lengstr(ll,9)]])
        msg = bot.send_message(m.chat.id,lengstr(ll,62),reply_markup=keyboard) 
        return	
################## Balans
    if m.text ==lengstr(ll,3):
        payload = (('token', 'iQPU9x7HuK3wTPGFeZeTFRqlX8JCaMiqY2VCNoUdaf0ZtPBM1u'), ('session',bdpol[k].session))
        r = requests.post('http://gate.paytochina.ru/rest/Personalarea/ListAccount', data=payload)
        r=r.json()
        print(r)
        cards=''
        chets=''
        for i in range(0,len(r['data'])):
            if 	r['data'][i]['account_type']=='card':
                cards=cards+r['data'][i]['title_en']+' '+r['data'][i]['amount']+' '+r['data'][i]['valut']+'\n'	
            if 	r['data'][i]['account_type']=='account':
                chets=chets+r['data'][i]['title_en']+' '+r['data'][i]['amount']+' '+r['data'][i]['valut']+'\n'        				
        msg = bot.send_message(m.chat.id,lengstr(ll,15)+cards) 
        msg = bot.send_message(m.chat.id,lengstr(ll,16)+chets) 	
################## Popolnit
    if m.text ==lengstr(ll,4):
        payload = {'token': 'iQPU9x7HuK3wTPGFeZeTFRqlX8JCaMiqY2VCNoUdaf0ZtPBM1u', 'login':bdpol[k].session}
        r = requests.get('http://gate.paytochina.ru/rest/Personalarea/urlPayment', params=payload)
        r=r.json()
        print(r)	
        msg = bot.send_message(m.chat.id,lengstr(ll,15)) 
        msg = bot.send_message(m.chat.id,lengstr(ll,16)) 
################## Vipiska po schetu
    if m.text ==lengstr(ll,3):
        payload = (('token', 'iQPU9x7HuK3wTPGFeZeTFRqlX8JCaMiqY2VCNoUdaf0ZtPBM1u'), ('session',bdpol[k].session))
        r = requests.post('http://gate.paytochina.ru/rest/Personalarea/Epitome', data=payload)
        r=r.json()
        print(r)
        cards=''
        for i in range(0,len(r['data'])):
            cards=cards+r['data'][i]['transactions_type']+'\n'+r['data'][i]['valut_from']+'->'+r['data'][i]['valut_to']+'\n'+r['data'][i]['title_from']+'->'+r['data'][i]['title_to']+'\n'+lengstr(ll,71)+' '+r['data'][i]['amount']+'\n'    				
        msg = bot.send_message(m.chat.id,cards) 		
################## FAQ menu
    if m.text ==lengstr(ll,8):
        keyboard = types.InlineKeyboardMarkup(row_width=1)
        keyboard.add(*[types.InlineKeyboardButton(text=name,callback_data='limits') for name in [lengstr(ll,63)]])
        keyboard.add(*[types.InlineKeyboardButton(text=name,callback_data='srok') for name in [lengstr(ll,67)]])
        keyboard.add(*[types.InlineKeyboardButton(text=name,callback_data='komis') for name in [lengstr(ll,64)]])
        keyboard.add(*[types.InlineKeyboardButton(text=name,callback_data='instr') for name in [lengstr(ll,65)]])
        msg = bot.send_message(m.chat.id,lengstr(ll,8),reply_markup=keyboard)
################## Poluchenie koda ot usera
    if bdpol[k].erpm==1:
        bdpol[k].erpm=m.text
        msg = bot.send_message(m.chat.id,lengstr(ll,12)) 
        bdpol[k].erp=1
        return
    if bdpol[k].erp==1:
        bdpol[k].erp=m.text
        msg = bot.send_message(m.chat.id,lengstr(ll,68)) 
        bdpol[k].erpkod=1
        return
    if bdpol[k].erpkod==1:
        payload = (('token', 'iQPU9x7HuK3wTPGFeZeTFRqlX8JCaMiqY2VCNoUdaf0ZtPBM1u'), ('code',m.text))
        r = requests.post('http://gate.paytochina.ru/rest/help/CodeVerification', data=payload)
        r=r.json()
        print(r)
        bdpol[k].erpkod=0
        msg = bot.send_message(m.chat.id,lengstr(ll,13)) 
        bdpol[k].passw=1
        return
    if bdpol[k].passw==1:
        bdpol[k].passw=m.text
        payload = (('token', 'iQPU9x7HuK3wTPGFeZeTFRqlX8JCaMiqY2VCNoUdaf0ZtPBM1u'), ('login',bdpol[k].erpm),('password',bdpol[k].passw),('phone',bdpol[k].erp))
        r = requests.post('http://gate.paytochina.ru/rest/user/SignUp', data=payload)
        r=r.json()
        print(r)
        if r['data']=='User Exists':
            msg = bot.send_message(m.chat.id,lengstr(ll,70)) 
            keyboard = types.InlineKeyboardMarkup(row_width=1)
            keyboard.add(*[types.InlineKeyboardButton(text=name,callback_data='reg') for name in [lengstr(ll,2)]])
            keyboard.add(*[types.InlineKeyboardButton(text=name,callback_data='login') for name in [lengstr(ll,61)]])
            msg = bot.send_message(m.chat.id,lengstr(ll,10) ,reply_markup=keyboard)		
        else:
            payload = (('token', 'iQPU9x7HuK3wTPGFeZeTFRqlX8JCaMiqY2VCNoUdaf0ZtPBM1u'), ('login',bdpol[k].erm),('password',bdpol[k].passw))
            r = requests.post('http://gate.paytochina.ru/rest/user/SignIn', data=payload)
            r=r.json()
            print(r)
            bdpol[k].session=r['data']['session']
            msg=bot.send_message(m.chat.id,lengstr(ll,69)) 	
            keyboard = types.ReplyKeyboardMarkup(resize_keyboard=True)
            keyboard.add(*[types.InlineKeyboardButton(text=name,callback_data=name) for name in [lengstr(ll,3),lengstr(ll,4),lengstr(ll,5),lengstr(ll,6),lengstr(ll,7),lengstr(ll,8),lengstr(ll,9)]])
            msg = bot.send_message(m.chat.id,lengstr(ll,62),reply_markup=keyboard) 			
        return
    if m.text== lengstr(ll,4) or m.text== lengstr(ll,5) or m.text== lengstr(ll,6):
            msg = bot.send_message(m.chat.id,'Не работает метод создания счета\карточки, по этому нельзя воспользоваться функцией.')        
################## Vivod valuti		
    if m.text ==lengstr(ll,6):	
        msg = bot.send_message(m.chat.id,lengstr(ll,48))
		
bot.remove_webhook()

bot.set_webhook(url=WEBHOOK_URL_BASE + WEBHOOK_URL_PATH,
                certificate=open(WEBHOOK_SSL_CERT, 'r'))

cherrypy.config.update({
    'server.socket_host': WEBHOOK_LISTEN,
    'server.socket_port': WEBHOOK_PORT,
    'server.ssl_module': 'builtin',
    'server.ssl_certificate': WEBHOOK_SSL_CERT,
    'server.ssl_private_key': WEBHOOK_SSL_PRIV
})

cherrypy.quickstart(WebhookServer(), WEBHOOK_URL_PATH, {'/': {}})