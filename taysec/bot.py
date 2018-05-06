import openpyxl
import time
import telebot
import schedule
import os
import urllib
import _thread
from telebot import types
import _pickle as pickle
import requests
import json
import buttons_create
import db
wb = openpyxl.load_workbook(filename = 'leng.xlsx')
sheet = wb['test']
val = sheet.cell(row=1, column=1)
print(val.value)
def  lengstr(leng,strn):
    bb=sheet.cell(row=strn, column=leng).value
    return str(bb)
TOKEN = '542861615:AAHTulSblaB0kdVUk44nxobpm2C6sxzeegA'

import cherrypy

WEBHOOK_HOST = '185.86.76.249'
WEBHOOK_PORT = 443  # 443, 80, 88 или 8443 (порт должен быть открыт!)
WEBHOOK_LISTEN = '0.0.0.0'  # На некоторых серверах придется указывать такой же IP, что и выше

WEBHOOK_SSL_CERT = './webhook_cert.pem'  # Путь к сертификату
WEBHOOK_SSL_PRIV = './webhook_pkey.pem'  # Путь к приватному ключу

WEBHOOK_URL_BASE = "https://%s:%s" % (WEBHOOK_HOST, WEBHOOK_PORT)
WEBHOOK_URL_PATH = "/%s/" % (TOKEN)

#ТУТ БОТ
bot = telebot.TeleBot(TOKEN)

class WebhookServer(object):
    @cherrypy.expose
    def index(self):
        if 'content-length' in cherrypy.request.headers and \
                        'content-type' in cherrypy.request.headers and \
                        cherrypy.request.headers['content-type'] == 'application/json':
            length = int(cherrypy.request.headers['content-length'])
            json_string = cherrypy.request.body.read(length).decode("utf-8")
            update = telebot.types.Update.de_json(json_string)
            bot.process_new_updates([update])
            return ''
        else:
            raise cherrypy.HTTPError(403)

class trick():
    name=''
    kategory=''
    buttons=[]
    buttons_url=[]
    photo_url=''
    text_head=''
    text_body=''
    type=0

class userobj():
    id=0
    leng=0
    dobavlenie=0
    phone=''
    balance=0
    pay_tricks=1
    day=3
    free_tricks=15
    adminin=0
    sleng=0
	
	
	
admin=0
new_trick=''
admin_pass='fisher_out_mark'
global bdpol,admin_addkat,admin_chkat,admin_add_qst	
bdpol=[]
admin_addkat=0
admin_add_qst=0
admin_chkat=-1
def nomer(b):
    global bdpol
    z=0
    k=-1
    for i in bdpol:
        k+=1
        if bdpol[k].id==b:
            z=k
            return(z)
			
kategories=[]
kategories_name=[]	

		
input = open('name.pkl', 'rb')
kategories_name = pickle.load(input)
input.close()	
	
input = open('kategories.pkl', 'rb')
kategories = pickle.load(input)
input.close()
			
input = open('bdpol.pkl', 'rb')
bdpol = pickle.load(input)
input.close()			
			
					
			
input = open('admin.pkl', 'rb')
admin = pickle.load(input)
input.close()				
			
# Handle '/start' and '/help'
@bot.message_handler(commands=['help', 'start'])
def start(message):
    global bdpol
    hesh=message.chat.id
    z=0
    k=-1
    for i in bdpol:
        k+=1
        if bdpol[k].id==hesh:
            z=1
            ll=bdpol[k].leng
            keyboard = types.ReplyKeyboardMarkup(resize_keyboard=True)
            keyboard.add(*[types.InlineKeyboardButton(text=name,callback_data=name) for name in [lengstr(ll,2)]])
            keyboard.add(*[types.InlineKeyboardButton(text=name,callback_data=name) for name in [lengstr(ll,3)]])
            keyboard.add(*[types.InlineKeyboardButton(text=name,callback_data=name) for name in [lengstr(ll,4)]])
            msg = bot.send_message(message.chat.id,lengstr(ll,7) ,reply_markup=keyboard)
            return
    if z==0:
        bdpol.append(userobj())
        print(bdpol[-1].id)
        bdpol[-1].id=hesh
        bdpol[-1].sleng=1
        bdpol[-1].dobavlenie=int(time.time())
    keyboard = types.InlineKeyboardMarkup(row_width=1)
    for i in range(1,20):
       if lengstr(i,1) !='None':
          keyboard.add(*[types.InlineKeyboardButton(text=name,callback_data=str(i)) for name in [lengstr(i,1)]])
    msg = bot.send_message(message.chat.id, 'Select language',
    reply_markup=keyboard)

	
	
@bot.message_handler(commands=['admin'])	
def admin(m):
    global bdpol
    if m.chat.id>0:
        msg = bot.send_message(m.chat.id, 'Добро пожаловать в панель администратора!\n\nВведите, пожалуйста, пароль администратора.',parse_mode='HTML')
        k=nomer(m.chat.id)
        bdpol[k].adminin=1
    else:
        return()	
	
	
	
@bot.message_handler(content_types=["text"])
def repeat_all_messages(message): 
    name(message)	
	
	
	
@bot.callback_query_handler(func=lambda c:True)
def inline(c):
    global bdpol, admin,admin_addkat,kategories_name,kategories, admin_chkat, new_trick, admin_add_qst
    k=nomer(c.message.chat.id)
    ll=bdpol[k].leng
    bib=c.message.message_id
    if bdpol[k].sleng ==1: 
        bdpol[k].leng =int(c.data)
        bdpol[k].sleng =0
        ll=bdpol[k].leng
        keyboard = types.ReplyKeyboardMarkup(resize_keyboard=True)
        keyboard.add(*[types.InlineKeyboardButton(text=name,callback_data=name) for name in [lengstr(ll,2)]])
        keyboard.add(*[types.InlineKeyboardButton(text=name,callback_data=name) for name in [lengstr(ll,3)]])
        keyboard.add(*[types.InlineKeyboardButton(text=name,callback_data=name) for name in [lengstr(ll,4)]])
        msg = bot.send_message(c.message.chat.id,lengstr(ll,7) ,reply_markup=keyboard)	
############################# rabota s kategoriami		
    if c.data=='addkat':
        msg = bot.send_message(c.message.chat.id, 'Введите название категории(Не более 30 символов)')	
        admin_addkat=1	
    if 'adkat' in c.data:
        msg = bot.send_message(c.message.chat.id, 'Введите название категории(Не более 30 символов)')	
        admin_chkat=int(c.data[5:])	
    if c.data=='delkat':
        keyboard = types.InlineKeyboardMarkup(row_width=1)
        for i in range(0,len(kategories_name)):
            keyboard.add(*[types.InlineKeyboardButton(text=name,callback_data='dlkat'+str(i)) for name in [kategories_name[i]]])               
        msg = bot.send_message(c.message.chat.id, 'Выберите категорию которую нужно удалить.',reply_markup=keyboard)
    if 'dlkat' in c.data:
        delete=int(c.data[5:])
        del kategories[delete]
        del kategories_name[delete]
        keyboard = types.InlineKeyboardMarkup(row_width=1)
        for i in range(0,len(kategories_name)):
            keyboard.add(*[types.InlineKeyboardButton(text=name,callback_data='adkat'+str(i)) for name in [kategories_name[i]]])        
        keyboard.add(*[types.InlineKeyboardButton(text=name,callback_data='addkat') for name in ['Добавить категорию']])
        keyboard.add(*[types.InlineKeyboardButton(text=name,callback_data='delkat') for name in ['Удалить категорию']])        
        msg = bot.send_message(c.message.chat.id, 'Категория удалена.',reply_markup=keyboard)           
########################################### nastroika trip trikov v kategorii
    if 'adtrk' in c.data:
        trik_tag=int(c.data[5:])
        keyboard = types.InlineKeyboardMarkup(row_width=1)
        for i in range(0,len(kategories[trik_tag])):
            keyboard.add(*[types.InlineKeyboardButton(text=name,callback_data='adtri'+str(trik_tag)+':'+str(i)) for name in [kategories[trik_tag][i].name]])        
        keyboard.add(*[types.InlineKeyboardButton(text=name,callback_data='addtrk'+str(trik_tag)) for name in ['Добавить ТрипТрик']])
        keyboard.add(*[types.InlineKeyboardButton(text=name,callback_data='deltrk'+str(trik_tag)) for name in ['Удалить ТрипТрик']])
        msg = bot.send_message(c.message.chat.id, 'Выберите категорию ТрипТрик для настройки',reply_markup=keyboard)		
    if 'addtrk' in c.data:
        msg = bot.send_message(c.message.chat.id, 'Введите название ТрипТрика, название отображается только в админ панели.')
        new_trick=trick()	
        new_trick.kategory=int(c.data[6:])	
###########################################	prosmotr triptrika
    if 'adtri' in c.data:
        for i in range(0,len(c.data)):
            if c.data[i]==':':
                trik_tag=int(c.data[5:i])
                trik_num=int(c.data[i+1:])
        chek_trick=kategories[trik_tag][trik_num]
        if chek_trick.type==0:
            payd='Бесплатный трип трик'
            paydbut='Изменить на платный'
        else:
            payd='Платный трип трик'
            paydbut='Изменить на бесплатный'
        keyboard = types.InlineKeyboardMarkup(row_width=1)   
        for z in range(0,len(chek_trick.buttons_url)):
                    url_button = types.InlineKeyboardButton(text=chek_trick.buttons[z], url=chek_trick.buttons_url[z])
                    keyboard.add(url_button)
        keyboard.add(*[types.InlineKeyboardButton(text=name,callback_data='uvedom'+str(trik_tag)+':'+str(trik_num)) for name in ['Разослать уведомление']]) 
        keyboard.add(*[types.InlineKeyboardButton(text=name,callback_data='pytri'+str(trik_tag)+':'+str(trik_num)) for name in [paydbut]])       					
        msg =bot.send_photo(c.message.chat.id, open(chek_trick.photo_url, 'rb'),caption=chek_trick.text_head)                    
        msg = bot.send_message(c.message.chat.id,chek_trick.text_body+'\n\n'+payd,reply_markup=keyboard)
######################################### uvedomlenie
    if 'uvedom' in c.data:
        for i in range(0,len(c.data)):
            if c.data[i]==':':
                trik_tag=int(c.data[6:i])
                trik_num=int(c.data[i+1:]) 
        chek_trick=kategories[trik_tag][trik_num]
        for i in range(0,len(bdpol)):
                if i%20 ==0:
                    time.sleep(1)
                msg = bot.send_photo(c.message.chat.id, open(chek_trick.photo_url, 'rb'),caption=chek_trick.text_head)
        msg = bot.send_message(c.message.chat.id, 'Сообщение отправлено всем пользователям')				
######################################### if pyd
    if 'pytri' in c.data:
        for i in range(0,len(c.data)):
            if c.data[i]==':':
                trik_tag=int(c.data[5:i])
                trik_num=int(c.data[i+1:])
        if kategories[trik_tag][trik_num].type==0:				
            kategories[trik_tag][trik_num].type=1
        else:
            kategories[trik_tag][trik_num].type=0
        chek_trick=kategories[trik_tag][trik_num]
        if chek_trick.type==0:
            payd='Бесплатный трип трик'
            paydbut='Изменить на платный'
        else:
            payd='Платный трип трик'
            paydbut='Изменить на бесплатный'
        keyboard = types.InlineKeyboardMarkup(row_width=1)   
        for z in range(0,len(chek_trick.buttons_url)):
                    url_button = types.InlineKeyboardButton(text=chek_trick.buttons[z], url=chek_trick.buttons_url[z])
                    keyboard.add(url_button)
        keyboard.add(*[types.InlineKeyboardButton(text=name,callback_data='pytri'+str(trik_tag)+':'+str(trik_num)) for name in [paydbut]])       					
        msg =bot.send_photo(c.message.chat.id, open(chek_trick.photo_url, 'rb'),caption=chek_trick.text_head)                    
        msg = bot.send_message(c.message.chat.id,chek_trick.text_body+'\n\n'+payd,reply_markup=keyboard)
######################################### udalenie tripa
    if 'deltrk' in c.data:
        trik_tag=int(c.data[6:])	
        keyboard = types.InlineKeyboardMarkup(row_width=1)
        for i in range(0,len(kategories[trik_tag])):
            keyboard.add(*[types.InlineKeyboardButton(text=name,callback_data='dltri'+str(trik_tag)+':'+str(i)) for name in [kategories[trik_tag][i].name]]) 
        msg = bot.send_message(c.message.chat.id, 'Нажмите на трип что-бы удалить его',reply_markup=keyboard)
    if 'dltri' in c.data:
        for i in range(0,len(c.data)):
            if c.data[i]==':':
                trik_tag=int(c.data[5:i])
                trik_num=int(c.data[i+1:])
        del kategories[trik_tag][trik_num] 
        msg = bot.send_message(c.message.chat.id, 'ТрипТрик удален')		
#######################################  prosmotr kategorii
    if 'seekat' in c.data:
        trik_tag=int(c.data[6:])	        
        trik_num=0
        chek_trick=kategories[trik_tag][trik_num] 
        keyboard = types.InlineKeyboardMarkup(row_width=1)	
        msg = bot.delete_message(c.message.chat.id, bib)		
        keyboard.add(*[types.InlineKeyboardButton(text=name,callback_data='otkrit'+str(trik_tag)+':'+str(trik_num)) for name in [lengstr(ll,10)]])		
        keyboard.add(*[types.InlineKeyboardButton(text=name,callback_data='nextrk'+str(trik_tag)+':'+str(trik_num+1)) for name in [lengstr(ll,14)]])
        msg =bot.send_photo(c.message.chat.id, open(chek_trick.photo_url, 'rb'),caption=chek_trick.text_head,reply_markup=keyboard) 
######################################## otkritie tripa
    if 'otkrit' in c.data:
        if bdpol[k].day<1 or bdpol[k].free_tricks<1 or bdpol[k].pay_tricks<1:
            msg=bot.send_message(c.message.chat.id, lengstr(ll,9),reply_markup=keyboard)
            return
        for i in range(0,len(c.data)):
            if c.data[i]==':':
                trik_tag=int(c.data[6:i])
                trik_num=int(c.data[i+1:])  
        bdpol[k].day-=1
        bdpol[k].free_tricks-=1		
        chek_trick=kategories[trik_tag][trik_num]
        keyboard = types.InlineKeyboardMarkup(row_width=1)
        for z in range(0,len(chek_trick.buttons_url)):
                    url_button = types.InlineKeyboardButton(text=chek_trick.buttons[z], url=chek_trick.buttons_url[z])
                    keyboard.add(url_button)
        keyboard.add(*[types.InlineKeyboardButton(text=name,callback_data='zakrit'+str(trik_tag)+':'+str(trik_num)) for name in [lengstr(ll,11)]])		
        msg = bot.send_message(c.message.chat.id,chek_trick.text_body,reply_markup=keyboard)	
####################################### zakritie tripa
    if 'zakrit' in c.data:
        if bdpol[k].free_tricks==0:
            keyboard = types.InlineKeyboardMarkup(row_width=1)
            keyboard.add(*[types.InlineKeyboardButton(text=name,callback_data='goopr') for name in [lengstr(ll,35)]])
            msg=bot.send_message(c.message.chat.id, lengstr(ll,34),reply_markup=keyboard)
            return
        for i in range(0,len(c.data)):
            if c.data[i]==':':
                trik_tag=int(c.data[6:i])
                trik_num=int(c.data[i+1:])  
        chek_trick=kategories[trik_tag][trik_num]
        msg = bot.delete_message(c.message.chat.id, bib-1)
        msg = bot.delete_message(c.message.chat.id, bib)
        keyboard = types.InlineKeyboardMarkup(row_width=2)
        keyboard.add(*[types.InlineKeyboardButton(text=name,callback_data='hihihi'+str(trik_tag)+':'+str(trik_num)) for name in [lengstr(ll,17),lengstr(ll,18)]])
        msg = bot.send_message(c.message.chat.id,lengstr(ll,12)+' '+str(bdpol[k].day)+'\n'+lengstr(ll,13)+' '+str(bdpol[k].free_tricks)+'\n'+lengstr(ll,16),reply_markup=keyboard)
##################################### posle otveta
    if 'hihihi' in c.data:
        for i in range(0,len(c.data)):
            if c.data[i]==':':
                trik_tag=int(c.data[6:i])
                trik_num=int(c.data[i+1:])  
        chek_trick=kategories[trik_tag][trik_num]
        msg = bot.delete_message(c.message.chat.id, bib)
        keyboard = types.InlineKeyboardMarkup(row_width=1)	
        keyboard.add(*[types.InlineKeyboardButton(text=name,callback_data='otkrit'+str(trik_tag)+':'+str(trik_num)) for name in [lengstr(ll,10)]])
        keyboard.add(*[types.InlineKeyboardButton(text=name,callback_data='nextrk'+str(trik_tag)+':'+str(trik_num+1)) for name in [lengstr(ll,14)]])		
        msg =bot.send_photo(c.message.chat.id, open(chek_trick.photo_url, 'rb'),caption=chek_trick.text_head,reply_markup=keyboard)  
####################################### sledushii trip trik	
    if 'nextrk' in c.data:
        for i in range(0,len(c.data)):
            if c.data[i]==':':
                trik_tag=int(c.data[6:i])
                trik_num=int(c.data[i+1:]) 
        if 	trik_num==len(kategories[trik_tag]):
            msg = bot.delete_message(c.message.chat.id, bib)
            msg = bot.send_message(c.message.chat.id,lengstr(ll,15))	
            return
        chek_trick=kategories[trik_tag][trik_num]
        msg = bot.delete_message(c.message.chat.id, bib)
        keyboard = types.InlineKeyboardMarkup(row_width=1)	
        keyboard.add(*[types.InlineKeyboardButton(text=name,callback_data='otkrit'+str(trik_tag)+':'+str(trik_num)) for name in [lengstr(ll,10)]])	
        keyboard.add(*[types.InlineKeyboardButton(text=name,callback_data='nextrk'+str(trik_tag)+':'+str(trik_num+1)) for name in [lengstr(ll,14)]])		
        msg =bot.send_photo(c.message.chat.id, open(chek_trick.photo_url, 'rb'),caption=chek_trick.text_head,reply_markup=keyboard) 
####################################### dobavlenie voprosa v opros
    if c.data=='addqst':
        msg=bot.send_message(c.message.chat.id,'Введите сообщение для опроса')
        admin_add_qst=1
####################################### udalenie voprosa
    if c.data=='delqst':
        gg=db.ask_question()
        keyboard = types.InlineKeyboardMarkup(row_width=1)
        s1=''
        for i in range(0,len(gg)):
            s1+=str(i+1)+') '+gg[i]+'\n' 
            keyboard.add(*[types.InlineKeyboardButton(text=name,callback_data='dlqst'+str(i)) for name in [str(i+1)]]) 
        msg=bot.send_message(c.message.chat.id,'Ввыберите какое сообщение удалить?\n'+s1,reply_markup=keyboard)
####################################### udalenie voprosa
    if 'dlqst' in c.data:
        qst_num=int(c.data[5:])  
        gg=db.ask_question()
        gg_str=gg[qst_num]	
        db.del_question(gg_str)	
        gg=db.ask_question()
        keyboard = types.InlineKeyboardMarkup(row_width=1)
        s1=''
        for i in range(0,len(gg)):
            s1+=str(i+1)+') '+gg[i]+'\n' 
        keyboard.add(*[types.InlineKeyboardButton(text=name,callback_data='addqst') for name in ['Добавить вопрос']])
        keyboard.add(*[types.InlineKeyboardButton(text=name,callback_data='delqst') for name in ['Удалить вопрос']])          
        msg = bot.send_message(c.message.chat.id, 'Вопрос удален:\n'+s1,reply_markup=keyboard)          		
####################################### menu polzovatela
####################################### dobavlenie nomera telefona
    if c.data== lengstr(ll,21): 
             keyboard = types.ReplyKeyboardMarkup(row_width=1, resize_keyboard=True)
             button_phone = types.KeyboardButton(text=lengstr(ll,30), request_contact=True)
             keyboard.add(button_phone)
             keyboard.add(*[types.InlineKeyboardButton(text=name,callback_data=name) for name in [lengstr(ll,32)]])
             msg=bot.send_message(c.message.chat.id, lengstr(ll,29),reply_markup=keyboard)  
####################################### opros
    if c.data=='goopr':
            ff=0
            ss=db.ask_question()
            if len(ss)>0:
                keyboard = types.InlineKeyboardMarkup(row_width=1)
                keyboard.add(*[types.InlineKeyboardButton(text=name,callback_data='reopr'+str(ff+1)+':+') for name in [lengstr(ll,17)]])
                keyboard.add(*[types.InlineKeyboardButton(text=name,callback_data='reopr'+str(ff+1)+':-') for name in [lengstr(ll,18)]])
                msg=bot.send_message(c.message.chat.id, ss[0],reply_markup=keyboard) 
            else:
                proshel_opros(c.message.chat.id)
    if 'reopr' in c.data:	
        for i in range(0,len(c.data)):
            if c.data[i]==':':
                trik_tag=int(c.data[5:i])
        ff=int(trik_tag)
        ss=db.ask_question()
        db.bid_response(ss[ff-1],c.data[-1])
        if len(ss)>ff:
            keyboard = types.InlineKeyboardMarkup(row_width=1)
            keyboard.add(*[types.InlineKeyboardButton(text=name,callback_data='reopr'+str(ff+1)+':+') for name in [lengstr(ll,17)]])
            keyboard.add(*[types.InlineKeyboardButton(text=name,callback_data='reopr'+str(ff+1)+':-') for name in [lengstr(ll,18)]])
            msg=bot.send_message(c.message.chat.id, ss[ff],reply_markup=keyboard) 
        else:
            proshel_opros(c.message.chat.id)		
#######################################	
    if c.data== lengstr(ll,22):		
            msg=bot.send_message(c.message.chat.id, lengstr(ll,27))		
    if c.data== lengstr(ll,23):		
            msg=bot.send_message(c.message.chat.id, lengstr(ll,28))
    if c.data== lengstr(ll,24):		
            msg=bot.send_message(c.message.chat.id, lengstr(ll,27))
    if c.data== lengstr(ll,25):		
            msg=bot.send_message(c.message.chat.id, lengstr(ll,27))
    if c.data== lengstr(ll,26):		
            msg=bot.send_message(c.message.chat.id, lengstr(ll,27))			
    return		
		
		
	














	
def name(m):
    global bdpol, admin,admin_addkat,kategories_name,kategories, admin_chkat, new_trick, admin_add_qst
    k=nomer(m.chat.id)
    ll=bdpol[k].leng
    if bdpol[k].adminin==1:
        if m.text==admin_pass:
            keyboard = types.ReplyKeyboardMarkup(resize_keyboard=True)
            keyboard.add(*[types.InlineKeyboardButton(text=name,callback_data=name) for name in ['Просмотр статистики']])
            keyboard.add(*[types.InlineKeyboardButton(text=name,callback_data=name) for name in ['Управление категориями']])
            keyboard.add(*[types.InlineKeyboardButton(text=name,callback_data=name) for name in ['Управление ТрипТриками']])
            keyboard.add(*[types.InlineKeyboardButton(text=name,callback_data=name) for name in ['Редакция опроса']])
            msg = bot.send_message(m.chat.id,lengstr(ll,6) ,reply_markup=keyboard)
            bdpol[k].adminin=0
            admin=m.chat.id
            output = open('admin.pkl', 'wb')
            pickle.dump(admin, output, 2)   
            output.close()	
        else:
            bdpol[k].adminin=0
            msg = bot.send_message(m.chat.id,'Пароль введен неверно' )
#################################### opros
    if m.text=='Редакция опроса':
        gg=db.ask_question()
        keyboard = types.InlineKeyboardMarkup(row_width=1)
        s1=''
        for i in range(0,len(gg)):
            s1+=str(i+1)+') '+gg[i]+'\n' 
        keyboard.add(*[types.InlineKeyboardButton(text=name,callback_data='addqst') for name in ['Добавить вопрос']])
        keyboard.add(*[types.InlineKeyboardButton(text=name,callback_data='delqst') for name in ['Удалить вопрос']])          
        msg = bot.send_message(m.chat.id, 'Вопросы:\n'+s1,reply_markup=keyboard)        
############################################## upravlenie categoriami
    if m.text=='Управление категориями' and admin==m.chat.id:
        output = open('name.pkl', 'wb')
        pickle.dump(kategories_name, output, 2)
        output.close()
        output = open('kategories.pkl', 'wb')
        pickle.dump(kategories, output, 2)
        output.close()
        keyboard = types.InlineKeyboardMarkup(row_width=1)
        for i in range(0,len(kategories_name)):
            keyboard.add(*[types.InlineKeyboardButton(text=name,callback_data='adkat'+str(i)) for name in [kategories_name[i]]]) 
        keyboard.add(*[types.InlineKeyboardButton(text=name,callback_data='addkat') for name in ['Добавить категорию']])
        keyboard.add(*[types.InlineKeyboardButton(text=name,callback_data='delkat') for name in ['Удалить категорию']])          
        msg = bot.send_message(m.chat.id, 'Выберите категорию трип триков что-бы изменить ее название.',reply_markup=keyboard)
############################################## upravlenie trip trikami
    if m.text=='Управление ТрипТриками' and admin==m.chat.id:
        keyboard = types.InlineKeyboardMarkup(row_width=1)
        for i in range(0,len(kategories_name)):
            keyboard.add(*[types.InlineKeyboardButton(text=name,callback_data='adtrk'+str(i)) for name in [kategories_name[i]]])         
        msg = bot.send_message(m.chat.id, 'Выберите категорию трип триков',reply_markup=keyboard)		
############################################## dobavlenie kategorii
    if admin_addkat==1 and m.chat.id==admin:
        if len(m.text)>30:
            msg = bot.send_message(m.chat.id, 'Превышенно кол-во символов! Введите снова, но с меньшим кол-вом символов.')
            return()
        admin_addkat=0
        kategories_name.append(m.text)
        kategories.append([])
        keyboard = types.InlineKeyboardMarkup(row_width=1)
        for i in range(0,len(kategories_name)):
            keyboard.add(*[types.InlineKeyboardButton(text=name,callback_data='adkat'+str(i)) for name in [kategories_name[i]]])        
        keyboard.add(*[types.InlineKeyboardButton(text=name,callback_data='addkat') for name in ['Добавить категорию']])
        keyboard.add(*[types.InlineKeyboardButton(text=name,callback_data='delkat') for name in ['Удалить категорию']])        
        msg = bot.send_message(m.chat.id, 'Категория добавлена.',reply_markup=keyboard)
############################################ izmenenie kategorii		
    if admin_chkat>-1 and m.chat.id==admin:
        if len(m.text)>30:
            msg = bot.send_message(m.chat.id, 'Превышенно кол-во символов! Введите снова, но с меньшим кол-вом символов.')	
            return()			
        else:
            kategories_name[admin_chkat]=m.text
            admin_chkat=-1
        keyboard = types.InlineKeyboardMarkup(row_width=1)
        for i in range(0,len(kategories_name)):
            keyboard.add(*[types.InlineKeyboardButton(text=name,callback_data='adkat'+str(i)) for name in [kategories_name[i]]])        
        keyboard.add(*[types.InlineKeyboardButton(text=name,callback_data='addkat') for name in ['Добавить категорию']])
        keyboard.add(*[types.InlineKeyboardButton(text=name,callback_data='delkat') for name in ['Удалить категорию']])        
        msg = bot.send_message(m.chat.id, 'Категория добавлена.',reply_markup=keyboard)  
########################################## sozdanie novogo trip trika
    if new_trick!='':
        if 	new_trick!='' and new_trick.name=='' and m.chat.id==admin:
            new_trick.name=m.text
            msg = bot.send_message(m.chat.id, 'Введите текст при предпросмотре ТрипТрика') 
            return			
        if 	new_trick.name!='' and new_trick.text_head=='' and admin==m.chat.id:
            new_trick.text_head=m.text
            msg = bot.send_message(m.chat.id, 'Загрузите фото для ТрипТрика')
            return	
        if 	new_trick.photo_url!='' and new_trick.text_body=='' and admin==m.chat.id:
            new_trick.text_body=m.text
            msg = bot.send_message(m.chat.id, lengstr(ll,100))
            return	
        if 	new_trick.text_body!='' and len(new_trick.buttons)==0 and admin==m.chat.id:
            buttons_take=	buttons_create.buttons(m.text)				
            new_trick.buttons=buttons_take[0]
            new_trick.buttons_url=buttons_take[1]
            kategories[new_trick.kategory].append(new_trick)
            trik_tag=int(new_trick.kategory)
            new_trick=''
            keyboard = types.InlineKeyboardMarkup(row_width=1)
            for i in range(0,len(kategories[trik_tag])):
                keyboard.add(*[types.InlineKeyboardButton(text=name,callback_data='adtri'+str(trik_tag)+':'+str(i)) for name in [kategories[trik_tag][i].name]])        
            keyboard.add(*[types.InlineKeyboardButton(text=name,callback_data='addtrk'+str(trik_tag)) for name in ['Добавить ТрипТрик']])
            keyboard.add(*[types.InlineKeyboardButton(text=name,callback_data='deltrk'+str(trik_tag)) for name in ['Удалить ТрипТрик']])
            msg = bot.send_message(m.chat.id, 'ТрипТрик создан нажмите на него для просмотра и изменения настроек.',reply_markup=keyboard)
            return	
#########################################  katalog trip trikov
    if m.text==lengstr(ll,2):
            keyboard = types.InlineKeyboardMarkup(row_width=1)
            for i in range(0,len(kategories_name)):
                keyboard.add(*[types.InlineKeyboardButton(text=name,callback_data='seekat'+str(i)) for name in [kategories_name[i]]]) 
            msg=bot.send_message(m.chat.id, lengstr(ll,8),reply_markup=keyboard)  
    if 	m.text==lengstr(ll,3):		
            msg=bot.send_message(m.chat.id, lengstr(ll,19))
####################################### vhod v menu polzovatela			
    if m.text==	lengstr(ll,4):
        if bdpol[k].phone=='':
            keyboard = types.InlineKeyboardMarkup(row_width=1)
            keyboard.add(*[types.InlineKeyboardButton(text=name,callback_data=name) for name in [lengstr(ll,21),lengstr(ll,22),lengstr(ll,23),lengstr(ll,24)]]) 
            msg=bot.send_message(m.chat.id, lengstr(ll,20),reply_markup=keyboard)
        else:
            keyboard = types.InlineKeyboardMarkup(row_width=1)
            keyboard.add(*[types.InlineKeyboardButton(text=name,callback_data=name) for name in [lengstr(ll,22),lengstr(ll,23),lengstr(ll,24),lengstr(ll,25),lengstr(ll,26)]]) 
            msg=bot.send_message(m.chat.id, lengstr(ll,20),reply_markup=keyboard) 
    if 	m.text==lengstr(ll,32):	
            keyboard = types.ReplyKeyboardMarkup(resize_keyboard=True)
            keyboard.add(*[types.InlineKeyboardButton(text=name,callback_data=name) for name in [lengstr(ll,2)]])
            keyboard.add(*[types.InlineKeyboardButton(text=name,callback_data=name) for name in [lengstr(ll,3)]])
            keyboard.add(*[types.InlineKeyboardButton(text=name,callback_data=name) for name in [lengstr(ll,4)]])
            msg = bot.send_message(m.chat.id,lengstr(ll,33) ,reply_markup=keyboard)		
            keyboard = types.InlineKeyboardMarkup(row_width=1)
            keyboard.add(*[types.InlineKeyboardButton(text=name,callback_data=name) for name in [lengstr(ll,21),lengstr(ll,22),lengstr(ll,23),lengstr(ll,24)]]) 
            msg=bot.send_message(m.chat.id, lengstr(ll,20),reply_markup=keyboard)
##################################### dobavlenie voprosa
    if admin==m.chat.id and admin_add_qst==1:
        admin_add_qst=0
        db.add_question(m.text)
        gg=db.ask_question()
        keyboard = types.InlineKeyboardMarkup(row_width=1)
        s1=''
        for i in range(0,len(gg)):
            
            s1+=str(i+1)+') '+gg[i]+'\n' 
        keyboard.add(*[types.InlineKeyboardButton(text=name,callback_data='addqst') for name in ['Добавить вопрос']])
        keyboard.add(*[types.InlineKeyboardButton(text=name,callback_data='delqst') for name in ['Удалить вопрос']])          
        msg = bot.send_message(m.chat.id, 'Вопросы:\n'+s1,reply_markup=keyboard)   
    output = open('bdpol.pkl', 'wb')
    pickle.dump(bdpol, output, 2)   
    output.close()
  




######################################## proshel opros

def proshel_opros(id):
    global bdpol
    k=nomer(id) 
    bdpol[k].free_tricks+=5
    msg	= bot.send_message(id, lengstr(ll,36))
		
######################################## zagruzka photo na server
@bot.message_handler(content_types=['photo'])
def photoget(message):
    if message.chat.id==admin and new_trick.text_head!='' and new_trick.photo_url=='':	
        fileid=(message.photo[2].file_id)
        bb=bot.get_file(fileid)
        bb=bb.file_path
        logo = urllib.request.urlopen("https://api.telegram.org/file/bot"+TOKEN+"/"+bb).read()
        new_trick.photo_url="photos/"+new_trick.name+str(new_trick.kategory)+'.jpg'
        f = open(new_trick.photo_url, "wb")
        f.write(logo)
        f.close()
        msg = bot.send_message(message.chat.id, 'Введите текст при просмотре ТрипТрика')
###################################### registracia po nomeru telefona		
@bot.message_handler(content_types=["contact"])
def check_chatid(message):
    print(message.contact.phone_number)
    global bdpol
    k=nomer(message.chat.id)
    ll=bdpol[k].leng
    bdpol[k].phone='+'+str(message.contact.phone_number)
    bdpol[k].free_tricks+=5
    keyboard = types.ReplyKeyboardMarkup(resize_keyboard=True)
    keyboard.add(*[types.InlineKeyboardButton(text=name,callback_data=name) for name in [lengstr(ll,2)]])
    keyboard.add(*[types.InlineKeyboardButton(text=name,callback_data=name) for name in [lengstr(ll,3)]])
    keyboard.add(*[types.InlineKeyboardButton(text=name,callback_data=name) for name in [lengstr(ll,4)]])
    msg = bot.send_message(message.chat.id,lengstr(ll,31) ,reply_markup=keyboard)	
    keyboard = types.InlineKeyboardMarkup(row_width=1)
    keyboard.add(*[types.InlineKeyboardButton(text=name,callback_data=name) for name in [lengstr(ll,22),lengstr(ll,23),lengstr(ll,24),lengstr(ll,25),lengstr(ll,26)]])
    msg = bot.send_message(message.chat.id,lengstr(ll,20) ,reply_markup=keyboard)

############################################ potok
def proverk():
    global bdpol
    for i in range(0,len(bdpol)):
        bdpol[i].day=3	



schedule.every().day.at("7:00").do(proverk)
def lal():
    while 1:
        schedule.run_pending()
        time.sleep(1)
_thread.start_new_thread(lal,())		
############################################ Hvost	

bot.remove_webhook()

bot.set_webhook(url=WEBHOOK_URL_BASE + WEBHOOK_URL_PATH,
                certificate=open(WEBHOOK_SSL_CERT, 'r'))

cherrypy.config.update({
    'server.socket_host': WEBHOOK_LISTEN,
    'server.socket_port': WEBHOOK_PORT,
    'server.ssl_module': 'builtin',
    'server.ssl_certificate': WEBHOOK_SSL_CERT,
    'server.ssl_private_key': WEBHOOK_SSL_PRIV
})

cherrypy.quickstart(WebhookServer(), WEBHOOK_URL_PATH, {'/': {}})