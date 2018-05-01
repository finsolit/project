import openpyxl
import time
import os
import urllib
import telebot
import sqlite3
import buttons_create
from telebot import types



wb = openpyxl.load_workbook(filename = 'leng.xlsx')
sheet = wb['test']
val = sheet.cell(row=1, column=1)
print(val.value)
def  lengstr(leng,strn):
    bb=sheet.cell(row=strn, column=leng).value
    return str(bb)
TOKEN = '507631866:AAHB0tjPBoeNXAABu9zAU7Zt4O8jbTgsZDE'

class post():
    user_id=0
    channel_id=0
    text=''
    reactions=[]
    document=''
    document_type=''
    buttons=[]
    buttons_url=[]
    post_id=0
    add_type=''
    pin=0
    mut=0
    def __init__(self):
        self.buttons_url=[]
        self.buttons=[]	
        self.reactions=[]


bot = telebot.TeleBot(TOKEN)
global add_channel, add_post, add_post_array,posts,change_post
posts=0
add_channel=[]
add_post=[]
add_post_array=[]
change_post=[]
# Handle '/start' and '/help'
@bot.message_handler(commands=['help', 'start'])
def start(message):
    keyboard = types.ReplyKeyboardMarkup(resize_keyboard=True)
    keyboard.add(*[types.InlineKeyboardButton(text=name,callback_data=name) for name in [lengstr(1,3),lengstr(1,4)]])
    keyboard.add(*[types.InlineKeyboardButton(text=name,callback_data=name) for name in [lengstr(1,5),lengstr(1,6)]])
    msg = bot.send_message(message.chat.id, lengstr(1,2),reply_markup=keyboard)

	
	
	
@bot.message_handler(content_types=["text"])
def repeat_all_messages(message): 
    name(message)	
	
	
	
@bot.callback_query_handler(func=lambda c:True)
def inline(c):
    global add_channel, add_post, add_post_array, posts,change_post
    ll=1
    bib=c.message.message_id
################################## Kanal 
    if c.data[-1]=='$':
        channel_id=c.data[:-1]
        channel_name=chan_name_func(channel_id)
        keyboard = types.InlineKeyboardMarkup(row_width=1)
        keyboard.add(*[types.InlineKeyboardButton(text=name,callback_data=str(channel_id)+'%') for name in ['Создать публикацию']])
        keyboard.add(*[types.InlineKeyboardButton(text=name,callback_data=str(channel_id)+'@') for name in ['Мои публикации']])
        keyboard.add(*[types.InlineKeyboardButton(text=name,callback_data=str(channel_id)+'&') for name in ['Автопостинг']])
        keyboard.add(*[types.InlineKeyboardButton(text=name,callback_data=str(channel_id)+'*') for name in ['Настройки']])
        msg =bot.send_message(c.message.chat.id, channel_name,reply_markup=keyboard)
        return		
################################# vse kanali
    if c.data =='Все каналы':
        channel_id=c.data[:-1]
        keyboard = types.InlineKeyboardMarkup(row_width=1)
        keyboard.add(*[types.InlineKeyboardButton(text=name,callback_data=name) for name in ['Создать публикацию']])
        keyboard.add(*[types.InlineKeyboardButton(text=name,callback_data=name) for name in ['Мои публикации']])
        keyboard.add(*[types.InlineKeyboardButton(text=name,callback_data=name) for name in ['Автопостинг']])
        keyboard.add(*[types.InlineKeyboardButton(text=name,callback_data=name) for name in ['Настройки']])
        msg =bot.send_message(c.message.chat.id, lengstr(ll,11),reply_markup=keyboard)
        return
############################### dobavlenie posta v kanale		
    if c.data[-1]=='%':
        add_post.append(c.message.chat.id)
        add_post_array.append(post())
        add_post_array[-1].user_id=c.message.chat.id
        add_post_array[-1].channel_id=int(c.data[:-1])
        add_post_array[-1].post_id=int(posts) 
        posts+=1
        keyboard = types.InlineKeyboardMarkup(row_width=1)
        keyboard.add(*[types.InlineKeyboardButton(text=name,callback_data=c.data[:-1]+'!') for name in ['Назад к каналу']])
        msg =bot.send_message(c.message.chat.id, lengstr(ll,12),reply_markup=keyboard) 
        return
############################## otmena dobavlenia posta v kanale		
    if c.data[-1]=='!':
        try:
            add_post.remove(c.message.chat.id)
        except Exception:
            a=0
        for i in range(0,len(add_post_array)):
            if 	add_post_array[i].user_id==c.message.chat.id:
                del add_post_array[i]
        channel_id=c.data[:-1]
        channel_name=chan_name_func(channel_id)
        keyboard = types.InlineKeyboardMarkup(row_width=1)
        keyboard.add(*[types.InlineKeyboardButton(text=name,callback_data=str(channel_id)+'%') for name in ['Создать публикацию']])
        keyboard.add(*[types.InlineKeyboardButton(text=name,callback_data=str(channel_id)+'@') for name in ['Мои публикации']])
        keyboard.add(*[types.InlineKeyboardButton(text=name,callback_data=str(channel_id)+'&') for name in ['Автопостинг']])
        keyboard.add(*[types.InlineKeyboardButton(text=name,callback_data=str(channel_id)+'*') for name in ['Настройки']])
        msg =bot.send_message(c.message.chat.id, channel_name,reply_markup=keyboard)  
        return
################################# publikacia posta
    if c.data[-1]=='o':
        for i in range(0,len(add_post_array)):
            if 	add_post_array[i].post_id==int(c.data[:-1]):
                print(add_post_array[i].channel_id)
                keyboard = types.InlineKeyboardMarkup()
                for z in range(0,len(add_post_array[i].buttons_url)):
                    print(add_post_array[i].buttons[z],add_post_array[i].buttons_url[z])
                    url_button = types.InlineKeyboardButton(text=add_post_array[i].buttons[z], url=add_post_array[i].buttons_url[z])
                    keyboard.add(url_button)
                if len(add_post_array[i].reactions)>0:
                    keyboard.add(*[types.InlineKeyboardButton(text=name+':0',callback_data='react') for name in add_post_array[i].reactions])   
                if 	add_post_array[i].mut==1:
                    mut=True
                else:
                    mut=False	
                if 	add_post_array[i].document_type=='':				
                    msg =bot.send_message(add_post_array[i].channel_id, add_post_array[i].text,reply_markup=keyboard,disable_notification=mut)
                if  add_post_array[i].document_type=='photo':
                    msg =bot.send_photo(add_post_array[i].channel_id, open('documents/'+add_post_array[i].document, 'rb'),caption=add_post_array[i].text,reply_markup=keyboard,disable_notification=mut)                    
                if  add_post_array[i].document_type=='audio':
                    msg =bot.send_audio(add_post_array[i].channel_id, open('documents/'+add_post_array[i].document, 'rb'),caption=add_post_array[i].text,reply_markup=keyboard,disable_notification=mut)
                if  add_post_array[i].document_type=='video':
                    msg =bot.send_video(add_post_array[i].channel_id, open('documents/'+add_post_array[i].document, 'rb'),caption=add_post_array[i].text,reply_markup=keyboard,disable_notification=mut)
                if  add_post_array[i].document_type=='document':
                    msg =bot.send_document(add_post_array[i].channel_id, open('documents/'+add_post_array[i].document, 'rb'),caption=add_post_array[i].text,reply_markup=keyboard,disable_notification=mut)

                joy=msg.message_id
                if 	add_post_array[i].pin==1:
                    msg=bot.pin_chat_message(add_post_array[i].channel_id,joy,disable_notification=mut)	               
                msg =bot.answer_callback_query(c.id,lengstr(ll,14)) 
                return				
################################# dobavlenie knopok k postu
    if c.data[-1]=='k':
        for i in range(0,len(add_post_array)):
            if 	add_post_array[i].post_id==int(c.data[:-1]):
                add_post_array[i].add_type='buttons'
                msg =bot.send_message(c.message.chat.id,lengstr(ll,15),disable_web_page_preview=True)
                change_post.append(c.message.chat.id)
                return					
################################# dobavlenie reakci				
    if c.data[-1]=='r':
        for i in range(0,len(add_post_array)):
            if 	add_post_array[i].post_id==int(c.data[:-1]):
                add_post_array[i].add_type='reactions'
                keyboard = types.InlineKeyboardMarkup(row_width=1)
                keyboard.add(*[types.InlineKeyboardButton(text=name,callback_data='addsmile'+name) for name in ['👍👎','😊😄😒','❤😐💔🤢','😄😊😔😱😡']])             
                msg =bot.send_message(c.message.chat.id,lengstr(ll,17),reply_markup=keyboard)
                change_post.append(c.message.chat.id)
                return	
################################ dobavlenie reakcii cherez smaili
    if 'addsmile' in c.data:
        for i in range(0,len(add_post_array)):
            if 	add_post_array[i].user_id==c.message.chat.id:
                if 	add_post_array[i].add_type=='reactions':
                    smiles=c.data[8:]
                    print(len(smiles))
                    add_post_array[i].reactions=[]
                    for z in range(0,len(smiles)):
                        print(smiles[z])
                        add_post_array[i].reactions.append(smiles[z])
                    post_id=add_post_array[i].post_id
                    channel_id=add_post_array[i].channel_id
                    keyboard = kukoard(post_id,channel_id)
						
                    msg =bot.send_message(c.message.chat.id,lengstr(ll,18),reply_markup=keyboard)
                    print(add_post_array[i].reactions)	
                    change_post.remove(c.message.chat.id)	
                    return	
################################## zakrepit
    if c.data[-1]=='z':
        for i in range(0,len(add_post_array)):
            if 	add_post_array[i].post_id==int(c.data[:-1]):
                if add_post_array[i].pin==0:
                    add_post_array[i].pin=1
                    msg =bot.answer_callback_query(c.id,'Пост будет закреплен')
                else:
                    add_post_array[i].pin=0	
                    msg =bot.answer_callback_query(c.id,'Пост не будет закреплен')					
                post_id=add_post_array[i].post_id
                channel_id=add_post_array[i].channel_id					
                keyboard = kukoard(post_id,channel_id)          
                msg =bot.edit_message_reply_markup(c.message.chat.id,bib,reply_markup=keyboard) 
                return	
################################## uvedomlenia
    if c.data[-1]=='u':
        for i in range(0,len(add_post_array)):
            if 	add_post_array[i].post_id==int(c.data[:-1]):
                if add_post_array[i].mut==0:
                    add_post_array[i].mut=1
                    msg =bot.answer_callback_query(c.id,'Пост не пришлет уведомление')
                else:
                    add_post_array[i].mut=0	
                    msg =bot.answer_callback_query(c.id,'Пост пришлет уведомление')					
                post_id=add_post_array[i].post_id
                channel_id=add_post_array[i].channel_id					
                keyboard = kukoard(post_id,channel_id)          
                msg =bot.edit_message_reply_markup(c.message.chat.id,bib,reply_markup=keyboard) 
                return
################################# publikacia posta
    if c.data[-1]=='p':
        for i in range(0,len(add_post_array)):
            if 	add_post_array[i].post_id==int(c.data[:-1]):
                print(add_post_array[i].channel_id)
                keyboard = types.InlineKeyboardMarkup()
                for z in range(0,len(add_post_array[i].buttons_url)):
                    print(add_post_array[i].buttons[z],add_post_array[i].buttons_url[z])
                    url_button = types.InlineKeyboardButton(text=add_post_array[i].buttons[z], url=add_post_array[i].buttons_url[z])
                    keyboard.add(url_button)
                if len(add_post_array[i].reactions)>0:
                    keyboard.add(*[types.InlineKeyboardButton(text=name+':0',callback_data='react') for name in add_post_array[i].reactions])   
                if 	add_post_array[i].mut==1:
                    mut=True
                else:
                    mut=False	
                if 	add_post_array[i].document_type=='':				
                    msg =bot.send_message(c.message.chat.id, add_post_array[i].text,reply_markup=keyboard,disable_notification=mut)
                if  add_post_array[i].document_type=='photo':
                    msg =bot.send_photo(c.message.chat.id, open('documents/'+add_post_array[i].document, 'rb'),caption=add_post_array[i].text,reply_markup=keyboard,disable_notification=mut)                    
                if  add_post_array[i].document_type=='audio':
                    msg =bot.send_audio(c.message.chat.id, open('documents/'+add_post_array[i].document, 'rb'),caption=add_post_array[i].text,reply_markup=keyboard,disable_notification=mut)
                if  add_post_array[i].document_type=='video':
                    msg =bot.send_video(c.message.chat.id, open('documents/'+add_post_array[i].document, 'rb'),caption=add_post_array[i].text,reply_markup=keyboard,disable_notification=mut)
                if  add_post_array[i].document_type=='document':
                    msg =bot.send_document(c.message.chat.id, open('documents/'+add_post_array[i].document, 'rb'),caption=add_post_array[i].text,reply_markup=keyboard,disable_notification=mut) 
                return
################################ nachat snachala
    #if c.data[-1]=='n':
        #for i in range(0,len(add_post_array)):
            #if 	add_post_array[i].post_id==int(c.data[:-1]):
                #add_post_array[i].reactions=[]
                #add_post_array[i].buttons=[]
                #add_post_array[i].buttons_url=[]
                #add_post_array[i].mut=0
                #add_post_array[i].pin=0
                #add_post.append(c.message.chat.id)
                #keyboard = types.InlineKeyboardMarkup(row_width=1)
                #keyboard.add(*[types.InlineKeyboardButton(text=name,callback_data=c.data[:-1]+'!') for name in ['Назад к каналу']])
                #msg =bot.send_message(c.message.chat.id, lengstr(ll,12),reply_markup=keyboard) 
                #return
###################################				
    msg =bot.answer_callback_query(c.id,'Эта функция пока не доступна')     				
    return		
		
		
		
		
		
def name(m):
    global add_channel, add_post, add_post_array, posts,change_post
    ll=1
#####################   Dobavlenie kanala
    if m.text==lengstr(ll,3) or m.text==lengstr(ll,4) or m.text==lengstr(ll,5) or m.text==lengstr(ll,6):
        try:
            add_post.remove(m.chat.id)
        except Exception:
            a=0
        for i in range(0,len(add_post_array)):
            if 	add_post_array[i].user_id==m.chat.id:
                del add_post_array[i]
    if m.text ==lengstr(ll,3):
        add_channel.append(m.chat.id)
        msg = bot.send_message(m.chat.id,lengstr(ll,7))
        return
    if m.chat.id in add_channel:
        new_id=m.forward_from_chat.id
        all_channel=chek_chan()
        if new_id not in all_channel:
            add_channel_func(m.chat.id,new_id,m.forward_from_chat.title)
            keyboard = types.InlineKeyboardMarkup(row_width=1)
            channel_list=chek_user_channels(m.chat.id)
            for i in range (0,len(channel_list[0])):
                keyboard.add(*[types.InlineKeyboardButton(text=name,callback_data=str(channel_list[0][i])+'$') for name in [channel_list[1][i]]])
            keyboard.add(*[types.InlineKeyboardButton(text=name,callback_data=name) for name in ['Все каналы']])
            msg = bot.send_message(m.chat.id,lengstr(ll,8),reply_markup=keyboard)
            add_channel.remove(m.chat.id)			
        else:
            msg = bot.send_message(m.chat.id,lengstr(ll,9))
            add_channel.remove(m.chat.id)
#####################   moi kanali
    if m.text==lengstr(ll,4):
            channel_list=chek_user_channels(m.chat.id)
            keyboard = types.InlineKeyboardMarkup(row_width=1)
            for i in range (0,len(channel_list[0])):
                keyboard.add(*[types.InlineKeyboardButton(text=name,callback_data=str(channel_list[0][i])+'$') for name in [channel_list[1][i]]])
            keyboard.add(*[types.InlineKeyboardButton(text=name,callback_data=name) for name in ['Все каналы']])
            msg = bot.send_message(m.chat.id,lengstr(ll,10),reply_markup=keyboard)
#####################  dobavlenie texta v novi post
    if m.chat.id in add_post:
        add_post.remove(m.chat.id)
        for i in range(0,len(add_post_array)):
            if 	add_post_array[i].user_id==m.chat.id:
                add_post_array[i].text=m.text
                post_id=add_post_array[i].post_id
                channel_id=add_post_array[i].channel_id
        keyboard = types.InlineKeyboardMarkup(row_width=2)
        keyboard.add(types.InlineKeyboardButton(text='Добавить кнопки',callback_data=str(post_id)+'k'),
        types.InlineKeyboardButton(text='Добавить реакцию',callback_data=str(post_id)+'r'),
        types.InlineKeyboardButton(text='Опубликовать',callback_data=str(post_id)+'o'),
        types.InlineKeyboardButton(text='Время публикации',callback_data=str(post_id)+'v'),
        types.InlineKeyboardButton(text='Сохранить',callback_data=str(post_id)+'s'),
        types.InlineKeyboardButton(text='Удалить',callback_data=str(post_id)+'d'),
        types.InlineKeyboardButton(text='Предпросмотр',callback_data=str(post_id)+'p'),
        types.InlineKeyboardButton(text='Начать сначала',callback_data=str(post_id)+'n'),
        types.InlineKeyboardButton(text='⚪️ Закрепить',callback_data=str(post_id)+'z'),
        types.InlineKeyboardButton(text='⚪️ Уведомления',callback_data=str(post_id)+'u'),
        types.InlineKeyboardButton(text='Вернуться к каналу',callback_data=str(channel_id)+'!'))
        msg =bot.send_message(m.chat.id, lengstr(ll,13),reply_markup=keyboard) 
############################### shapka izmeneni		
    if m.chat.id in change_post:
        for i in range(0,len(add_post_array)):
            if 	add_post_array[i].user_id==m.chat.id:
############################### dobavlenie knopok
                if 	add_post_array[i].add_type=='buttons':
                    buttons_take=	buttons_create.buttons(m.text)
                    print(buttons_take)					
                    add_post_array[i].buttons=buttons_take[0]
                    add_post_array[i].buttons_url=buttons_take[1]
                    change_post.remove(m.chat.id)
                    keyboard=kukoard(add_post_array[i].post_id,add_post_array[i].channel_id)					
                    msg =bot.send_message(m.chat.id, lengstr(ll,16),reply_markup=keyboard)					
############################### dobavlenie reakci					
                if 	add_post_array[i].add_type=='reactions':
                    smiles=m.text
                    add_post_array[i].reactions=[]
                    for z in range(0,len(smiles)):
                        add_post_array[i].reactions.append(smiles[z])
                    keyboard =kukoard(add_post_array[i].post_id,add_post_array[i].channel_id)
                    msg =bot.send_message(m.chat.id,lengstr(ll,18),reply_markup=keyboard)
                    print(add_post_array[i].reactions)
                    change_post.remove(m.chat.id)											
			
			
			
			
def add_channel_func(user_id,channel_id,title):
    conn = sqlite3.connect('BD.sqlite')
    cursor = conn.cursor()
	
    cursor.execute("insert into USERS values (:user_id, :channel_id, :title) ", {"user_id": user_id,"channel_id": channel_id,"title": title,})
    conn.commit()
    conn.close()
		
def chek_chan():
    conn = sqlite3.connect('BD.sqlite')
    cursor = conn.cursor()
	
    array=[]
    cursor.execute("SELECT CHANNEL FROM USERS")
    results = cursor.fetchall()
    for i in range(0,len(results)):
        array.append(results[i][0])
    conn.close()
    print(array)
    return(array)
	
def chek_user_channels(user_id):
    conn = sqlite3.connect('BD.sqlite')
    cursor = conn.cursor()
	
    array=[]
    array1=[]
    cursor.execute("SELECT * FROM USERS WHERE USER LIKE :user_id",{"user_id":user_id})
    results = cursor.fetchall()
    for i in range(0,len(results)):
        array.append(results[i][1])
        array1.append(results[i][2])		
    conn.close()
    print(array,array1)
    res=[]
    res.append(array)
    res.append(array1)	
    return(res)	
	
def chan_name_func(channel_id):
    conn = sqlite3.connect('BD.sqlite')
    cursor = conn.cursor()
	
    cursor.execute("SELECT TITLE FROM USERS WHERE CHANNEL LIKE :channel_id",{"channel_id":channel_id})
    results = cursor.fetchall()
    print(results)
    return(results)


def kukoard(post_id,channel_id):
                    for i in range(0,len(add_post_array)):
                        if 	add_post_array[i].post_id==post_id:
                            mut=add_post_array[i].mut
                            pin=add_post_array[i].pin
                    if mut==0:
                        mut_str='⚪️ '
                    else:
                        mut_str='🔘 '
                    if pin==0:
                        pin_str='⚪️ '
                    else:
                        pin_str='🔘 '
                    keyboard = types.InlineKeyboardMarkup(row_width=2)
                    keyboard.add(types.InlineKeyboardButton(text='Добавить кнопки',callback_data=str(post_id)+'k'),
                    types.InlineKeyboardButton(text='Добавить реакцию',callback_data=str(post_id)+'r'),
                    types.InlineKeyboardButton(text='Опубликовать',callback_data=str(post_id)+'o'),
                    types.InlineKeyboardButton(text='Время публикации',callback_data=str(post_id)+'v'),
                    types.InlineKeyboardButton(text='Сохранить',callback_data=str(post_id)+'s'),
                    types.InlineKeyboardButton(text='Удалить',callback_data=str(post_id)+'d'),
                    types.InlineKeyboardButton(text='Предпросмотр',callback_data=str(post_id)+'p'),
                    types.InlineKeyboardButton(text='Начать сначала',callback_data=str(post_id)+'n'),
                    types.InlineKeyboardButton(text=pin_str+'Закрепить',callback_data=str(post_id)+'z'),
                    types.InlineKeyboardButton(text=mut_str+'Уведомления',callback_data=str(post_id)+'u'),
                    types.InlineKeyboardButton(text='Вернуться к каналу',callback_data=str(channel_id)+'!'))
                    return(keyboard)
############################################# dostaem photo dla posta	
@bot.message_handler(content_types=['photo'])
def photoget(message):
    print(message)
    ll=1
    if message.chat.id in add_post:
        add_post.remove(message.chat.id)
        for i in range(0,len(add_post_array)):
            if 	add_post_array[i].user_id==message.chat.id:
                add_post_array[i].document_type='photo'
                add_post_array[i].text=message.caption
                print(message.text)
                add_post_array[i].document=str(add_post_array[i].post_id)+'.jpg'
                post_id=add_post_array[i].post_id
                channel_id=add_post_array[i].channel_id
                fileid=(message.photo[2].file_id)
                bb=bot.get_file(fileid)
                bb=bb.file_path
                logo = urllib.request.urlopen("https://api.telegram.org/file/bot"+TOKEN+"/"+bb).read()
                f = open("documents/"+add_post_array[i].document, "wb")
                f.write(logo)
                f.close()
                keyboard = types.InlineKeyboardMarkup(row_width=2)
                keyboard.add(types.InlineKeyboardButton(text='Добавить кнопки',callback_data=str(post_id)+'k'),
                types.InlineKeyboardButton(text='Добавить реакцию',callback_data=str(post_id)+'r'),
                types.InlineKeyboardButton(text='Опубликовать',callback_data=str(post_id)+'o'),
                types.InlineKeyboardButton(text='Время публикации',callback_data=str(post_id)+'v'),
                types.InlineKeyboardButton(text='Сохранить',callback_data=str(post_id)+'s'),
                types.InlineKeyboardButton(text='Удалить',callback_data=str(post_id)+'d'),
                types.InlineKeyboardButton(text='Предпросмотр',callback_data=str(post_id)+'p'),
                types.InlineKeyboardButton(text='Начать сначала',callback_data=str(post_id)+'n'),
                types.InlineKeyboardButton(text='⚪️ Закрепить',callback_data=str(post_id)+'z'),
                types.InlineKeyboardButton(text='⚪️ Уведомления',callback_data=str(post_id)+'u'),
                types.InlineKeyboardButton(text='Вернуться к каналу',callback_data=str(channel_id)+'!'))
                msg =bot.send_message(message.chat.id, lengstr(ll,13),reply_markup=keyboard) 
########################### dostaem audio	
@bot.message_handler(content_types=['audio'])
def photoget(message):
    print(message)
    ll=1
    if message.chat.id in add_post:
        msg =bot.send_message(message.chat.id, lengstr(ll,20))
        add_post.remove(message.chat.id)
        for i in range(0,len(add_post_array)):
            if 	add_post_array[i].user_id==message.chat.id:
                add_post_array[i].document_type='audio'
                add_post_array[i].text=message.caption
                print(message.text)
                add_post_array[i].document=str(add_post_array[i].post_id)+'.mp3'
                post_id=add_post_array[i].post_id
                channel_id=add_post_array[i].channel_id
                fileid=(message.audio.file_id)
                bb=bot.get_file(fileid)
                bb=bb.file_path
                logo = urllib.request.urlopen("https://api.telegram.org/file/bot"+TOKEN+"/"+bb).read()
                f = open("documents/"+add_post_array[i].document, "wb")
                f.write(logo)
                f.close()
                keyboard = types.InlineKeyboardMarkup(row_width=2)
                keyboard.add(types.InlineKeyboardButton(text='Добавить кнопки',callback_data=str(post_id)+'k'),
                types.InlineKeyboardButton(text='Добавить реакцию',callback_data=str(post_id)+'r'),
                types.InlineKeyboardButton(text='Опубликовать',callback_data=str(post_id)+'o'),
                types.InlineKeyboardButton(text='Время публикации',callback_data=str(post_id)+'v'),
                types.InlineKeyboardButton(text='Сохранить',callback_data=str(post_id)+'s'),
                types.InlineKeyboardButton(text='Удалить',callback_data=str(post_id)+'d'),
                types.InlineKeyboardButton(text='Предпросмотр',callback_data=str(post_id)+'p'),
                types.InlineKeyboardButton(text='Начать сначала',callback_data=str(post_id)+'n'),
                types.InlineKeyboardButton(text='⚪️ Закрепить',callback_data=str(post_id)+'z'),
                types.InlineKeyboardButton(text='⚪️ Уведомления',callback_data=str(post_id)+'u'),
                types.InlineKeyboardButton(text='Вернуться к каналу',callback_data=str(channel_id)+'!'))
                msg =bot.send_message(message.chat.id, lengstr(ll,13),reply_markup=keyboard)
####################### dostaem video
@bot.message_handler(content_types=['video'])
def photoget(message):
    print(message)
    ll=1
    if message.chat.id in add_post:
        msg =bot.send_message(message.chat.id, lengstr(ll,20))
        add_post.remove(message.chat.id)
        for i in range(0,len(add_post_array)):
            if 	add_post_array[i].user_id==message.chat.id:
                add_post_array[i].document_type='video'
                add_post_array[i].text=message.caption
                print(message.text)
                add_post_array[i].document=str(add_post_array[i].post_id)+'.mp4'
                post_id=add_post_array[i].post_id
                channel_id=add_post_array[i].channel_id
                fileid=(message.video.file_id)
                bb=bot.get_file(fileid)
                bb=bb.file_path
                logo = urllib.request.urlopen("https://api.telegram.org/file/bot"+TOKEN+"/"+bb).read()
                f = open("documents/"+add_post_array[i].document, "wb")
                f.write(logo)
                f.close()
                keyboard = types.InlineKeyboardMarkup(row_width=2)
                keyboard.add(types.InlineKeyboardButton(text='Добавить кнопки',callback_data=str(post_id)+'k'),
                types.InlineKeyboardButton(text='Добавить реакцию',callback_data=str(post_id)+'r'),
                types.InlineKeyboardButton(text='Опубликовать',callback_data=str(post_id)+'o'),
                types.InlineKeyboardButton(text='Время публикации',callback_data=str(post_id)+'v'),
                types.InlineKeyboardButton(text='Сохранить',callback_data=str(post_id)+'s'),
                types.InlineKeyboardButton(text='Удалить',callback_data=str(post_id)+'d'),
                types.InlineKeyboardButton(text='Предпросмотр',callback_data=str(post_id)+'p'),
                types.InlineKeyboardButton(text='Начать сначала',callback_data=str(post_id)+'n'),
                types.InlineKeyboardButton(text='⚪️ Закрепить',callback_data=str(post_id)+'z'),
                types.InlineKeyboardButton(text='⚪️ Уведомления',callback_data=str(post_id)+'u'),
                types.InlineKeyboardButton(text='Вернуться к каналу',callback_data=str(channel_id)+'!'))
                msg =bot.send_message(message.chat.id, lengstr(ll,13),reply_markup=keyboard)
######################### dostaem document
@bot.message_handler(content_types=['document'])
def photoget(message):
    print(message)
    ll=1
    if message.chat.id in add_post:
        msg =bot.send_message(message.chat.id, lengstr(ll,20))
        add_post.remove(message.chat.id)
        for i in range(0,len(add_post_array)):
            if 	add_post_array[i].user_id==message.chat.id:
                add_post_array[i].document_type='document'
                add_post_array[i].text=message.caption
                print(message.text)
                file_get_ex=''
                for z in range(0,len(message.document.file_name)):
                    if message.document.file_name[z]=='.':
                        file_get_ex=message.document.file_name[z:]
                add_post_array[i].document=str(add_post_array[i].post_id)+file_get_ex
                post_id=add_post_array[i].post_id
                channel_id=add_post_array[i].channel_id
                fileid=(message.document.file_id)
                bb=bot.get_file(fileid)
                bb=bb.file_path
                logo = urllib.request.urlopen("https://api.telegram.org/file/bot"+TOKEN+"/"+bb).read()
                f = open("documents/"+add_post_array[i].document, "wb")
                f.write(logo)
                f.close()
                keyboard = types.InlineKeyboardMarkup(row_width=2)
                keyboard.add(types.InlineKeyboardButton(text='Добавить кнопки',callback_data=str(post_id)+'k'),
                types.InlineKeyboardButton(text='Добавить реакцию',callback_data=str(post_id)+'r'),
                types.InlineKeyboardButton(text='Опубликовать',callback_data=str(post_id)+'o'),
                types.InlineKeyboardButton(text='Время публикации',callback_data=str(post_id)+'v'),
                types.InlineKeyboardButton(text='Сохранить',callback_data=str(post_id)+'s'),
                types.InlineKeyboardButton(text='Удалить',callback_data=str(post_id)+'d'),
                types.InlineKeyboardButton(text='Предпросмотр',callback_data=str(post_id)+'p'),
                types.InlineKeyboardButton(text='Начать сначала',callback_data=str(post_id)+'n'),
                types.InlineKeyboardButton(text='⚪️ Закрепить',callback_data=str(post_id)+'z'),
                types.InlineKeyboardButton(text='⚪️ Уведомления',callback_data=str(post_id)+'u'),
                types.InlineKeyboardButton(text='Вернуться к каналу',callback_data=str(channel_id)+'!'))
                msg =bot.send_message(message.chat.id, lengstr(ll,13),reply_markup=keyboard)

				
	
if __name__ == '__main__':

            bot.polling(none_stop=True)