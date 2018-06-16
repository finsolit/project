import openpyxl
import time
import os
from datetime import datetime
import urllib
import telebot
import _thread
import sqlite3
import schedule
import buttons_create
import json
import struct
import _pickle as pickle
from telebot import types
import cherrypy


wb = openpyxl.load_workbook(filename = 'leng1.xlsx')
sheet = wb['test']
val = sheet.cell(row=1, column=84)
print(val.value)
def  lengstr(leng,strn):
    bb=sheet.cell(row=strn, column=leng).value
    return str(bb)
TOKEN = '502527811:AAFmZk2ZhOrtToLsFprcLzucoPAtIwbYZEY'


WEBHOOK_HOST = '95.46.98.28'
WEBHOOK_PORT = 88  # 443, 80, 88 или 8443 (порт должен быть открыт!)
WEBHOOK_LISTEN = '0.0.0.0'  # На некоторых серверах придется указывать такой же IP, что и выше

WEBHOOK_SSL_CERT = './webhook_cert.pem'  # Путь к сертификату
WEBHOOK_SSL_PRIV = './webhook_pkey.pem'  # Путь к приватному ключу

WEBHOOK_URL_BASE = "https://%s:%s" % (WEBHOOK_HOST, WEBHOOK_PORT)
WEBHOOK_URL_PATH = "/%s/" % (TOKEN)

#ТУТ БОТ
bot = telebot.TeleBot(TOKEN)

class WebhookServer(object):
    @cherrypy.expose
    def index(self):
        if 'content-length' in cherrypy.request.headers and \
                        'content-type' in cherrypy.request.headers and \
                        cherrypy.request.headers['content-type'] == 'application/json':
            length = int(cherrypy.request.headers['content-length'])
            json_string = cherrypy.request.body.read(length).decode("utf-8")
            update = telebot.types.Update.de_json(json_string)
            bot.process_new_updates([update])
            return ''
        else:
            raise cherrypy.HTTPError(403)


class post():
    user_id=0
    channel_id=0
    text=''
    reactions=[]
    document=''
    document_type=''
    buttons=[]
    buttons_url=[]
    post_id=0
    add_type=''
    pin=0
    mut=-1
    saved=0
    delete_time=0
    forse_time=0
    def __init__(self):
        self.buttons_url=[]
        self.buttons=[]	
        self.reactions=[]


bot = telebot.TeleBot(TOKEN)
global add_channel, add_post,posts,change_post,reverse_time_id,reverse_time,server_time
server_time=3
input = open('posts.pkl', 'rb')
posts = pickle.load(input)
input.close()
add_channel=[]
add_post=[]
change_post=[]
reverse_time_id=[]
reverse_time=[]
# Handle '/start' and '/help'
@bot.message_handler(commands=['help', 'start'])
def start(message):
    ll=chek_leng(message.chat.id)
    if ll==1:
        keyboard = types.InlineKeyboardMarkup(row_width=3)
        for i in range(0,9):
            keyboard.add(types.InlineKeyboardButton(text=lengstr(3*i+1+1,1),callback_data='selleng'+str(3*i+1+1)),
            types.InlineKeyboardButton(text=lengstr(3*i+2+1,1),callback_data='selleng'+str(3*i+2+1)),
            types.InlineKeyboardButton(text=lengstr(3*i+3+1,1),callback_data='selleng'+str(3*i+3+1))) 
        keyboard.add(types.InlineKeyboardButton(text='>>',callback_data='lengsnext'+str(1)))			
        msg = bot.send_message(message.chat.id, 'Choose language',reply_markup=keyboard) 
        return		
    keyboard = types.ReplyKeyboardMarkup(resize_keyboard=True)
    keyboard.add(*[types.InlineKeyboardButton(text=name,callback_data=name) for name in [lengstr(ll,3),lengstr(ll,4)]])
    keyboard.add(*[types.InlineKeyboardButton(text=name,callback_data=name) for name in [lengstr(ll,5),lengstr(ll,6)]])
    msg = bot.send_message(message.chat.id, lengstr(ll,2),reply_markup=keyboard)

@bot.message_handler(commands=['userstat'])
def userstat(message):
    uc=str(users_count())
    cc=str(cannel_count())
    ccc=str(chat_mem_count())
    msg = bot.send_message(message.chat.id, lengstr(1,56)+uc+'\n'+lengstr(1,57)+cc+'\n'+lengstr(1,58)+ccc)
	
@bot.message_handler(commands=['use'])
def use(message):
    for i in range(0,100000000000):
        try:
            msg=bot.get_chat(i)
            print(msg)
        except Exception as e:
            print(e)
            print(i)			
	
@bot.message_handler(content_types=["text"])
def repeat_all_messages(message): 
    name(message)	
	
	
	
@bot.callback_query_handler(func=lambda c:True)
def inline(c):
    global add_channel, add_post, posts,change_post,reverse_time_id,reverse_time
    ll=chek_leng(c.message.chat.id)
    bib=c.message.message_id

################################## vibor iazika
    if 'selleng' in c.data:
        ll=int(c.data[7:])
        add_leng(c.message.chat.id,ll)
        keyboard = types.ReplyKeyboardMarkup(resize_keyboard=True)
        keyboard.add(*[types.InlineKeyboardButton(text=name,callback_data=name) for name in [lengstr(ll,3),lengstr(ll,4)]])
        keyboard.add(*[types.InlineKeyboardButton(text=name,callback_data=name) for name in [lengstr(ll,5),lengstr(ll,6)]])
        msg = bot.send_message(chat_id=c.message.chat.id, text=lengstr(ll,2),reply_markup=keyboard)          
################################## lengi
    if 'lengsnext' in c.data:
        numb=int(c.data[-1])
        keyboard = types.InlineKeyboardMarkup(row_width=3)
        for i in range(numb*9,numb*9+9):
            if lengstr(3*i+1+1,1)=='None':
                break
            keyboard.add(types.InlineKeyboardButton(text=lengstr(3*i+1+1,1),callback_data='selleng'+str(3*i+1+1)),
            types.InlineKeyboardButton(text=lengstr(3*i+2+1,1),callback_data='selleng'+str(3*i+2+1)),
            types.InlineKeyboardButton(text=lengstr(3*i+3+1,1),callback_data='selleng'+str(3*i+3+1))) 
        if numb == 0:
            keyboard.add(types.InlineKeyboardButton(text='>>',callback_data='lengsnext'+str(1)))	
        elif numb ==3:
            keyboard.add(types.InlineKeyboardButton(text='<<',callback_data='lengsnext'+str(2)))
        else:
            keyboard.add(types.InlineKeyboardButton(text='<<',callback_data='lengsnext'+str(numb-1)),
			types.InlineKeyboardButton(text='>>',callback_data='lengsnext'+str(numb+1)))
        msg = bot.edit_message_text(chat_id=c.message.chat.id, message_id=bib, text='Choose language',reply_markup=keyboard)  
        return	        
################################## Donate
    if c.data=='Donate':
        keyboard = types.InlineKeyboardMarkup()
        callback_button = types.InlineKeyboardButton(text="Копилка", url="https://t.me/Blockchain_info_bot?start=13jtTtFix1ji1j8dzk3WAeo6B1A3hY9FKX")
        keyboard.add(callback_button)
        msg = bot.send_photo(c.message.chat.id, 'AgADAgAD4KgxG1I8SEiCZ75xiZ942I50qw4ABL7xdThk0wYt3PoBAAEC',lengstr(ll,40),reply_markup=keyboard)
        return
################################## reakcii
    if 'roact' in c.data:
        for i in range(0,len(c.data)):
            if c.data[i]==':':
                num=int(c.data[5:i])
                res=reactions_forse(bib,c.message.chat.id,c.from_user.id,num)
                if res=='Y':
                    msg =bot.answer_callback_query(c.id,lengstr(ll,59))
                else:
                    msg =bot.answer_callback_query(c.id,lengstr(ll,60))                    
                return()                
################################## Kanal 
    if c.data[-1]=='$':
        channel_id=c.data[:-1]
        chek_channel_options(channel_id)
        channel_name=chan_name_func(channel_id)
        msg = bot.get_chat(channel_id)
        tmbl=chek_timeb(channel_id)
        print(tmbl)
        if tmbl == None:
            keyboard = types.InlineKeyboardMarkup(row_width=4)
            for i in range(0,6):
                keyboard.add(*[types.InlineKeyboardButton(text=str(name),callback_data='tp1n'+str(name)+':'+str(channel_id)) for name in [-12+(4*i)+1,-12+(4*i)+2,-12+(4*i)+3,-12+(4*i)+4]])
            keyboard.add(*[types.InlineKeyboardButton(text=name,callback_data=str(0)+'Gur') for name in [lengstr(ll,61)]])
            msg = bot.edit_message_text(chat_id=c.message.chat.id, message_id=bib, text=lengstr(ll,55),reply_markup=keyboard)       		
            return	           
        print(msg.id)
        try:
            print('https://t.me/'+msg.username)
            usrnm='https://t.me/'+msg.username
        except Exception:
            usrnm=lengstr(ll,70)
        usid=str(msg.id)
        print(msg)
        txt=lengstr(ll,69)+': ['+channel_name[0][0]+']('+usrnm+')\nChannelID: '+usid[4:]
        keyboard = types.InlineKeyboardMarkup(row_width=2)
        keyboard.add(types.InlineKeyboardButton(text=lengstr(ll,62),callback_data=str(channel_id)+'%'),
                     types.InlineKeyboardButton(text=lengstr(ll,63),callback_data=str(channel_id)+'@'))
        keyboard.add(types.InlineKeyboardButton(text=lengstr(ll,64),callback_data=str(channel_id)+'&'),
                     types.InlineKeyboardButton(text=lengstr(ll,65),callback_data=str(channel_id)+'rekl'))
        keyboard.add(types.InlineKeyboardButton(text=lengstr(ll,66),callback_data=str(channel_id)+'stat'),
                     types.InlineKeyboardButton(text=lengstr(ll,67),callback_data=str(channel_id)+'*'))
        keyboard.add(*[types.InlineKeyboardButton(text=name,callback_data=str(0)+'off') for name in [lengstr(ll,68)]])
        keyboard.add(*[types.InlineKeyboardButton(text=name,callback_data=str(0)+'Gur') for name in [lengstr(ll,61)]])
        msg = bot.edit_message_text(chat_id=c.message.chat.id, message_id=bib, text=txt,reply_markup=keyboard,parse_mode='Markdown',disable_web_page_preview=True)
        return
################################# Pervi vibor chasovogo poiasa	
    if c.data[:4]=='tp1n':
        for i in range(0,len(c.data)):
            if c.data[i]==':':
                channel_id=int(c.data[i+1:])
                belt=int(c.data[4:i])
        zz=chek_channel_options(channel_id)
        change_belt_t(channel_id,belt)				
        chek_channel_options(channel_id)
        channel_name=chan_name_func(channel_id)
        msg = bot.get_chat(channel_id)
        tmbl=chek_timeb(channel_id)
        print(tmbl)
        if tmbl == None:
            keyboard = types.InlineKeyboardMarkup(row_width=4)
            for i in range(0,6):
                keyboard.add(*[types.InlineKeyboardButton(text=str(name),callback_data='tp1n'+str(name)+':'+str(channel_id)) for name in [-12+(4*i)+1,-12+(4*i)+2,-12+(4*i)+3,-12+(4*i)+4]])
            keyboard.add(*[types.InlineKeyboardButton(text=name,callback_data=str(0)+'Gur') for name in [lengstr(ll,61)]])
            msg = bot.edit_message_text(chat_id=c.message.chat.id, message_id=bib, text=lengstr(ll,55),reply_markup=keyboard)       		
            return	           
        print(msg.id)
        try:
            print('https://t.me/'+msg.username)
            usrnm='https://t.me/'+msg.username
        except Exception:
            usrnm=lengstr(ll,70)
        usid=str(msg.id)
        print(msg)
        txt=lengstr(ll,69)+': ['+channel_name[0][0]+']('+usrnm+')\nChannelID: '+usid[4:]
        keyboard = types.InlineKeyboardMarkup(row_width=2)
        keyboard.add(types.InlineKeyboardButton(text=lengstr(ll,62),callback_data=str(channel_id)+'%'),
                     types.InlineKeyboardButton(text=lengstr(ll,63),callback_data=str(channel_id)+'@'))
        keyboard.add(types.InlineKeyboardButton(text=lengstr(ll,64),callback_data=str(channel_id)+'&'),
                     types.InlineKeyboardButton(text=lengstr(ll,65),callback_data=str(channel_id)+'rekl'))
        keyboard.add(types.InlineKeyboardButton(text=lengstr(ll,66),callback_data=str(channel_id)+'stat'),
                     types.InlineKeyboardButton(text=lengstr(ll,67),callback_data=str(channel_id)+'*'))
        keyboard.add(*[types.InlineKeyboardButton(text=name,callback_data=str(0)+'off') for name in [lengstr(ll,68)]])
        keyboard.add(*[types.InlineKeyboardButton(text=name,callback_data=str(0)+'Gur') for name in [lengstr(ll,61)]])
        msg = bot.edit_message_text(chat_id=c.message.chat.id, message_id=bib, text=txt,reply_markup=keyboard,parse_mode='Markdown',disable_web_page_preview=True)
        return	
################################# caption for files
    if 'cpt' in c.data:
                obj_post=find_post_cok(c.message.chat.id) 
                obj=obj_post[1]	
                obj.add_type='caption'
                dump_post_cok(c.message.chat.id,obj)
                keyboard = types.InlineKeyboardMarkup(row_width=1)
                keyboard.add(*[types.InlineKeyboardButton(text=name,callback_data='<') for name in [lengstr(ll,32)]])	
                try:
                    msg = bot.edit_message_text(chat_id=c.message.chat.id, message_id=bib, text=lengstr(ll,41),disable_web_page_preview=True,reply_markup=keyboard)
                except Exception:
                    msg=bot.delete_message(c.message.chat.id,bib)
                    msg = bot.send_message(chat_id=c.message.chat.id, text=lengstr(ll,41),disable_web_page_preview=True,reply_markup=keyboard)
                return	        
#################################otkluchenie kanala
    if 'off' in c.data:
        num=int(c.data[:-3])
        msg =bot.answer_callback_query(c.id,lengstr(ll,71)) 
        return
################################# statistika kanala
    if 'stat' in c.data:
        num=int(c.data[:-4])
        msg =bot.answer_callback_query(c.id,lengstr(ll,71)) 
        return
################################# reklama kanala
    if 'rekl' in c.data:
        num=int(c.data[:-4])
        msg =bot.answer_callback_query(c.id,lengstr(ll,71)) 
        return
################################# razdelenie kanalov
    if 'Gur' in c.data:
            num=int(c.data[:-3])
            channel_list=chek_user_channels(c.message.chat.id)
            keyboard = types.InlineKeyboardMarkup(row_width=2)
            if len(channel_list[0])>num*8+8:
                top=num*8+8  
            else:
                top=len(channel_list[0])			
            for i in range (num*8,top): 
                keyboard.add(*[types.InlineKeyboardButton(text=name,callback_data=str(channel_list[0][i])+'$') for name in [channel_list[1][i]]])
            if len(channel_list[0])>num*8+8:
                keyboard.add(*[types.InlineKeyboardButton(text=name,callback_data=str(num+1)+'Gur') for name in ['»']])
            if num>0:
                keyboard.add(*[types.InlineKeyboardButton(text=name,callback_data=str(num-1)+'Gur') for name in ['«']])			
            keyboard.add(*[types.InlineKeyboardButton(text=name,callback_data=name) for name in [lengstr(ll,72)]])
            msg = bot.edit_message_text(chat_id=c.message.chat.id, message_id=bib,text=lengstr(ll,73)+': '+str(len(channel_list[0])),reply_markup=keyboard)
            return        	
################################# vse kanali
    if c.data ==lengstr(ll,72):
        channel_id=c.data[:-1]
        keyboard = types.InlineKeyboardMarkup(row_width=1)
        keyboard.add(*[types.InlineKeyboardButton(text=name,callback_data=name) for name in [lengstr(ll,62)]])
        keyboard.add(*[types.InlineKeyboardButton(text=name,callback_data=name) for name in [lengstr(ll,63)]])
        keyboard.add(*[types.InlineKeyboardButton(text=name,callback_data=name) for name in [lengstr(ll,64)]])
        keyboard.add(*[types.InlineKeyboardButton(text=name,callback_data=name) for name in [lengstr(ll,67)]])
        msg = bot.edit_message_text(chat_id=c.message.chat.id, message_id=bib, text=lengstr(ll,11),reply_markup=keyboard)
        return
############################### knopki v kanale izmenenie

    if c.data[-2:]=='dp':
		
        channel_id=c.data[:-2]
        change_option_channel(4,0,channel_id) 
        zz=chek_channel_options(channel_id)
        keyboard=option_keyboard(zz)
        msg =bot.edit_message_reply_markup(c.message.chat.id,bib,reply_markup=keyboard)         		
        return	
		
    if c.data[-2:]=='dc':
		
        channel_id=c.data[:-2]
        keyboard = types.InlineKeyboardMarkup(row_width=2)
        keyboard.add(
                    types.InlineKeyboardButton(text=lengstr(ll,43),callback_data=channel_id+'dd'),
                    types.InlineKeyboardButton(text=lengstr(ll,44),callback_data=channel_id+'*'))
        msg =bot.edit_message_text(chat_id=c.message.chat.id, message_id=bib, text=lengstr(ll,45),reply_markup=keyboard)        		
        return	
    if c.data[-2:]=='dd':
		
        channel_id=c.data[:-2]
        delete_channel(channel_id)
        msg =bot.edit_message_text(chat_id=c.message.chat.id, message_id=bib, text=lengstr(ll,46))         		
        return	
    if c.data[-2:]=='ps':
		
        channel_id=c.data[:-2]
        change_option_channel(5,0,channel_id) 
        zz=chek_channel_options(channel_id)
        keyboard=option_keyboard(zz)
        msg =bot.edit_message_reply_markup(c.message.chat.id,bib,reply_markup=keyboard)         		
        return	

    if c.data[-2:]=='rp':
	
        channel_id=c.data[:-2]
        change_option_channel(6,0,channel_id) 
        zz=chek_channel_options(channel_id)
        keyboard=option_keyboard(zz)
        msg =bot.edit_message_reply_markup(c.message.chat.id,bib,reply_markup=keyboard)         		
        return	

    if c.data[-2:]=='mu':
	
        channel_id=c.data[:-2]
        change_option_channel(7,0,channel_id) 
        zz=chek_channel_options(channel_id)
        keyboard=option_keyboard(zz)
        msg =bot.edit_message_reply_markup(c.message.chat.id,bib,reply_markup=keyboard)         		
        return	
    if c.data[:4]=='tpin':
        for i in range(0,len(c.data)):
            if c.data[i]==':':
                channel_id=int(c.data[i+1:])
                belt=int(c.data[4:i])
        zz=chek_channel_options(channel_id)
        change_belt_t(channel_id,belt)				
        keyboard=option_keyboard(zz)      
        msg = bot.edit_message_text(chat_id=c.message.chat.id, message_id=bib, text=lengstr(ll,54),reply_markup=keyboard) 
        return
############################### nastroika chasovogo poiasa
    if c.data[-2:]=='tp':
        channel_id=int(c.data[:-2])
        keyboard = types.InlineKeyboardMarkup(row_width=4)
        for i in range(0,6):
            keyboard.add(*[types.InlineKeyboardButton(text=str(name),callback_data='tpin'+str(name)+':'+str(channel_id)) for name in [-12+(4*i)+1,-12+(4*i)+2,-12+(4*i)+3,-12+(4*i)+4]])
        keyboard.add(*[types.InlineKeyboardButton(text=name,callback_data=str(channel_id)+'*') for name in [lengstr(ll,32)]])
        msg = bot.edit_message_text(chat_id=c.message.chat.id, message_id=bib, text=lengstr(ll,53),reply_markup=keyboard)       		
        return		
############################### просмотр отложенного поста 
    if c.data[-7:]=='otsmotr':	
        try:
         gn=reverse_time_id.index(c.message.chat.id)
         reverse_time_id.remove(c.message.chat.id) 
         del reverse_time.append[gn]
        except Exception:
         ff=0		
        for i in range(0,len(c.data)):
            if c.data[i]==':':
                channel_id=int(c.data[:i])
                pst_id=int(c.data[i+1:-7])	
                numb=pst_id
        fg=channel_post_thread(channel_id)
        obj=pickle.loads(fg[pst_id][1])
        channel_id=obj.channel_id
        post_id=obj.post_id
        msg = bot.delete_message(c.message.chat.id, bib)
        keyboard = types.InlineKeyboardMarkup(row_width=2)
        keyboard.add(types.InlineKeyboardButton(text=lengstr(ll,74),callback_data=str(channel_id)+':'+str(numb)+'otred'),
        types.InlineKeyboardButton(text=lengstr(ll,75),callback_data=str(channel_id)+':'+str(numb)+'otdel'),
        types.InlineKeyboardButton(text=lengstr(ll,76),callback_data=str(channel_id)+':'+str(numb)+'otopl'),
        types.InlineKeyboardButton(text=lengstr(ll,77),callback_data=str(channel_id)+':'+str(numb)+'ottim'),
        types.InlineKeyboardButton(text=lengstr(ll,78),callback_data=str(channel_id)+'ot'))
        if 	obj.document_type=='':				
                    try:				
                      msg =bot.send_message(c.message.chat.id.text,reply_markup=keyboard,parse_mode='Markdown')
                    except Exception:
                      msg =bot.send_message(c.message.chat.id, obj.text,reply_markup=keyboard)
        if  obj.document_type=='photo':
            msg =bot.send_photo(c.message.chat.id, obj.document,caption=obj.text,reply_markup=keyboard)                    
        if  obj.document_type=='audio':
            msg =bot.send_audio(c.message.chat.id, obj.document,caption=obj.text,reply_markup=keyboard)
        if  obj.document_type=='video':
            msg =bot.send_video(c.message.chat.id, obj.document,caption=obj.text,reply_markup=keyboard)
        if  obj.document_type=='document':
            msg =bot.send_document(c.message.chat.id, obj.document,caption=obj.text,reply_markup=keyboard)
        if  obj.document_type=='voice':
            msg =bot.send_voice(c.message.chat.id, obj.document,reply_markup=keyboard) 
   
        return	
############################# izmenenie vremeni
    if c.data[-5:]=='ottim':	
                for i in range(0,len(c.data)):
                    if c.data[i]==':':
                        channel_id=int(c.data[:i])
                        pst_id=int(c.data[i+1:-5])	
                        numb=pst_id
                fg=channel_post_thread(channel_id)
                obj=pickle.loads(fg[pst_id][1])
                reverse_time_id.append(c.message.chat.id) 
                reverse_time.append(str(channel_id)+':'+str(pst_id))				
                keyboard = types.InlineKeyboardMarkup()  
                keyboard.add(*[types.InlineKeyboardButton(text=name,callback_data=str(channel_id)+':'+str(numb)+'otsmotr') for name in [lengstr(ll,32)]])
                try:				
                    msg =bot.edit_message_text(chat_id=c.message.chat.id, message_id=bib, text=lengstr(ll,33),reply_markup=keyboard)
                except Exception:
                    msg =bot.delete_message(chat_id=c.message.chat.id, message_id=bib)
                    msg =bot.send_message(chat_id=c.message.chat.id, text=lengstr(ll,33),reply_markup=keyboard)					
                return
############################# prosmotr otlojenogo posta
    if c.data[-5:]=='otred':	
                for i in range(0,len(c.data)):
                    if c.data[i]==':':
                        channel_id=int(c.data[:i])
                        pst_id=int(c.data[i+1:-5])	
                        numb=pst_id
                fg=channel_post_thread(channel_id)
                obj=pickle.loads(fg[pst_id][1])
                channel_id=obj.channel_id
                channel=chek_channel_options(channel_id)				
                print(obj.channel_id)
                keyboard = types.InlineKeyboardMarkup()
                for z in range(0,len(obj.buttons_url)):
                    print(obj.buttons[z],obj.buttons_url[z])
                    url_button = types.InlineKeyboardButton(text=obj.buttons[z], url=obj.buttons_url[z])
                    keyboard.add(url_button)
                if len(obj.reactions)>0:
                    keyboard.add(*[types.InlineKeyboardButton(text=name+':0',callback_data='‰') for name in obj.reactions])   
                elif channel[5]>0:
                    if len(channel[0])>0:
                        arreak=[]
                        for sv in range(0,len(channel[0])):
                            arreak.append(channel[0][sv])								
                        keyboard.add(*[types.InlineKeyboardButton(text=name+':0',callback_data='‰') for name in arreak])   
                if 	obj.mut==1:
                    mut=True
                elif obj.mut==-1 and channel[6]==1:
                    mut=True
                else:
                    mut=False	
                if channel[4]==1:
                    mut_ssilka=True
                else:
                    mut_ssilka=False
                keyboard.add(*[types.InlineKeyboardButton(text=name,callback_data=str(channel_id)+':'+str(numb)+'otsmotr') for name in [lengstr(ll,32)]])
                if 	obj.document_type=='':				
                    try:				
                      msg =bot.send_message(c.message.chat.id, obj.text,reply_markup=keyboard,disable_notification=mut,disable_web_page_preview=mut_ssilka,parse_mode='Markdown')
                    except Exception:
                      msg =bot.send_message(c.message.chat.id, obj.text,reply_markup=keyboard,disable_notification=mut,disable_web_page_preview=mut_ssilka)
                if  obj.document_type=='photo':
                    msg =bot.send_photo(c.message.chat.id, obj.document,caption=obj.text,reply_markup=keyboard,disable_notification=mut)                    
                if  obj.document_type=='audio':
                    msg =bot.send_audio(c.message.chat.id, obj.document,caption=obj.text,reply_markup=keyboard,disable_notification=mut)
                if  obj.document_type=='video':
                    msg =bot.send_video(c.message.chat.id, obj.document,caption=obj.text,reply_markup=keyboard,disable_notification=mut)
                if  obj.document_type=='document':
                    msg =bot.send_document(c.message.chat.id, obj.document,caption=obj.text,reply_markup=keyboard,disable_notification=mut) 
                if  obj.document_type=='voice':
                    msg =bot.send_voice(c.message.chat.id, obj.document,reply_markup=keyboard,disable_notification=mut) 				
                msg = bot.delete_message(chat_id=c.message.chat.id, message_id=bib)	
                return	
############################### udalenie posta iz otlojenih 
    if c.data[-5:]=='otdel':	
                for i in range(0,len(c.data)):
                    if c.data[i]==':':
                        channel_id=int(c.data[:i])
                        pst_id=int(c.data[i+1:-5])	
                        numb=pst_id
                fg=channel_post_thread(channel_id)
                obj=pickle.loads(fg[pst_id][1])
                delete_from_fors(channel_id,fg[pst_id][1])
                keyboard = types.InlineKeyboardMarkup()  
                keyboard.add(  types.InlineKeyboardButton(text=lengstr(ll,78),callback_data=str(channel_id)+'ot'))	
                try:				
                    msg =bot.edit_message_text(chat_id=c.message.chat.id, message_id=bib, text=lengstr(ll,49),reply_markup=keyboard)
                except Exception:
                    msg =bot.delete_message(chat_id=c.message.chat.id, message_id=bib)
                    msg =bot.send_message(chat_id=c.message.chat.id, text=lengstr(ll,49),reply_markup=keyboard)					
                return
############################### opublikovat nemedlenno
    if c.data[-5:]=='otopl':	
                for i in range(0,len(c.data)):
                    if c.data[i]==':':
                        channel_id=int(c.data[:i])
                        pst_id=int(c.data[i+1:-5])	
                        numb=pst_id
                fg=channel_post_thread(channel_id)
                obj=pickle.loads(fg[pst_id][1])
                delete_from_fors(channel_id,fg[pst_id][1])
                channel_id=obj.channel_id
                channel=chek_channel_options(channel_id)				
                keyboard = types.InlineKeyboardMarkup()
                for z in range(0,len(obj.buttons_url)):
                    print(obj.buttons[z],obj.buttons_url[z])
                    url_button = types.InlineKeyboardButton(text=obj.buttons[z], url=obj.buttons_url[z])
                    keyboard.add(url_button)
                if len(obj.reactions)>0:
                    keyboard.add(*[types.InlineKeyboardButton(text=name+':0',callback_data='roact'+str(obj.reactions.index(name))+':'+name) for name in obj.reactions])   
                elif channel[5]>0:
                    if len(channel[0])>0:
                        arreak=[]
                        for sv in range(0,len(channel[0])):
                            arreak.append(channel[0][sv])								
                        keyboard.add(*[types.InlineKeyboardButton(text=name+':0',callback_data='roact'+str(arreak.index(name))+':'+name) for name in arreak])   
                if 	obj.mut==1:
                    mut=True
                elif obj.mut==-1 and channel[6]==1:
                    mut=True
                else:
                    mut=False	
                if channel[4]==1:
                    mut_ssilka=True
                else:
                    mut_ssilka=False
                if 	obj.document_type=='':				
                    try:				
                      msg =bot.send_message(obj.channel_id, obj.text,reply_markup=keyboard,disable_notification=mut,disable_web_page_preview=mut_ssilka,parse_mode='Markdown')
                    except Exception:
                      msg =bot.send_message(obj.channel_id, obj.text,reply_markup=keyboard,disable_notification=mut,disable_web_page_preview=mut_ssilka)
                if  obj.document_type=='photo':
                    msg =bot.send_photo(obj.channel_id, obj.document,caption=obj.text,reply_markup=keyboard,disable_notification=mut)                    
                if  obj.document_type=='audio':
                    msg =bot.send_audio(obj.channel_id, obj.document,caption=obj.text,reply_markup=keyboard,disable_notification=mut)
                if  obj.document_type=='video':
                    msg =bot.send_video(obj.channel_id, obj.document,caption=obj.text,reply_markup=keyboard,disable_notification=mut)
                if  obj.document_type=='document':
                    msg =bot.send_document(obj.channel_id, obj.document,caption=obj.text,reply_markup=keyboard,disable_notification=mut)
                if  obj.document_type=='voice':
                    msg =bot.send_voice(obj.channel_id, obj.document,reply_markup=keyboard,disable_notification=mut) 
   
                joy=msg.message_id
                if len(obj.reactions)>0:
                    add_reactions(obj.channel_id,joy,obj.reactions)
                elif channel[5]>0:
                    if len(channel[0])>0:
                        add_reactions(obj.channel_id,joy,arreak)					
                if obj.delete_time>0:
                    tim=[0,3600,10800,21600,43200,86400,172800,604800,1209600,2592000]
                    add_delete_func(joy,obj.channel_id,tim[obj.delete_time])  
                elif channel[3]==1:
                    if channel[2]>0:
                        add_delete_func(joy,obj.channel_id,channel[2])  					
                if 	obj.pin==1:
                    msg=bot.pin_chat_message(obj.channel_id,joy,disable_notification=mut)
                keyboard = types.InlineKeyboardMarkup()  
                keyboard.add(  types.InlineKeyboardButton(text=lengstr(ll,78),callback_data=str(channel_id)+'ot'))				
                try:				
                    msg =bot.edit_message_text(chat_id=c.message.chat.id, message_id=bib, text=lengstr(ll,49),reply_markup=keyboard)
                except Exception:
                    msg =bot.delete_message(chat_id=c.message.chat.id, message_id=bib)
                    msg =bot.send_message(chat_id=c.message.chat.id, text=lengstr(ll,49),reply_markup=keyboard)					
                return               				
###############################
    if c.data[-2:]=='ot':
        channel_id=int(c.data[:-2])
        fg=channel_post_thread(channel_id)
        keyboard = types.InlineKeyboardMarkup(row_width=1)
        for i in range(0,len(fg)):
            time_pst=datetime.fromtimestamp(fg[i][0]).strftime('%Y-%m-%d %H:%M:%S')
            numb=str(i)
            keyboard.add(*[types.InlineKeyboardButton(text=time_pst,callback_data=str(channel_id)+':'+str(numb)+'otsmotr') for name in [lengstr(ll,32)]])		
        keyboard.add(*[types.InlineKeyboardButton(text=name,callback_data=str(channel_id)+'*') for name in [lengstr(ll,32)]])
        msg = bot.edit_message_text(chat_id=c.message.chat.id, message_id=bib, text=lengstr(ll,48),reply_markup=keyboard)
        return 		
############################### izmenenie reakci v kanale 
    if c.data[-2:]=='ra':
        channel_id=c.data[:-2]
        keyboard = types.InlineKeyboardMarkup(row_width=1)
        keyboard.add(*[types.InlineKeyboardButton(text=name,callback_data='adsh'+name+':'+channel_id) for name in ['👍👎','😊😄😒','❤😐💔🤢','😄😊😔😱😡']])
        keyboard.add(*[types.InlineKeyboardButton(text=name,callback_data=str(channel_id)+'*') for name in [lengstr(ll,32)]])		
        msg = bot.edit_message_text(chat_id=c.message.chat.id, message_id=bib, text=lengstr(ll,36),reply_markup=keyboard)
        return	
    if 'adsh' in c.data:
        for i in range(0,len(c.data)):
            if c.data[i]==':':
                reac=c.data[4:i]
                channel_id=c.data[i+1:]
        change_option_channel(1,reac,channel_id) 
        zz=chek_channel_options(channel_id)
        keyboard=option_keyboard(zz)      
        msg = bot.edit_message_text(chat_id=c.message.chat.id, message_id=bib, text=lengstr(ll,79),reply_markup=keyboard)          
        return 	

############################### Nastroika udaleni v opciah kanala
    if c.data[-2:]=='dl':
        channel_id=c.data[:-2]
        keyboard = types.InlineKeyboardMarkup(row_width=3)
        keyboard.add(
                    types.InlineKeyboardButton(text=lengstr(ll,23),callback_data='ady'+str(1)+':'+channel_id),
                    types.InlineKeyboardButton(text=lengstr(ll,24),callback_data='ady'+str(2)+':'+channel_id),
                    types.InlineKeyboardButton(text=lengstr(ll,25),callback_data='ady'+str(3)+':'+channel_id),
                    types.InlineKeyboardButton(text=lengstr(ll,26),callback_data='ady'+str(4)+':'+channel_id),
                    types.InlineKeyboardButton(text=lengstr(ll,27),callback_data='ady'+str(5)+':'+channel_id),
                    types.InlineKeyboardButton(text=lengstr(ll,28),callback_data='ady'+str(6)+':'+channel_id),
                    types.InlineKeyboardButton(text=lengstr(ll,29),callback_data='ady'+str(7)+':'+channel_id),
                    types.InlineKeyboardButton(text=lengstr(ll,30),callback_data='ady'+str(8)+':'+channel_id),
                    types.InlineKeyboardButton(text=lengstr(ll,31),callback_data='ady'+str(9)+':'+channel_id))	
        keyboard.add(*[types.InlineKeyboardButton(text=name,callback_data=str(channel_id)+'*') for name in [lengstr(ll,32)]])						
        msg = bot.edit_message_text(chat_id=c.message.chat.id, message_id=bib, text=lengstr(ll,22),reply_markup=keyboard)
        return   
    if 'ady' in c.data:
        for i in range(0,len(c.data)):
            if c.data[i]==':':
                reac=c.data[3:i]
                channel_id=c.data[i+1:]
        change_option_channel(3,reac,channel_id) 
        zz=chek_channel_options(channel_id)
        keyboard=option_keyboard(zz)      
        msg = bot.edit_message_text(chat_id=c.message.chat.id, message_id=bib, text=lengstr(ll,79),reply_markup=keyboard)          
        return 	    	
############################### dobavlenie posta v kanale		
    if c.data[-1]=='%':
        obj=post()
        obj.user_id=c.message.chat.id
        obj.channel_id=int(c.data[:-1])
        obj.post_id=int(posts) 
        add_post_cok(c.message.chat.id,obj.channel_id,obj.post_id,obj)
        posts+=1
        output = open('posts.pkl', 'wb')
        pickle.dump(posts, output, 2)
        output.close() 
        keyboard = types.InlineKeyboardMarkup(row_width=1)
        keyboard.add(*[types.InlineKeyboardButton(text=name,callback_data=c.data[:-1]+'!') for name in [lengstr(ll,80)]])
        msg = bot.edit_message_text(chat_id=c.message.chat.id, message_id=bib, text=lengstr(ll,12),reply_markup=keyboard,parse_mode='HTML') 
        return
############################## otmena dobavlenia posta v kanale		
    if c.data[-1]=='!':
        delete_cok(c.message.chat.id)
        channel_id=c.data[:-1]
        chek_channel_options(channel_id)
        channel_name=chan_name_func(channel_id)
        msg = bot.get_chat(channel_id)
        print(msg.id)
        tmbl=chek_timeb(channel_id)
        print(tmbl)
        print('@'+msg.username)
        usrnm='@'+msg.username
        usid=str(msg.id)
        print(msg)
        txt=lengstr(ll,69)+': ['+channel_name[0][0]+']('+usrnm+')\nChannelID: '+usid[4:]

        keyboard = types.InlineKeyboardMarkup(row_width=2)
        keyboard.add(types.InlineKeyboardButton(text=lengstr(ll,62),callback_data=str(channel_id)+'%'),
                     types.InlineKeyboardButton(text=lengstr(ll,63),callback_data=str(channel_id)+'@'))
        keyboard.add(types.InlineKeyboardButton(text=lengstr(ll,64),callback_data=str(channel_id)+'&'),
                     types.InlineKeyboardButton(text=lengstr(ll,65),callback_data=str(channel_id)+'rekl'))
        keyboard.add(types.InlineKeyboardButton(text=lengstr(ll,66),callback_data=str(channel_id)+'stat'),
                     types.InlineKeyboardButton(text=lengstr(ll,67),callback_data=str(channel_id)+'*'))
        keyboard.add(*[types.InlineKeyboardButton(text=name,callback_data=str(0)+'off') for name in [lengstr(ll,68)]])
        keyboard.add(*[types.InlineKeyboardButton(text=name,callback_data=str(0)+'Gur') for name in [lengstr(ll,61)]])	
        try:				
            msg =bot.edit_message_text(chat_id=c.message.chat.id, message_id=bib, text=txt,reply_markup=keyboard,parse_mode='Markdown',disable_web_page_preview=True)	
        except Exception:
                    msg = bot.delete_message(chat_id=c.message.chat.id, message_id=bib)
                    msg = bot.send_message(chat_id=c.message.chat.id, text=txt,reply_markup=keyboard,parsr_mode='Markdown',disable_web_page_preview=True)  
        return
################################# nastroiki kanala
    if c.data[-1]=='*':
        channel_id=int(c.data[:-1])	
        zz=chek_channel_options(channel_id)
        keyboard=option_keyboard(zz)
        msg = bot.edit_message_text(chat_id=c.message.chat.id, message_id=bib, text=lengstr(ll,79),reply_markup=keyboard)          
        return      
################################# spisok publikaci
    if c.data[-1]=='@':
        channel_id=int(c.data[:-1])	
        pos=post_in_channel(channel_id)
        keyboard = types.InlineKeyboardMarkup(row_width=1)
        for i in range(0,len(pos)):
            dd=pos[i][4]
            keyboard.add(*[types.InlineKeyboardButton(text=name,callback_data='selpost'+str(pos[i][0])) for name in [str(dd)]]) 
        if len(pos)>0:
            keyboard.add(*[types.InlineKeyboardButton(text=name,callback_data=c.data[:-1]+'rekl') for name in [ lengstr(ll,116)]])
        keyboard.add(*[types.InlineKeyboardButton(text=name,callback_data=c.data[:-1]+'!') for name in [lengstr(ll,80)]]) 
        if len(pos)>0:		
            try:			
                msg = bot.edit_message_text(chat_id=c.message.chat.id, message_id=bib, text=lengstr(ll,81)+str(len(pos)),reply_markup=keyboard)
            except Exception:
                msg = bot.delete_message(chat_id=c.message.chat.id, message_id=bib) 
                msg = bot.send_message(chat_id=c.message.chat.id, text=lengstr(ll,81)+str(len(pos)),reply_markup=keyboard)	
        else:
            try:			
                msg = bot.edit_message_text(chat_id=c.message.chat.id, message_id=bib, text=lengstr(ll,82),reply_markup=keyboard)
            except Exception:
                msg = bot.delete_message(chat_id=c.message.chat.id, message_id=bib) 
                msg = bot.send_message(chat_id=c.message.chat.id, text=lengstr(ll,82),reply_markup=keyboard)			
        return		
################################# Просмотр сохраненного поста
    if 'selpost' in c.data:
        inm=c.data[7:]
        obj=select_from_saved(inm)
        channel_id=obj.channel_id
        post_id=obj.post_id
        msg = bot.delete_message(c.message.chat.id, bib)
        keyboard = types.InlineKeyboardMarkup(row_width=2)
        keyboard.add(types.InlineKeyboardButton(text=lengstr(ll,83),callback_data='selred'+str(post_id)),
        types.InlineKeyboardButton(text=lengstr(ll,75),callback_data=str(post_id)+'l'),
        types.InlineKeyboardButton(text=lengstr(ll,84),callback_data='selopl'+str(post_id)),
        types.InlineKeyboardButton(text=lengstr(ll,85),callback_data=str(post_id)+'l'),
        types.InlineKeyboardButton(text=lengstr(ll,78),callback_data=str(channel_id)+'@'))
        if 	obj.document_type=='':				
                    try:				
                      msg =bot.send_message(c.message.chat.id.text,reply_markup=keyboard,parse_mode='Markdown')
                    except Exception:
                      msg =bot.send_message(c.message.chat.id, obj.text,reply_markup=keyboard)
        if  obj.document_type=='photo':
            msg =bot.send_photo(c.message.chat.id, obj.document,caption=obj.text,reply_markup=keyboard)                    
        if  obj.document_type=='audio':
            msg =bot.send_audio(c.message.chat.id, obj.document,caption=obj.text,reply_markup=keyboard)
        if  obj.document_type=='video':
            msg =bot.send_video(c.message.chat.id, obj.document,caption=obj.text,reply_markup=keyboard)
        if  obj.document_type=='document':
            msg =bot.send_document(c.message.chat.id, obj.document,caption=obj.text,reply_markup=keyboard)
        if  obj.document_type=='voice':
            msg =bot.send_voice(c.message.chat.id, obj.document,reply_markup=keyboard) 
   
        return
################################
    if 'selopl' in c.data:
                inm=c.data[6:]
                obj=select_from_saved(inm)
                channel_id=obj.channel_id
                channel=chek_channel_options(channel_id)				
                if obj.forse_time>0: 
                    inputt = pickle.dumps(obj)
                    add_forse_func(obj.forse_time,inputt,channel_id)
                    msg =bot.answer_callback_query(c.id,lengstr(ll,35)) 
                    return
                print(obj.channel_id)
                keyboard = types.InlineKeyboardMarkup()
                for z in range(0,len(obj.buttons_url)):
                    print(obj.buttons[z],obj.buttons_url[z])
                    url_button = types.InlineKeyboardButton(text=obj.buttons[z], url=obj.buttons_url[z])
                    keyboard.add(url_button)
                if len(obj.reactions)>0:
                    keyboard.add(*[types.InlineKeyboardButton(text=name+':0',callback_data='roact'+str(obj.reactions.index(name))+':'+name) for name in obj.reactions])   
                elif channel[5]>0:
                    if len(channel[0])>0:
                        arreak=[]
                        for sv in range(0,len(channel[0])):
                            arreak.append(channel[0][sv])								
                        keyboard.add(*[types.InlineKeyboardButton(text=name+':0',callback_data='roact'+str(arreak.index(name))+':'+name) for name in arreak])   
                if 	obj.mut==1:
                    mut=True
                elif obj.mut==-1 and channel[6]==1:
                    mut=True
                else:
                    mut=False	
                if channel[4]==1:
                    mut_ssilka=True
                else:
                    mut_ssilka=False
                if 	obj.document_type=='':				
                    try:				
                      msg =bot.send_message(obj.channel_id, obj.text,reply_markup=keyboard,disable_notification=mut,disable_web_page_preview=mut_ssilka,parse_mode='Markdown')
                    except Exception:
                      msg =bot.send_message(obj.channel_id, obj.text,reply_markup=keyboard,disable_notification=mut,disable_web_page_preview=mut_ssilka)
                if  obj.document_type=='photo':
                    msg =bot.send_photo(obj.channel_id, obj.document,caption=obj.text,reply_markup=keyboard,disable_notification=mut)                    
                if  obj.document_type=='audio':
                    msg =bot.send_audio(obj.channel_id, obj.document,caption=obj.text,reply_markup=keyboard,disable_notification=mut)
                if  obj.document_type=='video':
                    msg =bot.send_video(obj.channel_id, obj.document,caption=obj.text,reply_markup=keyboard,disable_notification=mut)
                if  obj.document_type=='document':
                    msg =bot.send_document(obj.channel_id, obj.document,caption=obj.text,reply_markup=keyboard,disable_notification=mut)
                if  obj.document_type=='voice':
                    msg =bot.send_voice(obj.channel_id, obj.document,reply_markup=keyboard,disable_notification=mut) 
   
                joy=msg.message_id
                if len(obj.reactions)>0:
                    add_reactions(obj.channel_id,joy,obj.reactions)
                elif channel[5]>0:
                    if len(channel[0])>0:
                        add_reactions(obj.channel_id,joy,arreak)					
                if obj.delete_time>0:
                    tim=[0,3600,10800,21600,43200,86400,172800,604800,1209600,2592000]
                    add_delete_func(joy,obj.channel_id,tim[obj.delete_time])  
                elif channel[3]==1:
                    if channel[2]>0:
                        add_delete_func(joy,obj.channel_id,channel[2])  					
                if 	obj.pin==1:
                    msg=bot.pin_chat_message(obj.channel_id,joy,disable_notification=mut)	               
                msg =bot.answer_callback_query(c.id,lengstr(ll,14)) 
                return						
################################# redaktirovanie sohranennogo posta
    if 'selred'	in c.data:
        inm=c.data[6:]
        obj=select_from_saved(inm) 
        add_post_cok(c.message.chat.id,obj.channel_id,obj.post_id,obj)        
        keyboard=kukoard(obj.post_id,obj.channel_id,c.message.chat.id)
        msg = bot.delete_message(c.message.chat.id,bib)
        msg = bot.send_message(c.message.chat.id,lengstr(ll,13),reply_markup=keyboard) 
        return	
################################# pered postom
    if c.data[-2:]=='oo':
                obj_post=find_post_cok(c.message.chat.id) 
                obj=obj_post[1]	
                post_id=obj.post_id
                channel_id=obj.channel_id
                opis=''
                msg = bot.delete_message(c.message.chat.id,bib)
                keyboard = types.InlineKeyboardMarkup()
                keyboard.add(types.InlineKeyboardButton(text=lengstr(ll,84),callback_data=str(post_id)+'o'))                
                keyboard.add(*[types.InlineKeyboardButton(text=name,callback_data='<') for name in [lengstr(ll,32)]])	 
                keyboard.add(types.InlineKeyboardButton(text=lengstr(ll,86),callback_data=str(channel_id)+'!'))				
                channel_list=chek_user_channels(c.message.chat.id)
                for i in range (0,len(channel_list[0])):
                    if channel_list[0][i]==channel_id:
                        opis= lengstr(ll,117)+channel_list[1][i]+'\n'
                if obj.text==None:
                    obj.text=opis
                else:
                    obj.text=str(opis)+str(obj.text)
                if 	obj.document_type=='':				
                    try:				
                      msg =bot.send_message(c.message.chat.id, obj.text,reply_markup=keyboard,parse_mode='Markdown')
                    except Exception:
                      msg =bot.send_message(c.message.chat.id, obj.text,reply_markup=keyboard)
                if  obj.document_type=='photo':
                    msg =bot.send_photo(c.message.chat.id, obj.document,caption=obj.text,reply_markup=keyboard)                    
                if  obj.document_type=='audio':
                    msg =bot.send_audio(c.message.chat.id, obj.document,caption=obj.text,reply_markup=keyboard)
                if  obj.document_type=='video':
                    msg =bot.send_video(c.message.chat.id, obj.document,caption=opis+obj.text,reply_markup=keyboard)
                if  obj.document_type=='document':
                    msg =bot.send_document(c.message.chat.id, obj.document,caption=obj.text,reply_markup=keyboard) 	
                if  obj.document_type=='voice':
                    msg =bot.send_voice(c.message.chat.id, obj.document,reply_markup=keyboard) 
                return                        	
################################# publikacia posta
    if c.data[-1]=='o':
                obj_post=find_post_cok(c.message.chat.id) 
                obj=obj_post[1]	
                channel_id=obj.channel_id
                channel=chek_channel_options(channel_id)				
                if obj.forse_time>0: 
                    inputt = pickle.dumps(obj)
                    dptime=int(chek_timeb(channel_id))
                    frstm=obj.forse_time-(server_time*60*60)+(dptime*60*60)
                    add_forse_func(frstm,inputt,channel_id)
                    msg =bot.answer_callback_query(c.id,lengstr(ll,35)) 
                    return
                print(obj.channel_id)
                keyboard = types.InlineKeyboardMarkup()
                for z in range(0,len(obj.buttons_url)):
                    print(obj.buttons[z],obj.buttons_url[z])
                    url_button = types.InlineKeyboardButton(text=obj.buttons[z], url=obj.buttons_url[z])
                    keyboard.add(url_button)
                if len(obj.reactions)>0:
                    keyboard.add(*[types.InlineKeyboardButton(text=name+':0',callback_data='roact'+str(obj.reactions.index(name))+':'+name) for name in obj.reactions])   
                elif channel[5]>0:
                    if len(channel[0])>0:
                        arreak=[]
                        for sv in range(0,len(channel[0])):
                            arreak.append(channel[0][sv])								
                        keyboard.add(*[types.InlineKeyboardButton(text=name+':0',callback_data='roact'+str(arreak.index(name))+':'+name) for name in arreak])   
                if 	obj.mut==1:
                    mut=True
                elif obj.mut==-1 and channel[6]==1:
                    mut=True
                else:
                    mut=False	
                if channel[4]==1:
                    mut_ssilka=True
                else:
                    mut_ssilka=False
                if 	obj.document_type=='':				
                    try:				
                      msg =bot.send_message(obj.channel_id, obj.text,reply_markup=keyboard,disable_notification=mut,disable_web_page_preview=mut_ssilka,parse_mode='Markdown')
                    except Exception:
                      msg =bot.send_message(obj.channel_id, obj.text,reply_markup=keyboard,disable_notification=mut,disable_web_page_preview=mut_ssilka)
                if  obj.document_type=='photo':
                    msg =bot.send_photo(obj.channel_id, obj.document,caption=obj.text,reply_markup=keyboard,disable_notification=mut)                    
                if  obj.document_type=='audio':
                    msg =bot.send_audio(obj.channel_id, obj.document,caption=obj.text,reply_markup=keyboard,disable_notification=mut)
                if  obj.document_type=='video':
                    msg =bot.send_video(obj.channel_id, obj.document,caption=obj.text,reply_markup=keyboard,disable_notification=mut)
                if  obj.document_type=='document':
                    msg =bot.send_document(obj.channel_id, obj.document,caption=obj.text,reply_markup=keyboard,disable_notification=mut)
                if  obj.document_type=='voice':
                    msg =bot.send_voice(obj.channel_id, obj.document,reply_markup=keyboard,disable_notification=mut) 
   
                joy=msg.message_id
                if len(obj.reactions)>0:
                    add_reactions(obj.channel_id,joy,obj.reactions)
                elif channel[5]>0:
                    if len(channel[0])>0:
                        add_reactions(obj.channel_id,joy,arreak)					
                if obj.delete_time>0:
                    tim=[0,3600,10800,21600,43200,86400,172800,604800,1209600,2592000]
                    add_delete_func(joy,obj.channel_id,tim[obj.delete_time])  
                elif channel[3]==1:
                    if channel[2]>0:
                        add_delete_func(joy,obj.channel_id,channel[2])  					
                if 	obj.pin==1:
                    msg=bot.pin_chat_message(obj.channel_id,joy,disable_notification=mut)	               
                msg =bot.answer_callback_query(c.id,lengstr(ll,14)) 
                return				
################################# dobavlenie knopok k postu
    if c.data[-1]=='k':
                obj_post=find_post_cok(c.message.chat.id) 
                obj=obj_post[1]	
                obj.add_type='buttons'
                dump_post_cok(c.message.chat.id,obj)
                keyboard = types.InlineKeyboardMarkup(row_width=1)
                keyboard.add(*[types.InlineKeyboardButton(text=name,callback_data='<') for name in [lengstr(ll,32)]])	
                try:
                    msg = bot.edit_message_text(chat_id=c.message.chat.id, message_id=bib, text=lengstr(ll,15),disable_web_page_preview=True,reply_markup=keyboard)
                except Exception:
                    msg=bot.delete_message(c.message.chat.id,bib)
                    msg = bot.send_message(chat_id=c.message.chat.id, text=lengstr(ll,15),disable_web_page_preview=True,reply_markup=keyboard)
                return		
################################# vrema publikacii
    if c.data[-1]=='v':
                obj_post=find_post_cok(c.message.chat.id) 
                obj=obj_post[1]	
                obj.add_type='vrema'
                dump_post_cok(c.message.chat.id,obj)
                keyboard = types.InlineKeyboardMarkup(row_width=1)
                keyboard.add(*[types.InlineKeyboardButton(text=name,callback_data='<') for name in [lengstr(ll,32)]])	
                try:
                    msg = bot.edit_message_text(chat_id=c.message.chat.id, message_id=bib, text=lengstr(ll,33),disable_web_page_preview=True,reply_markup=keyboard)
                except Exception:
                    msg=bot.delete_message(c.message.chat.id,bib)
                    msg = bot.send_message(chat_id=c.message.chat.id, text=lengstr(ll,33),disable_web_page_preview=True,reply_markup=keyboard)
                return						
################################# dobavlenie reakci				
    if c.data[-1]=='r':
                obj_post=find_post_cok(c.message.chat.id) 
                obj=obj_post[1]	
                obj.add_type='reactions'
                dump_post_cok(c.message.chat.id,obj)
                keyboard = types.InlineKeyboardMarkup(row_width=1)
                keyboard.add(*[types.InlineKeyboardButton(text=name,callback_data='addsmile'+name) for name in ['👍👎','😊😄😒','❤😐💔🤢','😄😊😔😱😡']]) 
                keyboard.add(*[types.InlineKeyboardButton(text=name,callback_data='<') for name in [lengstr(ll,32)]])	
                try:				
                    msg = bot.edit_message_text(chat_id=c.message.chat.id, message_id=bib, text=lengstr(ll,17),reply_markup=keyboard)
                except Exception:
                    msg=bot.delete_message(c.message.chat.id,bib)
                    msg = bot.send_message(chat_id=c.message.chat.id, message_id=bib, text=lengstr(ll,32),disable_web_page_preview=True,reply_markup=keyboard)
                return	
################################ dobavlenie reakcii cherez smaili
    if 'addsmile' in c.data:
                obj_post=find_post_cok(c.message.chat.id) 
                obj=obj_post[1]	
                if 	obj.add_type=='reactions':
                    smiles=c.data[8:]
                    print(len(smiles))
                    obj.reactions=[]
                    for z in range(0,len(smiles)):
                        print(smiles[z])
                        obj.reactions.append(smiles[z])
                    post_id=obj.post_id
                    channel_id=obj.channel_id
                    keyboard = kukoard(post_id,channel_id,c.message.chat.id)
                    dump_post_cok(c.message.chat.id,obj)						
                    msg = bot.edit_message_text(chat_id=c.message.chat.id, message_id=bib, text=lengstr(ll,18),reply_markup=keyboard)
                    print(obj.reactions)		
                    return	
################################## zakrepit
    if c.data[-1]=='z':
                obj_post=find_post_cok(c.message.chat.id) 
                obj=obj_post[1]	
                if obj.pin<=0:
                    obj.pin=1
                    msg =bot.answer_callback_query(c.id,lengstr(ll,87))
                else:
                    obj.pin=0	
                    msg =bot.answer_callback_query(c.id,lengstr(ll,88))					
                post_id=obj.post_id
                channel_id=obj.channel_id
                dump_post_cok(c.message.chat.id,obj)				
                keyboard = kukoard(post_id,channel_id,c.message.chat.id)          
                msg =bot.edit_message_reply_markup(c.message.chat.id,bib,reply_markup=keyboard) 
                return	
################################## uvedomlenia
    if c.data[-1]=='u':
                obj_post=find_post_cok(c.message.chat.id) 
                obj=obj_post[1]	
                if obj.mut<=0:
                    obj.mut=1
                    msg =bot.answer_callback_query(c.id,lengstr(ll,89))
                else:
                    obj.mut=0	
                    msg =bot.answer_callback_query(c.id,lengstr(ll,90))					
                post_id=obj.post_id
                channel_id=obj.channel_id
                dump_post_cok(c.message.chat.id,obj)					
                keyboard = kukoard(post_id,channel_id,c.message.chat.id)          
                msg =bot.edit_message_reply_markup(c.message.chat.id,bib,reply_markup=keyboard) 
                return
################################# udalenie posta
    if c.data[-1]=='d':
                obj_post=find_post_cok(c.message.chat.id) 
                obj=obj_post[1]	
                keyboard = types.InlineKeyboardMarkup(row_width=3)
                keyboard.add(
                    types.InlineKeyboardButton(text=lengstr(ll,23),callback_data='addelet'+str(1)),
                    types.InlineKeyboardButton(text=lengstr(ll,24),callback_data='addelet'+str(2)),
                    types.InlineKeyboardButton(text=lengstr(ll,25),callback_data='addelet'+str(3)),
                    types.InlineKeyboardButton(text=lengstr(ll,26),callback_data='addelet'+str(4)),
                    types.InlineKeyboardButton(text=lengstr(ll,27),callback_data='addelet'+str(5)),
                    types.InlineKeyboardButton(text=lengstr(ll,28),callback_data='addelet'+str(6)),
                    types.InlineKeyboardButton(text=lengstr(ll,29),callback_data='addelet'+str(7)),
                    types.InlineKeyboardButton(text=lengstr(ll,30),callback_data='addelet'+str(8)),
                    types.InlineKeyboardButton(text=lengstr(ll,31),callback_data='addelet'+str(9)))
                keyboard.add(*[types.InlineKeyboardButton(text=name,callback_data='<') for name in [lengstr(ll,32)]])	
                try:				
                    msg = bot.edit_message_text(chat_id=c.message.chat.id, message_id=bib, text=lengstr(ll,22),reply_markup=keyboard)
                except Exception:
                    msg=bot.delete_message(c.message.chat.id,bib)
                    msg = bot.send_message(chat_id=c.message.chat.id, text=lengstr(ll,33),disable_web_page_preview=True,reply_markup=keyboard)
                return	
################################# zapis na udalenie
    if 'addelet' in c.data:
                obj_post=find_post_cok(c.message.chat.id) 
                obj=obj_post[1]	
                funcs=int(c.data[7:])
                obj.delete_time=funcs
                dump_post_cok(c.message.chat.id,obj)				
                post_id=obj.post_id
                channel_id=obj.channel_id
                keyboard = kukoard(post_id,channel_id,c.message.chat.id)						
                msg = bot.edit_message_text(chat_id=c.message.chat.id, message_id=bib, text=lengstr(ll,18),reply_markup=keyboard)	
                return				
################################# publikacia posta
    if c.data[-1]=='p':
                obj_post=find_post_cok(c.message.chat.id) 
                obj=obj_post[1]	
                channel_id=obj.channel_id
                channel=chek_channel_options(channel_id)				
                print(obj.channel_id)
                keyboard = types.InlineKeyboardMarkup()
                for z in range(0,len(obj.buttons_url)):
                    print(obj.buttons[z],obj.buttons_url[z])
                    url_button = types.InlineKeyboardButton(text=obj.buttons[z], url=obj.buttons_url[z])
                    keyboard.add(url_button)
                if len(obj.reactions)>0:
                    keyboard.add(*[types.InlineKeyboardButton(text=name+':0',callback_data='‰') for name in obj.reactions])   
                elif channel[5]>0:
                    if len(channel[0])>0:
                        arreak=[]
                        for sv in range(0,len(channel[0])):
                            arreak.append(channel[0][sv])								
                        keyboard.add(*[types.InlineKeyboardButton(text=name+':0',callback_data='‰') for name in arreak])   
                if 	obj.mut==1:
                    mut=True
                elif obj.mut==-1 and channel[6]==1:
                    mut=True
                else:
                    mut=False	
                if channel[4]==1:
                    mut_ssilka=True
                else:
                    mut_ssilka=False
                keyboard.add(*[types.InlineKeyboardButton(text=name,callback_data='<') for name in [lengstr(ll,32)]])
                if 	obj.document_type=='':				
                    try:				
                      msg =bot.send_message(c.message.chat.id, obj.text,reply_markup=keyboard,disable_notification=mut,disable_web_page_preview=mut_ssilka,parse_mode='Markdown')
                    except Exception:
                      msg =bot.send_message(c.message.chat.id, obj.text,reply_markup=keyboard,disable_notification=mut,disable_web_page_preview=mut_ssilka)
                if  obj.document_type=='photo':
                    msg =bot.send_photo(c.message.chat.id, obj.document,caption=obj.text,reply_markup=keyboard,disable_notification=mut)                    
                if  obj.document_type=='audio':
                    msg =bot.send_audio(c.message.chat.id, obj.document,caption=obj.text,reply_markup=keyboard,disable_notification=mut)
                if  obj.document_type=='video':
                    msg =bot.send_video(c.message.chat.id, obj.document,caption=obj.text,reply_markup=keyboard,disable_notification=mut)
                if  obj.document_type=='document':
                    msg =bot.send_document(c.message.chat.id, obj.document,caption=obj.text,reply_markup=keyboard,disable_notification=mut) 
                if  obj.document_type=='voice':
                    msg =bot.send_voice(c.message.chat.id, obj.document,reply_markup=keyboard,disable_notification=mut) 
                keyboard=kukoard(obj.post_id,obj.channel_id,c.message.chat.id)					
                msg = bot.delete_message(chat_id=c.message.chat.id, message_id=bib)	
                return
################################ nachat snachala
    if c.data[-1]=='n':
                delete_cok(c.message.chat.id)        
                obj=post()
                obj.user_id=c.message.chat.id
                obj.channel_id=int(c.data[:-1])
                obj.post_id=int(posts) 
                add_post_cok(c.message.chat.id,obj.channel_id,obj.post_id,obj)
                posts+=1
                output = open('posts.pkl', 'wb')
                pickle.dump(posts, output, 2)
                output.close() 				
                keyboard = types.InlineKeyboardMarkup(row_width=1)
                keyboard.add(*[types.InlineKeyboardButton(text=name,callback_data=c.data[:-1]+'!') for name in [lengstr(ll,86)]])
                msg = bot.edit_message_text(chat_id=c.message.chat.id, message_id=bib, text=lengstr(ll,12),reply_markup=keyboard) 
                return
################################### sohranenie posta
    if c.data[-1]=='s':
                obj_post=find_post_cok(c.message.chat.id) 
                obj=obj_post[1]	
                if obj.saved==1:
                        msg =bot.answer_callback_query(c.id,lengstr(ll,91))					         
                        return	                        
                obj.saved=1
                dump_post_cok(c.message.chat.id,obj)
                save_post(obj.post_id,c.message.chat.id,obj.channel_id,obj)
                msg =bot.answer_callback_query(c.id,lengstr(ll,92))					         
                return	
################################### udalenie posta
    if c.data[-1]=='l':
        try:
                obj_post=find_post_cok(c.message.chat.id) 
                obj=obj_post[1]	
                if obj.saved==0:
                        msg =bot.answer_callback_query(c.id,lengstr(ll,93))					         
                        return	                        
                obj.saved=0
                delete_post(obj.post_id)
                msg =bot.answer_callback_query(c.id,lengstr(ll,94))					         
                return	
        except Exception:
            obj=select_from_saved(int(c.data[:-1]))
            delete_post(int(c.data[:-1]))	
            channel_id=obj.channel_id
            keyboard = types.InlineKeyboardMarkup(row_width=1)
            keyboard.add(*[types.InlineKeyboardButton(text=name,callback_data=str(channel_id)+'@') for name in [lengstr(ll,95)]])			
            try:			
                msg = bot.edit_message_text(chat_id=c.message.chat.id, message_id=bib, text=lengstr(ll,94),reply_markup=keyboard)
            except Exception:
                msg = bot.delete_message(chat_id=c.message.chat.id, message_id=bib) 
                msg = bot.send_message(chat_id=c.message.chat.id, text=lengstr(ll,94),reply_markup=keyboard)			
        return	
##############################################################		
    if c.data=='š':
            add_channel.remove(c.message.chat.id) 
            msg = bot.edit_message_text(chat_id=c.message.chat.id, message_id=bib, text=lengstr(ll,37))		
            return	
################################### nazad v poste
    if c.data=='<':
                obj_post=find_post_cok(c.message.chat.id) 
                obj=obj_post[1]					
                post_id=obj.post_id
                channel_id=obj.channel_id	
                obj.add_type=''
                dump_post_cok(c.message.chat.id,obj)				
                keyboard = kukoard(post_id,channel_id,c.message.chat.id)  
                try:				
                    msg =bot.edit_message_text(chat_id=c.message.chat.id, message_id=bib, text=lengstr(ll,38),reply_markup=keyboard)	
                except Exception:
                    msg = bot.delete_message(chat_id=c.message.chat.id, message_id=bib)
                    msg = bot.send_message(chat_id=c.message.chat.id, text=lengstr(ll,38),reply_markup=keyboard)
                return				
###################################				
    msg =bot.answer_callback_query(c.id,lengstr(ll,71))     				
    return		
		
		
		
		
		
		
		
		
		
		
		
		
		
		
		
def name(m):
    global add_channel, add_post, posts,change_post,reverse_time_id,reverse_time
    ll=chek_leng(m.chat.id)
#####################   Dobavlenie kanala
    if m.text==lengstr(ll,3) or m.text==lengstr(ll,4) or m.text==lengstr(ll,5) or m.text==lengstr(ll,6):
        try:
           gn=reverse_time_id.index(m.chat.id)
           reverse_time_id.remove(m.chat.id) 
           del reverse_time.append[gn]
        except Exception:
           ff=0		
        delete_cok(m.chat.id)  
        try:		
            add_channel.remove(m.chat.id) 		
        except Exception:
            ff=0
    if m.text ==lengstr(ll,3):
        add_channel.append(m.chat.id)
        msg = bot.send_message(m.chat.id,lengstr(ll,7))
        return
    if m.text ==lengstr(ll,6):
        keyboard = types.InlineKeyboardMarkup(row_width=3)
        keyboard.add(types.InlineKeyboardButton(text="Support", url="https://t.me/FohbotSupportbot"),types.InlineKeyboardButton(text='Donate',callback_data='Donate'),types.InlineKeyboardButton(text=lengstr(ll,96), url="https://t.me/Fohbot_News/3"))
        keyboard.add(*[types.InlineKeyboardButton(text='Choose language',callback_data='lengsnext'+str(0)) for name in [lengstr(ll,72)]])
        msg = bot.send_message(m.chat.id,lengstr(ll,39),reply_markup=keyboard)
        return
########################### dobavit kanal
    if m.chat.id in add_channel:
        try:
           new_id=m.forward_from_chat.id
        except Exception:
           if m.text.isdigit():
               new_id=m.text
           else:
               msg=bot.send_message(m.chat.id,lengstr(ll,47))
        all_channel=chek_chan()
        if new_id not in all_channel:
            add_channel_func(m.chat.id,new_id,m.forward_from_chat.title)
            keyboard = types.InlineKeyboardMarkup(row_width=1)
            channel_list=chek_user_channels(m.chat.id)
            for i in range (0,len(channel_list[0])):
                keyboard.add(*[types.InlineKeyboardButton(text=name,callback_data=str(channel_list[0][i])+'$') for name in [channel_list[1][i]]])
            keyboard.add(*[types.InlineKeyboardButton(text=name,callback_data=name) for name in [lengstr(ll,72)]])
            msg = bot.send_message(m.chat.id,lengstr(ll,8),reply_markup=keyboard)
            add_channel.remove(m.chat.id)			
        else:
            msg = bot.send_message(m.chat.id,lengstr(ll,9))
            add_channel.remove(m.chat.id)
        return
#####################   izmenenie vremeni
    if m.chat.id in reverse_time_id:
        gn=reverse_time_id.index(m.chat.id)
        strtime=reverse_time[gn] 
        for i in range(0,len(strtime)):
                    if strtime[i]==':':
                        channel_id=int(strtime[:i])
                        pst_id=int(strtime[i+1:])
        fg=channel_post_thread(channel_id)
        m_obj=fg[pst_id][1]						
        new_time=m.text	
        try:
                        ss=datetime.strptime(new_time, '%d.%m.%Y.%H.%M')
        except Exception:
                        msg =bot.send_message(m.chat.id,lengstr(ll,34),reply_markup=keyboard)                        

        ll=time.mktime(ss.timetuple())	
        change_forse_time(channel_id,m_obj,ll)
        keyboard = types.InlineKeyboardMarkup()  
        keyboard.add(*[types.InlineKeyboardButton(text=name,callback_data=str(channel_id)+':'+str(numb)+'otsmotr') for name in [lengstr(ll,32)]])
        msg = bot.send_message(m.chat.id,lengstr(ll,52),reply_markup=keyboard)        		
#####################   moi kanali
    if m.text==lengstr(ll,4):
            channel_list=chek_user_channels(m.chat.id)
            keyboard = types.InlineKeyboardMarkup(row_width=2)
            if len(channel_list[0])>8:
                top=8  
            else:
                top=len(channel_list[0])			
            for i in range (0,top): 
                keyboard.add(*[types.InlineKeyboardButton(text=name,callback_data=str(channel_list[0][i])+'$') for name in [channel_list[1][i]]])
            if len(channel_list[0])>8:
                keyboard.add(*[types.InlineKeyboardButton(text=name,callback_data=str(1)+'Gur') for name in ['»']])                
            keyboard.add(*[types.InlineKeyboardButton(text=name,callback_data=name) for name in [lengstr(ll,72)]])
            msg = bot.send_message(m.chat.id,lengstr(ll,73)+': '+str(len(channel_list[0])),reply_markup=keyboard)
            return
#####################  dobavlenie texta v novi post
    obj_post=find_post_cok(m.chat.id)
    print(obj_post)
    try:
        if obj_post[0]=='ADD':
                obj_post[1].text=m.text.replace('~','`')
                
                dump_post_cok(m.chat.id,obj_post[1])
                post_id=obj_post[1].post_id
                channel_id=obj_post[1].channel_id
                keyboard = types.InlineKeyboardMarkup(row_width=2)
                keyboard.add(types.InlineKeyboardButton(text=lengstr(ll,97),callback_data=str(post_id)+'k'),
                types.InlineKeyboardButton(text=lengstr(ll,98),callback_data=str(post_id)+'r'),
                types.InlineKeyboardButton(text=lengstr(ll,84),callback_data=str(post_id)+'oo'),
                types.InlineKeyboardButton(text=lengstr(ll,99),callback_data=str(post_id)+'v'),
                types.InlineKeyboardButton(text=lengstr(ll,100),callback_data=str(post_id)+'s'),
                types.InlineKeyboardButton(text=lengstr(ll,101),callback_data=str(post_id)+'d'),
                types.InlineKeyboardButton(text=lengstr(ll,74),callback_data=str(post_id)+'p'),
                types.InlineKeyboardButton(text=lengstr(ll,102),callback_data=str(post_id)+'n'),
                types.InlineKeyboardButton(text='⚪️'+lengstr(ll,103),callback_data=str(post_id)+'z'),
                types.InlineKeyboardButton(text='⚪️'+lengstr(ll,104),callback_data=str(post_id)+'u'),
                types.InlineKeyboardButton(text=lengstr(ll,105),callback_data=str(post_id)+'l'),
                types.InlineKeyboardButton(text=lengstr(ll,86),callback_data=str(channel_id)+'!'))
                try:
                    msg =bot.send_message(m.chat.id, obj_post[1].text,reply_markup=keyboard,parse_mode='Markdown')
                except Exception:
                    msg =bot.send_message(m.chat.id, obj_post[1].text,reply_markup=keyboard)                    
                return				
############################### shapka izmeneni		
        obj=obj_post[1]
############################### caption add
        if 	obj.add_type=='caption':
                    obj.add_type==''					
                    obj.text=m.text
                    dump_post_cok(m.chat.id,obj)
                    keyboard=kukoard(obj.post_id,obj.channel_id,m.chat.id)					
                    msg =bot.send_message(m.chat.id, lengstr(ll,42),reply_markup=keyboard)
############################### dobavlenie knopok
        if 	obj.add_type=='buttons':
                    buttons_take=	buttons_create.buttons(m.text)
                    print(buttons_take)	
                    obj.add_type==''					
                    obj.buttons=buttons_take[0]
                    obj.buttons_url=buttons_take[1]
                    dump_post_cok(m.chat.id,obj)
                    keyboard=kukoard(obj.post_id,obj.channel_id,m.chat.id)					
                    msg =bot.send_message(m.chat.id, lengstr(ll,16),reply_markup=keyboard)					
############################### dobavlenie reakci					
        if 	obj.add_type=='reactions':
                    smiles=m.text
                    obj.reactions=[]
                    for z in range(0,len(smiles)):
                        obj.reactions.append(smiles[z])
                    keyboard =kukoard(obj.post_id,obj.channel_id,m.chat.id)
                    msg =bot.send_message(m.chat.id,lengstr(ll,18),reply_markup=keyboard)
                    print(obj.reactions)
                    dump_post_cok(m.chat.id,obj)						
        if 	obj.add_type=='vrema':
                    dd=m.text
                    obj.forse_time=0
                    try:
                        ss=datetime.strptime(dd, '%d.%m.%Y.%H.%M')
                    except Exception:
                        msg =bot.send_message(m.chat.id,lengstr(ll,34),reply_markup=keyboard)                        
                    print(ss)
                    ll=time.mktime(ss.timetuple())
                    obj.forse_time=ll  
                    print(obj.forse_time)					
                    keyboard =kukoard(obj.post_id,obj.channel_id,m.chat.id)
                    dump_post_cok(m.chat.id,obj)
                    msg =bot.send_message(m.chat.id,lengstr(ll,34),reply_markup=keyboard)	
    except Exception:
        return	


def   add_leng(id,ll):
    conn = sqlite3.connect('BD.sqlite')
    cursor = conn.cursor()
	
    array=[]
    array1=[]
    cursor.execute("SELECT * FROM LENGS WHERE USER_ID = :user_id",{"user_id":id})
    results = cursor.fetchall()
    conn.close()
    if len(results)==0:	
        conn = sqlite3.connect('BD.sqlite')
        cursor = conn.cursor()
        cursor.execute("insert into LENGS values (:message_id, :channel_id) ", {"message_id": id,"channel_id": ll})
        conn.commit()
        conn.close()   
    else:
        conn = sqlite3.connect('BD.sqlite')
        cursor = conn.cursor()
        cursor.execute("UPDATE LENGS SET LENG = :ll WHERE USER_ID = :cennel_id",{"cennel_id":id,"ll":ll})

        conn.commit()
        conn.close()          	




		
def chek_leng(id):
    conn = sqlite3.connect('BD.sqlite')
    cursor = conn.cursor()
	
    array=[]
    array1=[]
    cursor.execute("SELECT * FROM LENGS WHERE USER_ID = :user_id",{"user_id":id})
    results = cursor.fetchall()
    conn.close()
    if len(results)==0:	
       return(1)	
    return(results[0][1])



		
def chek_timeb(channel_id):
    conn = sqlite3.connect('CHANNELS.sqlite')
    cursor = conn.cursor()

    cursor.execute("SELECT TIMEB FROM CHANNELS WHERE CHANNEL_ID = :cennel_id ",{"cennel_id":channel_id})
  
    results = cursor.fetchall()			
    conn.close()
    print(results[0][0])	
    return(results[0][0])    	
		
		
def         change_belt_t(channel_id,belt):
    conn = sqlite3.connect('CHANNELS.sqlite')
    cursor = conn.cursor()
    cursor.execute("UPDATE CHANNELS SET TIMEB = :ll WHERE CHANNEL_ID = :cennel_id",{"cennel_id":channel_id,"ll":belt})

    conn.commit()
    conn.close() 		
		
def   change_forse_time(channel_id,m_obj,ll):
    conn = sqlite3.connect('Thread.sqlite')
    cursor = conn.cursor()
    cursor.execute("UPDATE FORSEDER SET FORSE_TIME = :ll WHERE CHANNEL_ID = :cennel_id AND MESSAGE_OBJECT = :m_obj",{"cennel_id":channel_id,"m_obj":m_obj,"ll":ll})

    conn.commit()
    conn.close() 		
		
		
		
		
		
		
def   delete_from_fors(channel_id,m_obj):
    conn = sqlite3.connect('Thread.sqlite')
    cursor = conn.cursor()
    cursor.execute("DELETE FROM FORSEDER WHERE CHANNEL_ID = :cennel_id AND MESSAGE_OBJECT = :m_obj",{"cennel_id":channel_id,"m_obj":m_obj})

    conn.commit()
    conn.close() 

		
def   channel_post_thread(channel_id):		
    conn = sqlite3.connect('Thread.sqlite')
    cursor = conn.cursor()
    cursor.execute("SELECT * FROM FORSEDER WHERE CHANNEL_ID = :ch_id", {"ch_id": channel_id,})
    results = cursor.fetchall()	    
    conn.close() 
    return(results)	
		
        	
def delete_channel(channel_id):
    conn = sqlite3.connect('CHANNELS.sqlite')
    cursor = conn.cursor()
    cursor.execute("DELETE FROM CHANNELS WHERE CHANNEL_ID = :cennel_id ",{"cennel_id":channel_id})

    conn.commit()
    conn.close() 	
    conn = sqlite3.connect('BD.sqlite')
    cursor = conn.cursor()
    cursor.execute("DELETE FROM POSTS WHERE CHANNEL_ID = :cennel_id ",{"cennel_id":channel_id})

    conn.commit()
    conn.close()
    conn = sqlite3.connect('BD.sqlite')
    cursor = conn.cursor()
    cursor.execute("DELETE FROM USERS WHERE CHANNEL = :cennel_id ",{"cennel_id":channel_id})

    conn.commit()
    conn.close()
    conn = sqlite3.connect('Thread.sqlite')
    cursor = conn.cursor()
    cursor.execute("DELETE FROM DELETER WHERE CHANNEL_ID = :cennel_id ",{"cennel_id":channel_id})

    conn.commit()
    conn.close()
    conn = sqlite3.connect('Thread.sqlite')
    cursor = conn.cursor()
    cursor.execute("DELETE FROM POSTS WHERE CHANNEL_ID = :cennel_id ",{"cennel_id":channel_id})

    conn.commit()
    conn.close()
    conn = sqlite3.connect('Thread.sqlite')
    cursor = conn.cursor()
    cursor.execute("DELETE FROM REACTIONS WHERE CHENNEL_ID = :cennel_id ",{"cennel_id":channel_id})

    conn.commit()
    conn.close()	
			
def add_delete_func(message_id,channel_id,delete_time):
    conn = sqlite3.connect('Thread.sqlite')
    cursor = conn.cursor()
    delete_time=delete_time+int(time.time())
    print(message_id)
    print(channel_id)
    print(delete_time)
    cursor.execute("insert into DELETER values (:message_id, :channel_id, :delete_time) ", {"message_id": message_id,"channel_id": channel_id,"delete_time": delete_time,})
    conn.commit()
    conn.close()    


def add_forse_func(forse_time,message_object,ch_id):
    conn = sqlite3.connect('Thread.sqlite')
    cursor = conn.cursor()
    cursor.execute("insert into FORSEDER values (:forse_time, :message_object, :ch_id) ", {"forse_time": forse_time,"message_object": message_object,"ch_id": ch_id})
    conn.commit()
    conn.close() 	
			
def add_channel_func(user_id,channel_id,title):
    conn = sqlite3.connect('BD.sqlite')
    cursor = conn.cursor()
	
    cursor.execute("insert into USERS values (:user_id, :channel_id, :title) ", {"user_id": user_id,"channel_id": channel_id,"title": title,})
    msg = bot.get_chat(channel_id)
    print(msg.id)
    try:
            print('https://t.me/'+msg.username)
            usrnm='https://t.me/'+msg.username
    except Exception:
            usrnm=lengstr(ll,70)
    usid=str(msg.id)
    print(msg)
    txt='New channel: ['+title+']('+usrnm+')\nChannelID: '+usid[4:]    
    bot.send_message(-1001237388190,txt,parse_mode='Markdown')
    conn.commit()
    conn.close()
		
def chek_chan():
    conn = sqlite3.connect('BD.sqlite')
    cursor = conn.cursor()
	
    array=[]
    cursor.execute("SELECT CHANNEL FROM USERS")
    results = cursor.fetchall()
    for i in range(0,len(results)):
        array.append(results[i][0])
    conn.close()
    print(array)
    return(array)
	
def chek_user_channels(user_id):
    conn = sqlite3.connect('BD.sqlite')
    cursor = conn.cursor()
	
    array=[]
    array1=[]
    cursor.execute("SELECT * FROM USERS WHERE USER = :user_id",{"user_id":user_id})
    results = cursor.fetchall()
    for i in range(0,len(results)):
        array.append(results[i][1])
        array1.append(results[i][2])		
    conn.close()
    print(array,array1)
    res=[]
    res.append(array)
    res.append(array1)	
    return(res)	
	
def chan_name_func(channel_id):
    conn = sqlite3.connect('BD.sqlite')
    cursor = conn.cursor()
	
    cursor.execute("SELECT TITLE FROM USERS WHERE CHANNEL = :channel_id",{"channel_id":channel_id})
    results = cursor.fetchall()
    conn.close()
    print(results)
    return(results)


	
def add_reactions(channel_id,message_id,reactions): 
    conn = sqlite3.connect('Thread.sqlite')
    cursor = conn.cursor()
    aa=[]
    for i in range(0,len(reactions)):
        aa.append(0)
    numbers =pickle.dumps(aa)
    users=[]
    users = pickle.dumps(users)
    reactions=pickle.dumps(reactions)
    cursor.execute("insert into REACTIONS values (:channel_id, :message_id, :users,:reactions,:numbers) ", {"channel_id": channel_id,"message_id": message_id,"users": users,"reactions": reactions,"numbers": numbers})
    conn.commit()
    conn.close()   
  
def  reactions_forse(message_id,cennel_id,user_id,num):  
    conn = sqlite3.connect('Thread.sqlite')
    cursor = conn.cursor()

    cursor.execute("SELECT * FROM REACTIONS WHERE CHENNEL_ID = :cennel_id AND MESSAGE_ID = :message_id",{"cennel_id":cennel_id,"message_id":message_id})
  
    results = cursor.fetchall()	    
    conn.close()
    print(results)
    users=pickle.loads(results[0][2])
    reactions=pickle.loads(results[0][3])
    numbers=pickle.loads(results[0][4])
    if user_id not in users:
        numbers[num]+=1
        users.append(user_id)
        keyboard = types.InlineKeyboardMarkup()        
        keyboard.add(*[types.InlineKeyboardButton(text=str(reactions[i])+':'+str(numbers[i]),callback_data='roact'+str(i)+':'+str(reactions[i])) for i in range(0,len(reactions))]) 
        msg =bot.edit_message_reply_markup(cennel_id,message_id,reply_markup=keyboard)		
		

        conn = sqlite3.connect('Thread.sqlite')
        cursor = conn.cursor()
        numbers =pickle.dumps(numbers)
        users = pickle.dumps(users)
        reactions=pickle.dumps(reactions)
        cursor.execute("UPDATE REACTIONS SET USRES = :users, NUMBERS = :numbers WHERE CHENNEL_ID = :cennel_id AND MESSAGE_ID = :message_id ", {"cennel_id": cennel_id,"message_id": message_id,"users": users,"numbers": numbers})
        conn.commit()
        conn.close()  
        return('Y')
    else:
        return('N')	
  
  
def save_post(post_id,user_id,channel_id,postr):
    conn = sqlite3.connect('SAVES.sqlite')
    cursor = conn.cursor()
    inputt = pickle.dumps(postr)
    times=str(datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
    print(times)
    cursor.execute("insert into SAVED_POST values (:post_id, :user_id, :channel_id, :message_object, :time) ", {"post_id": post_id,"user_id": user_id,"channel_id": channel_id,"message_object": inputt,"time": times,})
    conn.commit()
    conn.close()   
  
  
def delete_post(post_id):
    conn = sqlite3.connect('SAVES.sqlite')
    cursor = conn.cursor()
    cursor.execute("DELETE FROM SAVED_POST WHERE POST_ID = :post_id",{"post_id":post_id})
    conn.commit()    
    conn.close()   
  

	
def post_in_channel(channel_id):
    conn = sqlite3.connect('SAVES.sqlite')
    cursor = conn.cursor()

    cursor.execute("SELECT * FROM SAVED_POST WHERE CHANNEL_ID = :cennel_id ",{"cennel_id":channel_id})
  
    results = cursor.fetchall()	    
    conn.close() 
    return(results)	
  
  
  
def select_from_saved(inm):
    conn = sqlite3.connect('SAVES.sqlite')
    cursor = conn.cursor()

    cursor.execute("SELECT * FROM SAVED_POST WHERE POST_ID = :cennel_id ",{"cennel_id":inm})
  
    results = cursor.fetchall()	
    res=pickle.loads(results[0][3])    
    conn.close() 
    return(res)	

def  chek_channel_options(channel_id):
    conn = sqlite3.connect('CHANNELS.sqlite')
    cursor = conn.cursor()

    cursor.execute("SELECT * FROM CHANNELS WHERE CHANNEL_ID = :cennel_id ",{"cennel_id":channel_id})
  
    results = cursor.fetchall()	
    if len(results)==0:
        cursor.execute("insert into CHANNELS values (:a1, :a2, :a3,:a4,:a5,:a6,:a7,:a8) ", {"a1": '',"a2": '',"a3": 0,"a4": 0,"a5": 0,"a6": 0,"a7": 0,"a8": channel_id})
        conn.commit()	
        cursor.execute("SELECT * FROM CHANNELS WHERE CHANNEL_ID = :cennel_id ",{"cennel_id":channel_id})
  
        results = cursor.fetchall()	
    res=results[0]
    conn.close() 
    return(res)	

	
def  change_option_channel(num,znach,channel):
    conn = sqlite3.connect('CHANNELS.sqlite')
    cursor = conn.cursor()
    cursor.execute("SELECT * FROM CHANNELS WHERE CHANNEL_ID = :cennel_id ",{"cennel_id":channel})
  
    results = cursor.fetchall()	   
    if num==1:
            cursor.execute("UPDATE CHANNELS SET REACTIONS = :nom WHERE CHANNEL_ID = :cennel_id ",{"nom":znach, "cennel_id":channel})  
    if num==3:
            tim=[0,3600,10800,21600,43200,86400,172800,604800,1209600,2592000]
            delete_time=tim[int(znach)]
            cursor.execute("UPDATE CHANNELS SET DELETE_TIME = :nom WHERE CHANNEL_ID = :cennel_id ",{"nom":delete_time, "cennel_id":channel}) 			
    if num==4:
        if results[0][3]==0:
            cursor.execute("UPDATE CHANNELS SET DELETE_IN = :nom WHERE CHANNEL_ID = :cennel_id ",{"nom":1, "cennel_id":channel})        
        else:
            cursor.execute("UPDATE CHANNELS SET DELETE_IN = :nom WHERE CHANNEL_ID = :cennel_id ",{"nom":0, "cennel_id":channel}) 
    if num==5:
        if results[0][4]==0:
            cursor.execute("UPDATE CHANNELS SET SSILKA_IN = :nom WHERE CHANNEL_ID = :cennel_id ",{"nom":1, "cennel_id":channel})        
        else:
            cursor.execute("UPDATE CHANNELS SET SSILKA_IN = :nom WHERE CHANNEL_ID = :cennel_id ",{"nom":0, "cennel_id":channel})
    if num==6:
        if results[0][5]==0:
            cursor.execute("UPDATE CHANNELS SET REACTIONS_IN = :nom WHERE CHANNEL_ID = :cennel_id ",{"nom":1, "cennel_id":channel})        
        else:
            cursor.execute("UPDATE CHANNELS SET REACTIONS_IN = :nom WHERE CHANNEL_ID = :cennel_id ",{"nom":0, "cennel_id":channel})
    if num==7:
        if results[0][6]==0:
            cursor.execute("UPDATE CHANNELS SET MUT_IN = :nom WHERE CHANNEL_ID = :cennel_id ",{"nom":1, "cennel_id":channel})        
        else:
            cursor.execute("UPDATE CHANNELS SET MUT_IN = :nom WHERE CHANNEL_ID = :cennel_id ",{"nom":0, "cennel_id":channel})			
    conn.commit()
    conn.close() 




def   add_post_cok(user_id,channel_id,post_id,obj):
    conn = sqlite3.connect('CHANNELS.sqlite')
    cursor = conn.cursor()    
    obj1 = pickle.dumps(obj)
    cursor.execute("insert into COK values (:a1, :a2, :a3,:a4,:a5) ", {"a1": user_id,"a2": channel_id,"a3": post_id,"a4": obj1,"a5": 'ADD'})
    conn.commit()
    conn.close() 

def find_post_cok(user_id):
    try:
        conn = sqlite3.connect('CHANNELS.sqlite')
        cursor = conn.cursor()
        cursor.execute("SELECT * FROM COK WHERE USER_ID = :cennel_id ",{"cennel_id":user_id})
        results = cursor.fetchall()	
        res=[]
        res.append(results[0][4])
        obj=pickle.loads(results[0][3])
        res.append(obj)
        cursor.execute("UPDATE COK SET CHANGE = :nom WHERE USER_ID = :cennel_id ",{"nom":'', "cennel_id":user_id})
        conn.commit()
        conn.close() 
        return(res)
    except Exception:
        return()
	
	
def  dump_post_cok(user_id,obj):	
    conn = sqlite3.connect('CHANNELS.sqlite')
    cursor = conn.cursor()
    obj1 = pickle.dumps(obj)
    cursor.execute("UPDATE COK SET POST_OBJ = :nom WHERE USER_ID = :cennel_id ",{"nom":obj1, "cennel_id":user_id})
    conn.commit()
    conn.close() 



def  delete_cok(user_id):
    conn = sqlite3.connect('CHANNELS.sqlite')
    cursor = conn.cursor()
    cursor.execute("DELETE FROM COK WHERE USER_ID = :cennel_id ",{"cennel_id":user_id})
    conn.commit()
    conn.close()     

	
	
def users_count():
    conn = sqlite3.connect('BD.sqlite')
    cursor = conn.cursor()
	

    cursor.execute("SELECT USER FROM USERS ")
    results = cursor.fetchall()	
    results=list(set(results))	
    conn.close()
    return(len(results)) 

	
def cannel_count():
    conn = sqlite3.connect('BD.sqlite')
    cursor = conn.cursor()
	

    cursor.execute("SELECT * FROM USERS ")
    results = cursor.fetchall()		
    conn.close()
    return(len(results))  

def chat_mem_count():
    conn = sqlite3.connect('BD.sqlite')
    cursor = conn.cursor()
	
    sum=0
    cursor.execute("SELECT * FROM USERS ")
    results = cursor.fetchall()	
    for i in range(0,len(results)):
        try:
           msg=bot.get_chat_members_count(results[i][1])
           sum+=int(msg)
        except Exception:
           zzs=0		
    conn.close()
    return(sum)  


def option_keyboard(channel):
                    channel_id=channel[7]
                    if channel[3]==0:
                        uk='⚪️'
                    else:
                        uk='🔘'
                    if channel[4]==0:
                        pk='⚪️'
                    else:
                        pk='🔘'
                    if channel[5]==0:
                        rk='⚪️'
                    else:
                        rk='🔘'
                    if channel[6]==0:
                        zk='⚪️'
                    else:
                        zk='🔘'						
                    keyboard = types.InlineKeyboardMarkup(row_width=2)
                    keyboard.add(types.InlineKeyboardButton(text=lengstr(ll,106),callback_data=str(channel_id)+'ot'),
                    types.InlineKeyboardButton(text=lengstr(ll,107),callback_data=str(channel_id)+'ra'),
                    types.InlineKeyboardButton(text=lengstr(ll,108),callback_data=str(channel_id)+'tp'),
                    types.InlineKeyboardButton(text=lengstr(ll,109),callback_data=str(channel_id)+'dl'))
                    keyboard.add(types.InlineKeyboardButton(text=uk+lengstr(ll,110),callback_data=str(channel_id)+'dp'))
                    keyboard.add(types.InlineKeyboardButton(text=pk+lengstr(ll,111),callback_data=str(channel_id)+'ps'))
                    keyboard.add(types.InlineKeyboardButton(text=rk+lengstr(ll,112),callback_data=str(channel_id)+'rp'))
                    keyboard.add(types.InlineKeyboardButton(text=zk+lengstr(ll,113),callback_data=str(channel_id)+'mu'))
                    keyboard.add(types.InlineKeyboardButton(text=lengstr(ll,114),callback_data=str(channel_id)+'dc'))
                    keyboard.add(types.InlineKeyboardButton(text=lengstr(ll,80),callback_data=str(channel_id)+'$'))		
                    return(keyboard)
def kukoard(post_id,channel_id,user_id):
                    obj_post=find_post_cok(user_id) 
                    obj=obj_post[1]	
                    mut=obj.mut
                    pin=obj.pin
                    if mut<=0:
                        mut_str='⚪️ '
                    else:
                        mut_str='🔘 '
                    if pin<=0:
                        pin_str='⚪️ '
                    else:
                        pin_str='🔘 '
                    print(channel_id)
                    keyboard = types.InlineKeyboardMarkup(row_width=2)
                    if obj.text==None and (obj.document_type=='photo' or obj.document_type=='video' or obj.document_type=='document'):
                        keyboard.add(types.InlineKeyboardButton(text=lengstr(ll,115),callback_data=str(post_id)+'cpt'))
                    keyboard.add(types.InlineKeyboardButton(text=lengstr(ll,97),callback_data=str(post_id)+'k'),
                    types.InlineKeyboardButton(text=lengstr(ll,98),callback_data=str(post_id)+'r'),
                    types.InlineKeyboardButton(text=lengstr(ll,84),callback_data=str(post_id)+'oo'),
                    types.InlineKeyboardButton(text=lengstr(ll,99),callback_data=str(post_id)+'v'),
                    types.InlineKeyboardButton(text=lengstr(ll,100),callback_data=str(post_id)+'s'),
                    types.InlineKeyboardButton(text=lengstr(ll,101),callback_data=str(post_id)+'d'),
                    types.InlineKeyboardButton(text=lengstr(ll,74),callback_data=str(post_id)+'p'),
                    types.InlineKeyboardButton(text=lengstr(ll,102),callback_data=str(post_id)+'n'),
                    types.InlineKeyboardButton(text=pin_str+lengstr(ll,103),callback_data=str(post_id)+'z'),
                    types.InlineKeyboardButton(text=mut_str+lengstr(ll,104),callback_data=str(post_id)+'u'),
                    types.InlineKeyboardButton(text=lengstr(ll,105),callback_data=str(post_id)+'l'),
                    types.InlineKeyboardButton(text=lengstr(ll,86),callback_data=str(channel_id)+'!'))
                    return(keyboard)
					
					
					
					
					
					
					
					
					
					
					
					
					
############################################# dostaem photo dla posta	
@bot.message_handler(content_types=['photo'])
def photoget(message):
    print(message)
    ll=1
    if message.chat.id in add_channel:
        new_id=message.forward_from_chat.id
        all_channel=chek_chan()
        if new_id not in all_channel:
            add_channel_func(message.chat.id,new_id,message.forward_from_chat.title)
            keyboard = types.InlineKeyboardMarkup(row_width=1)
            channel_list=chek_user_channels(message.chat.id)
            for i in range (0,len(channel_list[0])):
                keyboard.add(*[types.InlineKeyboardButton(text=name,callback_data=str(channel_list[0][i])+'$') for name in [channel_list[1][i]]])
            keyboard.add(*[types.InlineKeyboardButton(text=name,callback_data=name) for name in [lengstr(ll,72)]])
            msg = bot.send_message(message.chat.id,lengstr(ll,8),reply_markup=keyboard)
            add_channel.remove(message.chat.id)			
        else:
            msg = bot.send_message(message.chat.id,lengstr(ll,9))
            add_channel.remove(message.chat.id)
        return
    obj_post=find_post_cok(message.chat.id)
    print(obj_post)
    if obj_post[0]=='ADD':
                obj=obj_post[1]
                obj.document_type='photo'
                obj.text=message.caption
                post_id=obj.post_id
                channel_id=obj.channel_id
                obj.document=message.photo[2].file_id
                print(obj.document)
                dump_post_cok(message.chat.id,obj)
                keyboard = types.InlineKeyboardMarkup(row_width=2)
                if obj.text==None:
                    keyboard.add(types.InlineKeyboardButton(text=lengstr(ll,115),callback_data=str(post_id)+'cpt'))
                keyboard.add(types.InlineKeyboardButton(text=lengstr(ll,97),callback_data=str(post_id)+'k'),
                types.InlineKeyboardButton(text=lengstr(ll,98),callback_data=str(post_id)+'r'),
                types.InlineKeyboardButton(text=lengstr(ll,84),callback_data=str(post_id)+'oo'),
                types.InlineKeyboardButton(text=lengstr(ll,99),callback_data=str(post_id)+'v'),
                types.InlineKeyboardButton(text=lengstr(ll,100),callback_data=str(post_id)+'s'),
                types.InlineKeyboardButton(text=lengstr(ll,101),callback_data=str(post_id)+'d'),
                types.InlineKeyboardButton(text=lengstr(ll,74),callback_data=str(post_id)+'p'),
                types.InlineKeyboardButton(text=lengstr(ll,102),callback_data=str(post_id)+'n'),
                types.InlineKeyboardButton(text='⚪️'+ lengstr(ll,103),callback_data=str(post_id)+'z'),
                types.InlineKeyboardButton(text='⚪️'+ lengstr(ll,104),callback_data=str(post_id)+'u'),
                types.InlineKeyboardButton(text=lengstr(ll,105),callback_data=str(post_id)+'l'),
                types.InlineKeyboardButton(text=lengstr(ll,86),callback_data=str(channel_id)+'!'))
                msg =bot.send_photo(message.chat.id, obj.document,caption=obj.text,reply_markup=keyboard) 
########################### dostaem audio	
@bot.message_handler(content_types=['audio'])
def photoget(message):
    print(message)
    ll=1
    if message.chat.id in add_channel:
        new_id=message.forward_from_chat.id
        all_channel=chek_chan()
        if new_id not in all_channel:
            add_channel_func(message.chat.id,new_id,message.forward_from_chat.title)
            keyboard = types.InlineKeyboardMarkup(row_width=1)
            channel_list=chek_user_channels(message.chat.id)
            for i in range (0,len(channel_list[0])):
                keyboard.add(*[types.InlineKeyboardButton(text=name,callback_data=str(channel_list[0][i])+'$') for name in [channel_list[1][i]]])
            keyboard.add(*[types.InlineKeyboardButton(text=name,callback_data=name) for name in [lengstr(ll,72)]])
            msg = bot.send_message(message.chat.id,lengstr(ll,8),reply_markup=keyboard)
            add_channel.remove(message.chat.id)			
        else:
            msg = bot.send_message(message.chat.id,lengstr(ll,9))
            add_channel.remove(message.chat.id)
        return
    obj_post=find_post_cok(message.chat.id)
    print(obj_post)
    if obj_post[0]=='ADD':
                obj=obj_post[1]
                obj.document_type='audio'
                obj.text=message.caption
                post_id=obj.post_id
                channel_id=obj.channel_id
                obj.document=message.audio.file_id
                dump_post_cok(message.chat.id,obj)
                keyboard = types.InlineKeyboardMarkup(row_width=2)
                keyboard.add(types.InlineKeyboardButton(text=lengstr(ll,97),callback_data=str(post_id)+'k'),
                types.InlineKeyboardButton(text=lengstr(ll,98),callback_data=str(post_id)+'r'),
                types.InlineKeyboardButton(text=lengstr(ll,84),callback_data=str(post_id)+'oo'),
                types.InlineKeyboardButton(text=lengstr(ll,99),callback_data=str(post_id)+'v'),
                types.InlineKeyboardButton(text=lengstr(ll,100),callback_data=str(post_id)+'s'),
                types.InlineKeyboardButton(text=lengstr(ll,101),callback_data=str(post_id)+'d'),
                types.InlineKeyboardButton(text=lengstr(ll,74),callback_data=str(post_id)+'p'),
                types.InlineKeyboardButton(text=lengstr(ll,102),callback_data=str(post_id)+'n'),
                types.InlineKeyboardButton(text='⚪️'+ lengstr(ll,103),callback_data=str(post_id)+'z'),
                types.InlineKeyboardButton(text='⚪️'+ lengstr(ll,104),callback_data=str(post_id)+'u'),
                types.InlineKeyboardButton(text=lengstr(ll,105),callback_data=str(post_id)+'l'),
                types.InlineKeyboardButton(text=lengstr(ll,86),callback_data=str(channel_id)+'!'))
                msg =bot.send_audio(message.chat.id, obj.document,caption=obj.text,reply_markup=keyboard) 
####################### dostaem video
@bot.message_handler(content_types=['video'])
def photoget(message):
    print(message)
    ll=1
    if message.chat.id in add_channel:
        new_id=message.forward_from_chat.id
        all_channel=chek_chan()
        if new_id not in all_channel:
            add_channel_func(message.chat.id,new_id,message.forward_from_chat.title)
            keyboard = types.InlineKeyboardMarkup(row_width=1)
            channel_list=chek_user_channels(message.chat.id)
            for i in range (0,len(channel_list[0])):
                keyboard.add(*[types.InlineKeyboardButton(text=name,callback_data=str(channel_list[0][i])+'$') for name in [channel_list[1][i]]])
            keyboard.add(*[types.InlineKeyboardButton(text=name,callback_data=name) for name in [lengstr(ll,72)]])
            msg = bot.send_message(message.chat.id,lengstr(ll,8),reply_markup=keyboard)
            add_channel.remove(message.chat.id)			
        else:
            msg = bot.send_message(message.chat.id,lengstr(ll,9))
            add_channel.remove(message.chat.id)
        return
    obj_post=find_post_cok(message.chat.id)
    if obj_post[0]=='ADD':
                obj=obj_post[1]
                obj.document_type='video'
                obj.text=message.caption
                post_id=obj.post_id
                channel_id=obj.channel_id
                obj.document=message.video.file_id
                dump_post_cok(message.chat.id,obj)
                keyboard = types.InlineKeyboardMarkup(row_width=2)
                if obj.text==None:
                    keyboard.add(types.InlineKeyboardButton(text=lengstr(ll,115),callback_data=str(post_id)+'cpt'))
                keyboard.add(types.InlineKeyboardButton(text=lengstr(ll,97),callback_data=str(post_id)+'k'),
                types.InlineKeyboardButton(text=lengstr(ll,98),callback_data=str(post_id)+'r'),
                types.InlineKeyboardButton(text=lengstr(ll,84),callback_data=str(post_id)+'oo'),
                types.InlineKeyboardButton(text=lengstr(ll,99),callback_data=str(post_id)+'v'),
                types.InlineKeyboardButton(text=lengstr(ll,100),callback_data=str(post_id)+'s'),
                types.InlineKeyboardButton(text=lengstr(ll,101),callback_data=str(post_id)+'d'),
                types.InlineKeyboardButton(text=lengstr(ll,74),callback_data=str(post_id)+'p'),
                types.InlineKeyboardButton(text=lengstr(ll,102),callback_data=str(post_id)+'n'),
                types.InlineKeyboardButton(text='⚪️'+ lengstr(ll,103),callback_data=str(post_id)+'z'),
                types.InlineKeyboardButton(text='⚪️'+ lengstr(ll,104),callback_data=str(post_id)+'u'),
                types.InlineKeyboardButton(text=lengstr(ll,105),callback_data=str(post_id)+'l'),
                types.InlineKeyboardButton(text=lengstr(ll,86),callback_data=str(channel_id)+'!'))
                msg =bot.send_video(message.chat.id, obj.document,caption=obj.text,reply_markup=keyboard) 
######################### dostaem document
@bot.message_handler(content_types=['document'])
def photoget(message):
    print(message)
    ll=1
    if message.chat.id in add_channel:
        new_id=message.forward_from_chat.id
        all_channel=chek_chan()
        if new_id not in all_channel:
            add_channel_func(message.chat.id,new_id,message.forward_from_chat.title)
            keyboard = types.InlineKeyboardMarkup(row_width=1)
            channel_list=chek_user_channels(message.chat.id)
            for i in range (0,len(channel_list[0])):
                keyboard.add(*[types.InlineKeyboardButton(text=name,callback_data=str(channel_list[0][i])+'$') for name in [channel_list[1][i]]])
            keyboard.add(*[types.InlineKeyboardButton(text=name,callback_data=name) for name in [lengstr(ll,72)]])
            msg = bot.send_message(message.chat.id,lengstr(ll,8),reply_markup=keyboard)
            add_channel.remove(message.chat.id)			
        else:
            msg = bot.send_message(message.chat.id,lengstr(ll,9))
            add_channel.remove(message.chat.id)
        return
    obj_post=find_post_cok(message.chat.id)
    if obj_post[0]=='ADD':
                obj=obj_post[1]
                obj.document_type='document'
                obj.text=message.caption
                post_id=obj.post_id
                channel_id=obj.channel_id
                obj.document=message.document.file_id
                dump_post_cok(message.chat.id,obj)
                keyboard = types.InlineKeyboardMarkup(row_width=2)
                if obj.text==None:
                    keyboard.add(types.InlineKeyboardButton(text=lengstr(ll,115),callback_data=str(post_id)+'cpt'))
                keyboard.add(types.InlineKeyboardButton(text=lengstr(ll,97),callback_data=str(post_id)+'k'),
                types.InlineKeyboardButton(text=lengstr(ll,98),callback_data=str(post_id)+'r'),
                types.InlineKeyboardButton(text=lengstr(ll,84),callback_data=str(post_id)+'oo'),
                types.InlineKeyboardButton(text=lengstr(ll,99),callback_data=str(post_id)+'v'),
                types.InlineKeyboardButton(text=lengstr(ll,100),callback_data=str(post_id)+'s'),
                types.InlineKeyboardButton(text=lengstr(ll,101),callback_data=str(post_id)+'d'),
                types.InlineKeyboardButton(text=lengstr(ll,74),callback_data=str(post_id)+'p'),
                types.InlineKeyboardButton(text=lengstr(ll,102),callback_data=str(post_id)+'n'),
                types.InlineKeyboardButton(text='⚪️'+ lengstr(ll,103),callback_data=str(post_id)+'z'),
                types.InlineKeyboardButton(text='⚪️'+ lengstr(ll,104),callback_data=str(post_id)+'u'),
                types.InlineKeyboardButton(text=lengstr(ll,105),callback_data=str(post_id)+'l'),
                types.InlineKeyboardButton(text=lengstr(ll,86),callback_data=str(channel_id)+'!'))
                msg =bot.send_document(message.chat.id, obj.document,caption=obj.text,reply_markup=keyboard) 
############################################# dostaem voice
@bot.message_handler(content_types=['voice'])
def photoget(message):
    print(message)
    ll=1
    if message.chat.id in add_channel:
        new_id=message.forward_from_chat.id
        all_channel=chek_chan()
        if new_id not in all_channel:
            add_channel_func(message.chat.id,new_id,message.forward_from_chat.title)
            keyboard = types.InlineKeyboardMarkup(row_width=1)
            channel_list=chek_user_channels(message.chat.id)
            for i in range (0,len(channel_list[0])):
                keyboard.add(*[types.InlineKeyboardButton(text=name,callback_data=str(channel_list[0][i])+'$') for name in [channel_list[1][i]]])
            keyboard.add(*[types.InlineKeyboardButton(text=name,callback_data=name) for name in [lengstr(ll,72)]])
            msg = bot.send_message(message.chat.id,lengstr(ll,8),reply_markup=keyboard)
            add_channel.remove(message.chat.id)			
        else:
            msg = bot.send_message(message.chat.id,lengstr(ll,9))
            add_channel.remove(message.chat.id)
        return
    obj_post=find_post_cok(message.chat.id)
    print(obj_post)
    if obj_post[0]=='ADD':
                obj=obj_post[1]
                obj.document_type='voice'
                post_id=obj.post_id
                channel_id=obj.channel_id
                obj.document=message.voice.file_id
                dump_post_cok(message.chat.id,obj)
                keyboard = types.InlineKeyboardMarkup(row_width=2)
                keyboard.add(types.InlineKeyboardButton(text=lengstr(ll,97),callback_data=str(post_id)+'k'),
                types.InlineKeyboardButton(text=lengstr(ll,98),callback_data=str(post_id)+'r'),
                types.InlineKeyboardButton(text=lengstr(ll,84),callback_data=str(post_id)+'oo'),
                types.InlineKeyboardButton(text=lengstr(ll,99),callback_data=str(post_id)+'v'),
                types.InlineKeyboardButton(text=lengstr(ll,100),callback_data=str(post_id)+'s'),
                types.InlineKeyboardButton(text=lengstr(ll,101),callback_data=str(post_id)+'d'),
                types.InlineKeyboardButton(text=lengstr(ll,74),callback_data=str(post_id)+'p'),
                types.InlineKeyboardButton(text=lengstr(ll,102),callback_data=str(post_id)+'n'),
                types.InlineKeyboardButton(text='⚪️'+ lengstr(ll,103),callback_data=str(post_id)+'z'),
                types.InlineKeyboardButton(text='⚪️'+ lengstr(ll,104),callback_data=str(post_id)+'u'),
                types.InlineKeyboardButton(text=lengstr(ll,105),callback_data=str(post_id)+'l'),
                types.InlineKeyboardButton(text=lengstr(ll,86),callback_data=str(channel_id)+'!'))
                msg =bot.send_voice(message.chat.id, obj.document,reply_markup=keyboard) 

def deleterer():
    conn = sqlite3.connect('Thread.sqlite')
    cursor = conn.cursor()
	
    array=[]
    now_time=int(time.time())
    cursor.execute("SELECT * FROM DELETER WHERE DELETE_TIME <= :time",{"time":now_time})
    results = cursor.fetchall()	
    print(results)
    cursor.execute("DELETE FROM DELETER WHERE DELETE_TIME <= :time",{"time":now_time})
    conn.commit()    
    conn.close()
    for i in range(0,len(results)):
        msg=bot.delete_message(results[i][1],results[i][0])	
        
def forserer():
    conn = sqlite3.connect('Thread.sqlite')
    cursor = conn.cursor()
	
    array=[]
    now_time=int(time.time())
    cursor.execute("SELECT * FROM FORSEDER WHERE FORSE_TIME <= :time",{"time":now_time})
    results = cursor.fetchall()	
    print(results)
    cursor.execute("DELETE FROM FORSEDER WHERE FORSE_TIME <= :time",{"time":now_time})
    conn.commit()    
    conn.close()
    for i in range(0,len(results)):
                obj=pickle.loads(results[i][1]) 
                channel_id=obj.channel_id
                channel=chek_channel_options(channel_id)				
                print(obj.channel_id)
                keyboard = types.InlineKeyboardMarkup()
                for z in range(0,len(obj.buttons_url)):
                    print(obj.buttons[z],obj.buttons_url[z])
                    url_button = types.InlineKeyboardButton(text=obj.buttons[z], url=obj.buttons_url[z])
                    keyboard.add(url_button)
                if len(obj.reactions)>0:
                    keyboard.add(*[types.InlineKeyboardButton(text=name+':0',callback_data='roact'+str(obj.reactions.index(name))+':'+name) for name in obj.reactions])   
                elif channel[5]>0:
                    if len(channel[0])>0:
                        arreak=[]
                        for sv in range(0,len(channel[0])):
                            arreak.append(channel[0][sv])								
                        keyboard.add(*[types.InlineKeyboardButton(text=name+':0',callback_data='roact'+str(arreak.index(name))+':'+name) for name in arreak])   
                if 	obj.mut==1:
                    mut=True
                elif obj.mut==-1 and channel[6]==1:
                    mut=True
                else:
                    mut=False	
                if channel[4]==1:
                    mut_ssilka=True
                else:
                    mut_ssilka=False
                if 	obj.document_type=='':
                    try:				
                      msg =bot.send_message(obj.channel_id, obj.text,reply_markup=keyboard,disable_notification=mut,disable_web_page_preview=mut_ssilka,parse_mode='Markdown')
                    except Exception:
                      msg =bot.send_message(obj.channel_id, obj.text,reply_markup=keyboard,disable_notification=mut,disable_web_page_preview=mut_ssilka)					
                if  obj.document_type=='photo':
                    msg =bot.send_photo(obj.channel_id, obj.document,caption=obj.text,reply_markup=keyboard,disable_notification=mut)                    
                if  obj.document_type=='audio':
                    msg =bot.send_audio(obj.channel_id, obj.document,caption=obj.text,reply_markup=keyboard,disable_notification=mut)
                if  obj.document_type=='video':
                    msg =bot.send_video(obj.channel_id, obj.document,caption=obj.text,reply_markup=keyboard,disable_notification=mut)
                if  obj.document_type=='document':
                    msg =bot.send_document(obj.channel_id, obj.document,caption=obj.text,reply_markup=keyboard,disable_notification=mut)
                if  obj.document_type=='voice':
                    msg =bot.send_voice(obj.channel_id, obj.document,reply_markup=keyboard,disable_notification=mut)   
                joy=msg.message_id
                if len(obj.reactions)>0:
                    add_reactions(obj.channel_id,joy,obj.reactions)
                elif channel[5]>0:
                    if len(channel[0])>0:
                        add_reactions(obj.channel_id,joy,arreak)					
                if obj.delete_time>0:
                    tim=[0,3600,10800,21600,43200,86400,172800,604800,1209600,2592000]
                    add_delete_func(joy,obj.channel_id,tim[obj.delete_time])  
                elif channel[3]==1:
                    if channel[2]>0:
                        add_delete_func(joy,obj.channel_id,channel[2])  					
                if 	obj.pin==1:
                    msg=bot.pin_chat_message(obj.channel_id,joy,disable_notification=mut)	               
                return	
		
                	
        



schedule.every(1).minutes.do(deleterer)
schedule.every(1).minutes.do(forserer)
def lal():
    while 1:
        schedule.run_pending()
        time.sleep(1)
_thread.start_new_thread(lal,())




bot.remove_webhook()

bot.set_webhook(url=WEBHOOK_URL_BASE + WEBHOOK_URL_PATH,
                certificate=open(WEBHOOK_SSL_CERT, 'r'))

cherrypy.config.update({
    'server.socket_host': WEBHOOK_LISTEN,
    'server.socket_port': WEBHOOK_PORT,
    'server.ssl_module': 'builtin',
    'server.ssl_certificate': WEBHOOK_SSL_CERT,
    'server.ssl_private_key': WEBHOOK_SSL_PRIV
})

cherrypy.quickstart(WebhookServer(), WEBHOOK_URL_PATH, {'/': {}})