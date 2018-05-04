import openpyxl
import time
import threading
import telebot
from telebot import types
import _pickle as pickle
import requests
import json
import random
import string
import os
import xlwt
import urllib
import pochta
global wb, wb1, sheet, sheet1
global arraypaidphone, arraypaidemail
TOKEN = '500239333:AAEpjOsc00JC1_2cw_Kaq_--VIdv_QzSMTA'
WEBHOOK_HOST = '95.46.98.126'
WEBHOOK_PORT = 80  # 443, 80, 88 или 8443 (порт должен быть открыт!)
WEBHOOK_LISTEN = '0.0.0.0'  # На некоторых серверах придется указывать такой же IP, что и выше

WEBHOOK_SSL_CERT = './webhook_cert.pem'  # Путь к сертификату
WEBHOOK_SSL_PRIV = './webhook_pkey.pem'  # Путь к приватному ключу

WEBHOOK_URL_BASE = "https://%s:%s" % (WEBHOOK_HOST, WEBHOOK_PORT)
WEBHOOK_URL_PATH = "/%s/" % (TOKEN)

bot = telebot.TeleBot(TOKEN)


class WebhookServer(object):
    @cherrypy.expose
    def index(self):
        if 'content-length' in cherrypy.request.headers and \
                        'content-type' in cherrypy.request.headers and \
                        cherrypy.request.headers['content-type'] == 'application/json':
            length = int(cherrypy.request.headers['content-length'])
            json_string = cherrypy.request.body.read(length).decode("utf-8")
            update = telebot.types.Update.de_json(json_string)
            bot.process_new_updates([update])
            return ''
        else:
            raise cherrypy.HTTPError(403)





arraypaidphone=[]
arraypaidemail=[]
wb = openpyxl.load_workbook(filename = '1.xlsx')
sheet = wb['test']
wb1 = openpyxl.load_workbook(filename = '2.xlsx')
sheet1 = wb1['test']
def  tour(leng,strn):
    bb=sheet.cell(row=strn, column=leng).value
    return str(bb)
def  info(leng,strn):
    bb=sheet1.cell(row=strn, column=leng).value
    if bb==None:
        bb=''
    return str(bb)
TOKEN = '507631866:AAHB0tjPBoeNXAABu9zAU7Zt4O8jbTgsZDE'

bot = telebot.TeleBot(TOKEN)
admin_password='QWERTY123456!@#$%^'
class userobj():
    id=0
    email=''
    bonus=0
    shfr=''
    spis=0
    phone=''
    phone1=''
    name=''
    koment=''
    adminin=0
    addtext=0
    regpoz=0
    arrayreg=[]
    kudaedu=''
    regprosh=0    
    def __init__(self):
        self.arrayreg=[]
global bdpol, reflist, admin
admin=0
reflist=[]	
bdpol=[]
def nomer(b):
    global bdpol
    z=0
    k=-1
    for i in bdpol:
        k+=1
        if bdpol[k].id==b:
            z=k
            return(z)
			
				
			
			
			
# Handle '/start' and '/help'
@bot.message_handler(commands=['help', 'start'])
def start(message):
    global arraypaidphone, arraypaidemail
    print(arraypaidphone,arraypaidemail)
    print(message)
    if message.chat.id<0:
        return()
    global bdpol
    hesh=message.chat.id
    z=0
    k=-1
    for i in bdpol:
        k+=1
        if bdpol[k].id==hesh:
            z=1
    if z==0:
        bdpol.append(userobj())
        print(bdpol[-1].id)
        bdpol[-1].id=hesh
        bdpol[-1].start=1
    keyboard = types.ReplyKeyboardMarkup(resize_keyboard=True)
    keyboard.add(*[types.InlineKeyboardButton(text=name,callback_data=name) for name in [info(1,13)]])
    keyboard.add(*[types.InlineKeyboardButton(text=name,callback_data=name) for name in [info(1,2), info(1,3),'Бонусы']])
    msg = bot.send_message(message.chat.id, info(1,1),
    reply_markup=keyboard,parse_mode='HTML')

	
	
@bot.message_handler(commands=['admin'])	
def admin(m):
    global bdpol
    if m.chat.id>0:
        msg = bot.send_message(m.chat.id, 'Добро пожаловать в панель администратора!\n\nВведите, пожалуйста, пароль администратора.',parse_mode='HTML')
        k=nomer(m.chat.id)
        bdpol[k].adminin=1
    else:
        return()	
	
	
	
@bot.message_handler(commands=['tours'])	
def tours(message):
    s=''
    for i in range(1,100):
        if  tour(i,1)=='None':
            break
        s=s+str(i)+') <b>'+tour(i,1)+'</b>\n'+tour(i,4)+'\n'+tour(i,5)+'\n'+tour(i,14)+'\n\n'	
    msg = bot.send_message(message.chat.id, info(1,14)+'\n'+s+info(1,23),parse_mode='HTML',disable_web_page_preview=True)	    	
	
	
	
	
	
	
	
	
	
@bot.message_handler(content_types=["text"])
def repeat_all_messages(message): 
    name(message)	
	
	
	
@bot.callback_query_handler(func=lambda c:True)
def inline(c):
    global bdpol, admin
    if c.message.chat.id<0:
        return()
    k=nomer(c.message.chat.id)
    bib=c.message.message_id
    if c.data =='Назад' :
        msg = bot.delete_message(c.message.chat.id, bib)
    if c.data[:-1] in info(1,15):
        keyboard = types.InlineKeyboardMarkup(row_width=1)
        keyboard.add(*[types.InlineKeyboardButton(text=name,callback_data=name) for name in ['Назад']])
        msg = bot.send_message(c.message.chat.id,tour(int(c.data[-1]),5)+'\n\n'+tour(int(c.data[-1]),6),reply_markup=keyboard,parse_mode='HTML')
        return
    if c.data[:-1] in info(1,16):
        keyboard = types.InlineKeyboardMarkup(row_width=1)
        keyboard.add(*[types.InlineKeyboardButton(text=name,callback_data=name) for name in ['Назад']])
        msg = bot.send_message(c.message.chat.id,tour(int(c.data[-1]),7),reply_markup=keyboard,parse_mode='HTML')
        return
    if c.data[:-1] in info(1,17):
        keyboard = types.InlineKeyboardMarkup(row_width=1)
        keyboard.add(*[types.InlineKeyboardButton(text=name,callback_data=name) for name in ['Назад']])
        msg = bot.send_message(c.message.chat.id,tour(int(c.data[-1]),8),reply_markup=keyboard,parse_mode='HTML')
        return
    if c.data[:-1] in info(1,18):
        keyboard = types.InlineKeyboardMarkup(row_width=1)
        keyboard.add(*[types.InlineKeyboardButton(text=name,callback_data=name) for name in ['Назад']])
        msg = bot.send_message(c.message.chat.id,tour(int(c.data[-1]),9),reply_markup=keyboard,parse_mode='HTML')
        return
    if c.data[:-1] in info(1,19):
        keyboard = types.InlineKeyboardMarkup(row_width=1)
        keyboard.add(*[types.InlineKeyboardButton(text=name,callback_data=name) for name in ['Назад']])
        msg = bot.send_message(c.message.chat.id,tour(int(c.data[-1]),10),reply_markup=keyboard,parse_mode='HTML')
        return
    if c.data[:-1] in info(1,20):
        keyboard = types.InlineKeyboardMarkup(row_width=1)
        keyboard.add(*[types.InlineKeyboardButton(text=name,callback_data=name) for name in ['Назад']])
        msg = bot.send_message(c.message.chat.id,tour(int(c.data[-1]),11),reply_markup=keyboard,parse_mode='HTML')
        return
    if c.data[:-1] in info(1,21):
        keyboard = types.InlineKeyboardMarkup(row_width=1)
        keyboard.add(*[types.InlineKeyboardButton(text=name,callback_data=name) for name in ['Назад']])
        msg = bot.send_message(c.message.chat.id,tour(int(c.data[-1]),12),reply_markup=keyboard,parse_mode='HTML')
        return
    if c.data[:-1] in info(1,22):
        keyboard = types.InlineKeyboardMarkup(row_width=1)
        keyboard.add(*[types.InlineKeyboardButton(text=name,callback_data=name) for name in ['Назад']])
        msg = bot.send_message(c.message.chat.id,tour(int(c.data[-1]),13),reply_markup=keyboard,parse_mode='HTML')
        return
    if c.data[:-1] in info(1,24):
        if bdpol[k].shfr!='':
             msg = bot.send_message(c.message.chat.id, info(1,26)+'\n'+tour(int(c.data[-1]),14)+'/?&SQF_S=$'+bdpol[k].shfr,parse_mode='HTML')
             return			 
        if bdpol[k].email=='':
                        msg = bot.send_message(c.message.chat.id, info(1,25),parse_mode='HTML')
                        bdpol[k].email=10+int(c.data[-1])	
                        return()
        else:
             msg = bot.send_message(c.message.chat.id, info(1,26)+'\n'+tour(int(c.data[-1]),14)+'/?&SQF_S=$'+bdpol[k].shfr,parse_mode='HTML')		
    if c.data[1:]=='klmno':
        zz=int(c.data[0])
        keyboard = types.InlineKeyboardMarkup(row_width=2)
        keyboard.add(*[types.InlineKeyboardButton(text=name,callback_data=name+c.data[0]) for name in [info(1,24),info(1,49)]])
        keyboard.add(types.InlineKeyboardButton(text='Подробней', url=tour(zz,14)))
        msg = bot.send_message(c.message.chat.id,'<b>'+tour(zz,1)+'</b>\n\n'+tour(zz,2)+'\n\n'+tour(zz,3)+'\n\n'+tour(zz,4)+'\n\n'+tour(zz,14) ,reply_markup=keyboard,parse_mode='HTML',disable_web_page_preview=True)	        	
        return	
    if c.data == info(1,31):
        msg = bot.send_message(c.message.chat.id, info(1,25),parse_mode='HTML')
        bdpol[k].email=2	
        return()	        	
    if c.data == info(1,27):
        msg = bot.send_message(c.message.chat.id, info(1,46),parse_mode='HTML')
        return()
    if c.data == info(1,47):
        msg = bot.send_message(c.message.chat.id, info(1,48),parse_mode='HTML')
        return()
    if c.data == info(1,44):
        msg = bot.send_message(c.message.chat.id, info(1,45)+'\n'+'http://www.surfinua.com/surftrips/?&SQF_S=$'+bdpol[k].shfr,parse_mode='HTML')  
  ############ registracia na poezdku		
    if c.data[:-1] == info(1,49):
        if bdpol[k].phone!='' and bdpol[k].name!='' and bdpol[k].email!='' :
            bdpol[k].regpoz=4
            msg = bot.send_message(c.message.chat.id, info(1,53),parse_mode='HTML')   
            return() 

        if bdpol[k].phone!='' and bdpol[k].name!='':
            bdpol[k].regpoz=3
            msg = bot.send_message(m.chat.id, info(1,53),parse_mode='HTML')   
            return()			
        keyboard = types.ReplyKeyboardMarkup(row_width=1, resize_keyboard=True)
        button_phone = types.KeyboardButton(text="Отправить контакт", request_contact=True)
        keyboard.add(button_phone)
        keyboard.add(*[types.InlineKeyboardButton(text=name,callback_data=name) for name in [info(1,52)]])
        bdpol[k].regpoz=1
        bdpol[k].kudaedu=tour(int(c.data[-1]),16)
        msg = bot.send_message(c.message.chat.id, info(1,50),parse_mode='HTML',reply_markup=keyboard)
        return()        	
############### bd OPLATIVSHIH
    if c.data=='Загрузить список оплативших':
        wads = openpyxl.load_workbook(filename = 'bdpaid.xlsx')
        sheetka = wads['test']	
        for i in range(1,len(arraypaidphone)):
            sheetka.cell(row=i, column=1).value = arraypaidphone[i-1]
            sheetka.cell(row=i, column=2).value = arraypaidemail[i-1]	
        wads.save('bdpaid.xlsx')
        msg = bot.send_document(c.message.chat.id,open('bdpaid.xlsx', 'rb'))	
############## bd polzovatelei		
    if c.data=='Загрузить список пользователей':
        wads = openpyxl.load_workbook(filename = 'bduser.xlsx')
        sheetka = wads['test']	
        for i in range(0,len(bdpol)):
            sheetka.cell(row=i+1, column=1).value = bdpol[i].email
            sheetka.cell(row=i+1, column=2).value = bdpol[i].phone
            sheetka.cell(row=i+1, column=3).value = bdpol[i].bonus
        wads.save('bduser.xlsx')
        msg = bot.send_document(c.message.chat.id,open('bduser.xlsx', 'rb'))			
    #if c.data=='Редактировать пользователя':		
        #msg = bot.send_message(c.message.chat.id, 'Напишите телефон пользователя',parse_mode='HTML')		
		
		
		
		
		
		
		
		
		
		
		
		
def name(m):
    global bdpol, admin, spisem, spisko
    if m.chat.id<0:
        return()
    k=nomer(m.chat.id)
    if bdpol[k].spis==1:
        bdpol[k].spis=0
        try:
            spisan=int(m.text)
        except Exception:
            spisan=0
        if spisan==0:
            msg= bot.send_message(m.chat.id, info(1,34),parse_mode='HTML')
            return()
        elif spisan>bdpol[k].bonus:
            msg= bot.send_message(m.chat.id, info(1,35),parse_mode='HTML') 
            return()			
        else:
            msg = bot.send_message(m.chat.id, info(1,36),parse_mode='HTML')	
            msg = bot.send_message(m.chat.id, bdpol[k].email+' '+info(1,38)+' '+str(spisan),parse_mode='HTML')	
            return()	
    if bdpol[k].adminin==1:
        if m.text==admin_password:
            bdpol[k].is_admin=1
            bdpol[k].adminin=0
            admin=m.chat.id
            keyboard = types.ReplyKeyboardMarkup(resize_keyboard=True)
            keyboard.add(*[types.InlineKeyboardButton(text=name,callback_data=name) for name in ['Списать бонусы','Изменить тексты','Изменить туры','БД оплативших','БД пользователей']])   
            msg = bot.send_message(m.chat.id, info(1,37),reply_markup=keyboard,parse_mode='HTML')
            adminid=m.chat.id
        else: 
            bdpol[k].adminin=0 
            msg = bot.send_message(m.chat.id, 'Пароль введен не верно.',parse_mode='HTML')	\
############################### BD oplativshih
    if m.text=='БД пользователей':
        keyboard = types.InlineKeyboardMarkup(row_width=1)	
        keyboard.add(*[types.InlineKeyboardButton(text=name,callback_data=name) for name in ['Загрузить список пользователей']])
        msg = bot.send_message(m.chat.id,'Тут вы можете работать с БД пользователей.',reply_markup=keyboard,parse_mode='HTML')  
############################### BD polzovatelei	
    if m.text=='БД оплативших':
        keyboard = types.InlineKeyboardMarkup(row_width=1)	
        keyboard.add(*[types.InlineKeyboardButton(text=name,callback_data=name) for name in ['Загрузить список оплативших']])
        msg = bot.send_message(m.chat.id,'Тут вы можете работать с БД оплативших услугу.',reply_markup=keyboard,parse_mode='HTML') 	
############################### otmena registracii
    if m.text==info(1,52):
        bdpol[k].regpoz=0
        keyboard = types.ReplyKeyboardMarkup(resize_keyboard=True)
        keyboard.add(*[types.InlineKeyboardButton(text=name,callback_data=name) for name in [info(1,13)]])
        keyboard.add(*[types.InlineKeyboardButton(text=name,callback_data=name) for name in [info(1,2), info(1,3),'Бонусы']])
        msg = bot.send_message(m.chat.id, info(1,52),
        reply_markup=keyboard,parse_mode='HTML')   
############################### Ruchnaia registracia
    if bdpol[k].regpoz==1 and m.text == info(1,51):
            keyboard = types.ReplyKeyboardMarkup(resize_keyboard=True)
            keyboard.add(*[types.InlineKeyboardButton(text=name,callback_data=name) for name in [info(1,52)]])
            msg = bot.send_message(m.chat.id, info(1,55),reply_markup=keyboard,parse_mode='HTML')	
            return()
    if 	bdpol[k].regpoz==1:
        bdpol[k].regpoz=2
        bdpol[k].name=m.text
        msg = bot.send_message(m.chat.id, info(1,56),parse_mode='HTML')	
        return()
    if 	bdpol[k].regpoz==2:
        bdpol[k].regpoz=3
        bdpol[k].phone=m.text
        msg = bot.send_message(m.chat.id, info(1,57),parse_mode='HTML')	
        return()	
################################ Ruchnoi vvod telefona
    if bdpol[k].phone1==1 and m.text != "Отправить контакт":
        bdpol[k].phone1=m.text
        bdpol[k].shfr=refrand()	
        keyboard = types.InlineKeyboardMarkup(row_width=1)	
        keyboard.add(*[types.InlineKeyboardButton(text=name,callback_data=name) for name in [info(1,27),info(1,31),info(1,47),info(1,44)]])
        msg = bot.send_message(m.chat.id, info(1,28)+str(bdpol[k].bonus)+'\n'+info(1,29)+str(bdpol[k].bonus)+'\n'+info(1,30)+str(bdpol[k].bonus)+'\n'+info(1,32)+str(bdpol[k].phone1),reply_markup=keyboard,parse_mode='HTML')	 		
############################## ADMIN COMADNI
    if m.text=='Изменить тексты':
        if bdpol[k].id==admin:
            msg = bot.send_document(m.chat.id,open('2.xlsx', 'rb') ,caption=info(1,39),parse_mode='HTML')
            bdpol[k].addtext=1
    if m.text=='Изменить туры':
        if bdpol[k].id==admin:
            msg = bot.send_document(m.chat.id,open('1.xlsx', 'rb') ,caption=info(1,39),parse_mode='HTML')
            bdpol[k].addtext=2
    if m.text=='Списать бонусы':
        if bdpol[k].id==admin:
            msg = bot.send_message(m.chat.id,info(1,42),parse_mode='HTML')
            bdpol[k].addtext=3
            return()
    if bdpol[k].addtext==3:
            bdpol[k].addtext=4        
            spisem=m.text
            for i in range (0,len(bdpol)):
                if spisem == bdpol[i].email:
                    msg = bot.send_message(m.chat.id,info(1,33),parse_mode='HTML')
                    return()
            msg = bot.send_message(m.chat.id,'Данные введены неверно',parse_mode='HTML')
            bdpol[k].addtext=0  
    if bdpol[k].addtext==4:
        bdpol[k].addtext=0        
        try:
            spisan=int(m.text)
        except Exception:
            spisan=0
        if spisan==0:
            msg= bot.send_message(m.chat.id, info(1,34),parse_mode='HTML')
            return()
        elif spisan>bdpol[k].bonus:
            msg= bot.send_message(m.chat.id, info(1,35),parse_mode='HTML') 
            return()			
        else:
            for i in range (0,len(bdpol[k])):
                if spisem == bdpol[i].email:
                    msg = bot.send_document(m.chat.id,info(1,43),parse_mode='HTML')
                    bdpol[i].bonus=bdpol[i].bonus-spisan
                    return()
            msg = bot.send_document(m.chat.id,'Данные введены неверно',parse_mode='HTML')
            bdpol[k].addtext=0 
############################# personalni link 
    if m.text == info(1,44):
        msg = bot.send_message(m.chat.id, info(1,45)+'\n'+'http://www.surfinua.com/surftrips/?&SQF_S=$'+bdpol[k].shfr,parse_mode='HTML')        	
############################################
   
			
    if m.text == 'Бонусы':
        if bdpol[k].phone=='':
                        keyboard = types.ReplyKeyboardMarkup(row_width=1, resize_keyboard=True)
                        button_phone = types.KeyboardButton(text="Отправить контакт", request_contact=True)
                        keyboard.add(button_phone)
                        msg = bot.send_message(m.chat.id, info(1,25),reply_markup=keyboard,parse_mode='HTML')
                        bdpol[k].phone=1	
                        return()	
        keyboard = types.InlineKeyboardMarkup(row_width=1)	
        keyboard.add(*[types.InlineKeyboardButton(text=name,callback_data=name) for name in [info(1,27),info(1,47),info(1,44)]])			
        msg = bot.send_message(m.chat.id, info(1,28)+str(bdpol[k].bonus)+'\n'+info(1,29)+str(len(bdpol[k].arrayreg))+'\n'+info(1,30)+str(bdpol[k].regprosh)+'\n'+info(1,32)+str(bdpol[k].phone)+'\n'+info(1,59)+str(bdpol[k].email),reply_markup=keyboard)
    if bdpol[k].email!='':
        print('oh')
        try:
            zz=int(bdpol[k].email)
            if zz>10:
                 bdpol[k].email=m.text
                 bdpol[k].shfr=refrand()	
       
                 msg = bot.send_message(m.chat.id, info(1,26)+'\n'+tour(int(zz-10),14)+'/?&SQF_S='+bdpol[k].shfr,parse_mode='HTML')	
        except Exception:
            zz=''				
    if bdpol[k].email==1:
        bdpol[k].email=m.text
        bdpol[k].shfr=refrand()	
        keyboard = types.InlineKeyboardMarkup(row_width=1)	
        keyboard.add(*[types.InlineKeyboardButton(text=name,callback_data=name) for name in [info(1,27),info(1,31),info(1,47),info(1,44)]])
        msg = bot.send_message(m.chat.id, info(1,28)+str(bdpol[k].bonus)+'\n'+info(1,29)+str(bdpol[k].bonus)+'\n'+info(1,30)+str(bdpol[k].bonus)+'\n'+info(1,32)+str(bdpol[k].email),reply_markup=keyboard,parse_mode='HTML')		
    if bdpol[k].email==2:
        bdpol[k].email=m.text
        keyboard = types.InlineKeyboardMarkup(row_width=1)	
        keyboard.add(*[types.InlineKeyboardButton(text=name,callback_data=name) for name in [info(1,27),info(1,31),info(1,47),info(1,44)]])
        msg = bot.send_message(m.chat.id, info(1,28)+str(bdpol[k].bonus)+'\n'+info(1,29)+str(len(bdpol[k].arrayreg))+'\n'+info(1,30)+str(bdpol[k].regprosh)+'\n'+info(1,32)+str(bdpol[k].phone1)+'\n'+info(1,59)+str(bdpol[k].email))	
    if m.text==info(1,2):
       msg =    msg = bot.send_message(m.chat.id, info(1,12))
    if m.text==info(1,3):
       msg =    msg = bot.send_message(m.chat.id, '<b>'+info(1,4)+'</b>\n'+info(1,5)+'\n\n<b>'+info(1,6)+'</b>\n'+info(1,7)+'\n\n<b>'+info(1,8)+'</b>\n'+info(1,9)+'\n\n<b>'+info(1,10)+'</b>\n'+info(1,11),parse_mode='HTML')
    if m.text==info(1,13):
        keyboard = types.InlineKeyboardMarkup(row_width=1)
        s=''
        for i in range(1,100):
            if  tour(i,1)=='None':
                break
            s=s+str(i)+') <b>'+tour(i,1)+'</b>\n'+tour(i,4)+'\n'+tour(i,5)+'\n'+tour(i,14)+'\n\n'	
            keyboard.add(*[types.InlineKeyboardButton(text=name,callback_data=str(i)+'klmno') for name in [tour(i,6)]])			
        msg = bot.send_message(m.chat.id, info(1,14)+'\n'+s,reply_markup=keyboard,parse_mode='HTML',disable_web_page_preview=True)
############################### Rigistracia na poezdku 
    if bdpol[k].regpoz==3:
        bdpol[k].email=m.text
        bdpol[k].regpoz=4
        msg = bot.send_message(m.chat.id, info(1,53),parse_mode='HTML')   
        return()		
    if bdpol[k].regpoz==4:
        bdpol[k].koment=m.text
        bdpol[k].regpoz=0
        keyboard = types.ReplyKeyboardMarkup(resize_keyboard=True)
        keyboard.add(*[types.InlineKeyboardButton(text=name,callback_data=name) for name in [info(1,13)]])
        keyboard.add(*[types.InlineKeyboardButton(text=name,callback_data=name) for name in [info(1,2), info(1,3),'Бонусы']])
        msg = bot.send_message(m.chat.id, info(1,54),reply_markup=keyboard,parse_mode='HTML')  
        print(bdpol[k].name,bdpol[k].email,bdpol[k].phone,bdpol[k].koment,bdpol[k].kudaedu)
        pochta.sendmail(bdpol[k].name,bdpol[k].email,bdpol[k].phone,bdpol[k].koment,bdpol[k].kudaedu)       		

		
@bot.message_handler(content_types=['document'])
def photoget(message):
    global wb, wb1, sheet, sheet1
    k=nomer(message.chat.id)
    if bdpol[k].addtext==1:	
        fileid=(message.document.file_id)
        bb=bot.get_file(fileid)
        bb=bb.file_path
        logo = urllib.request.urlopen("https://api.telegram.org/file/bot"+TOKEN+"/"+bb).read()
        f = open('2.xlsx', "wb")
        f.write(logo)
        f.close()
        wb1 = openpyxl.load_workbook(filename = '2.xlsx')
        sheet1 = wb1['test']
        bdpol[k].addtext=0
        msg = bot.send_message(message.chat.id, info(1,41)) 
    if bdpol[k].addtext==2:	
        fileid=(message.document.file_id)
        bb=bot.get_file(fileid)
        bb=bb.file_path
        logo = urllib.request.urlopen("https://api.telegram.org/file/bot"+TOKEN+"/"+bb).read()
        f = open('1.xlsx', "wb")
        f.write(logo)
        f.close()
        wb = openpyxl.load_workbook(filename = '1.xlsx')
        sheet = wb1['test']
        bdpol[k].addtext=0
        msg = bot.send_message(message.chat.id, info(1,41))        	
		
		
		
		
		
		
		
		
def refrand():
    gh=1
    while gh==1:
            chars=string.ascii_uppercase + string.digits
            zz=''.join(random.choice(chars) for _ in range(8))   
            if zz not in reflist:
                gh=0			
    print(zz) 
    reflist.append(zz)	
    return(zz)	
	
####################################################### otsilaem kontakt	
@bot.message_handler(content_types=["contact"])
def check_chatid(message):
    print(message.contact.phone_number)
    global bdpol
    global title
    k=nomer(message.chat.id)
    if bdpol[k].regpoz==1:
           bdpol[k].regpoz=3
           bdpol[k].phone=str(message.contact.phone_number)
           bdpol[k].name=message.from_user.first_name+' '+message.from_user.last_name 
           keyboard = types.ReplyKeyboardMarkup(resize_keyboard=True)
           keyboard.add(*[types.InlineKeyboardButton(text=name,callback_data=name) for name in [info(1,52)]])
           msg = bot.send_message(message.chat.id, info(1,57) ,reply_markup=keyboard)		   
    if bdpol[k].phone==1:		   
        bdpol[k].phone=str(message.contact.phone_number)
        bdpol[k].shfr=refrand()	
        bdpol[k].name=message.from_user.first_name+' '+message.from_user.last_name 
        keyboard = types.ReplyKeyboardMarkup(resize_keyboard=True)
        keyboard.add(*[types.InlineKeyboardButton(text=name,callback_data=name) for name in [info(1,13)]])
        keyboard.add(*[types.InlineKeyboardButton(text=name,callback_data=name) for name in [info(1,2), info(1,3),'Бонусы']])
        msg = bot.send_message(message.chat.id, 'Регистрация прошла успешно!',
        reply_markup=keyboard,parse_mode='HTML')
        keyboard = types.InlineKeyboardMarkup(row_width=1)	
        keyboard.add(*[types.InlineKeyboardButton(text=name,callback_data=name) for name in [info(1,27),info(1,47),info(1,44)]])
        msg = bot.send_message(message.chat.id, info(1,28)+str(bdpol[k].bonus)+'\n'+info(1,29)+str(bdpol[k].bonus)+'\n'+info(1,30)+str(bdpol[k].bonus)+'\n'+info(1,32)+str(bdpol[k].phone),reply_markup=keyboard,parse_mode='HTML')	
    output = open('bdpol.pkl', 'wb')
    pickle.dump(bdpol, output, 2)
    output.close()

	
	
######################################## Skachivanie i chtenie tablici

def chek_table(name):
    global arraypaidphone, arraypaidemail
    dls = "https://docs.google.com/spreadsheets/d/"+name+"/export?format=xlsx&id="+name
    resp = requests.get(dls)

    output = open('test.xlsx', 'wb')
    output.write(resp.content)
    output.close()





    zz = openpyxl.load_workbook(filename = 'test.xlsx', read_only=True, data_only=True)
    shut = zz['Table']

    kk=True
    i=2
    print(shut.cell(row=3, column=6).value)
    while kk:
        i+=1
        ff=shut.cell(row=i, column=6).value
        if ff==None:
            kk=False
        else:
            if str(ff)=='1':
                if str(shut.cell(row=i, column=24).value) not in arraypaidphone or str(shut.cell(row=i, column=25).value) not in arraypaidemail:
                    arraypaidphone.append(str(shut.cell(row=i, column=24).value)) 
                    arraypaidemail.append(str(shut.cell(row=i, column=25).value)) 			

	
	
	
	
	
	
	
	
def check_reg(name,lor):
    global bdpol
    dls = "https://docs.google.com/spreadsheets/d/"+name+"/export?format=xlsx&id="+name
    resp = requests.get(dls)

    output = open('test.xlsx', 'wb')
    output.write(resp.content)
    output.close()





    zz = openpyxl.load_workbook(filename = 'test.xlsx', read_only=True, data_only=True)
    shut = zz['Table']

    kk=True
    i=2
    print(shut.cell(row=3, column=6).value)
    while kk:
        i+=1
        ff=shut.cell(row=i, column=6).value
        bb=str(shut.cell(row=i, column=46).value)
        if bb[0]=='$':
            for j in range(0,len(bdpol)):
                if bdpol[j].shfr==bb[1:] and str(shut.cell(row=i, column=24).value) not in bdpol[j].arrayreg:
                    bdpol[j].arrayreg.append(str(shut.cell(row=i, column=24).value))
                if 	str(shut.cell(row=i, column=24).value) not in arraypaidphone and int(shut.cell(row=i, column=6).value)==1 and bdpol[j].shfr==bb[1:]:
                        bdpol[j].bonus+=int(tour(lor,18))
                        bdpol[j].regprosh+=1
                        msg = bot.send_message(bdpol[j].id, info(1,58)) 						
        if ff==None:
            kk=False   	
	
	
	
	
	
	
########################  cheking function
def potok():
    while True:
        for i in range(1,100):
            if tour(i,17)=='None':
                break
            else:
                print(tour(i,17))
                check_reg(str(tour(i,17)),i)
                chek_table(str(tour(i,17))) 
        time.sleep(30)				

################################## potok s chekom tablic

t1 = threading.Thread(target=potok, args=())
t1.start()

##################################	
	
		
bot.remove_webhook()

bot.set_webhook(url=WEBHOOK_URL_BASE + WEBHOOK_URL_PATH,
                certificate=open(WEBHOOK_SSL_CERT, 'r'))

cherrypy.config.update({
    'server.socket_host': WEBHOOK_LISTEN,
    'server.socket_port': WEBHOOK_PORT,
    'server.ssl_module': 'builtin',
    'server.ssl_certificate': WEBHOOK_SSL_CERT,
    'server.ssl_private_key': WEBHOOK_SSL_PRIV
})

cherrypy.quickstart(WebhookServer(), WEBHOOK_URL_PATH, {'/': {}})
