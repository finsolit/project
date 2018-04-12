import requests
import json
import xlwt
import urllib
import openpyxl
dls = "https://docs.google.com/spreadsheets/d/1VU3eJFVrqo8njDZsf2BITM3HoFMFYGyQXm0Dmq7_usg/export?format=xlsx&id=1VU3eJFVrqo8njDZsf2BITM3HoFMFYGyQXm0Dmq7_usg"
resp = requests.get(dls)

output = open('test.xlsx', 'wb')
output.write(resp.content)
output.close()
print('all ok')





zz = openpyxl.load_workbook(filename = 'test.xlsx', read_only=True, data_only=True)
shut = zz['Table']

kk=True
i=2
arraypaidphone=[]
arraypaidemail=[]
arraydec=[]
print(shut.cell(row=3, column=6).value)
while kk:
    i+=1
    ff=shut.cell(row=i, column=46).value
    arraydec.append(ff) 
    ro=shut.cell(row=i, column=6).value
    if ro==None:
        kk=False	
print(arraydec)
a=input()